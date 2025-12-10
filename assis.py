# -*- coding: utf-8 -*-
# 指定文件编码为UTF-8，确保支持中文

#
# import os  # 导入操作系统模块，用于处理文件和目录路径
# import zipfile  # 导入zipfile模块，用于创建和操作ZIP文件
# import tkinter as tk  # 导入Tkinter模块，用于创建GUI界面
# from tkinter import ttk, messagebox, filedialog  # 导入Tkinter的子模块，用于创建表格、消息框和文件对话框
# import pandas as pd  # 导入Pandas模块，用于处理Excel文件
# from fontTools.misc.timeTools import timestampNow
# from netmiko import ConnectHandler, NetMikoTimeoutException, NetMikoAuthenticationException  # 导入netmiko模块，用于网络设备连接
# import threading  # 导入threading模块，用于多线程操作
# from datetime import datetime  # 导入datetime模块，用于处理日期和时间
# from ping3 import ping  # 导入ping3模块，用于执行Ping测试
# from concurrent.futures import ThreadPoolExecutor  # 导入ThreadPoolExecutor模块，用于线程池管理
# import requests  # 导入requests模块，用于发送HTTP请求（如Server酱通知）
# from apscheduler.schedulers.background import BackgroundScheduler  # 导入APScheduler模块，用于定时任务管理
# from apscheduler.jobstores.memory import MemoryJobStore  # 导入MemoryJobStore模块，用于存储定时任务
# from apscheduler.executors.pool import ThreadPoolExecutor as APSPThreadPoolExecutor  # 导入线程池执行器
# from apscheduler.triggers.cron import CronTrigger  # 导入CronTrigger模块，用于设置定时任务的触发条件
# import json  # 导入json模块，用于处理JSON格式的数据
# import time  # 导入time模块，用于时间相关的操作
# from tkinter.scrolledtext import ScrolledText  # 导入ScrolledText模块，用于创建带滚动条的文本框
# import openpyxl  # 导入openpyxl模块，用于处理Excel文件
# from openpyxl.styles import PatternFill  # 导入PatternFill模块，用于设置Excel单元格的背景色
# from tkinter import ttk  # 再次导入ttk模块，确保使用最新版本
# from tkinter import simpledialog  # 导入 simpledialog 模块
# # 在导入模块部分添加
# from tkinter import PanedWindow
# from cryptography.fernet import Fernet  # 加密與解密
# import re
# import difflib
# import windnd  # 实现拖入功能，调用此库
# from PIL import Image, ImageTk
# import math
# from plyer import notification
# import time
# import socket
import os
import threading
import time
import json
import re
import socket
import math
import requests
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from tkinter.scrolledtext import ScrolledText
from tkinter import PanedWindow
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import pandas as pd
from cryptography.fernet import Fernet
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.jobstores.memory import MemoryJobStore
from apscheduler.executors.pool import ThreadPoolExecutor as APSPThreadPoolExecutor
from apscheduler.triggers.cron import CronTrigger
from ping3 import ping
import matplotlib

matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from netmiko import ConnectHandler, NetMikoTimeoutException, NetMikoAuthenticationException
from concurrent.futures import ThreadPoolExecutor
from difflib import SequenceMatcher
from plyer import notification
import windnd
from PIL import Image, ImageTk

# import matplotlib matplotlib.use('Agg') from import pyplot as plt
# import matplotlib
# matplotlib.use('TkAgg')
matplotlib.use('Agg')
import matplotlib.pyplot as plt

plt.rcParams['font.sans-serif'] = ['SimHei']  # 设置中文字体
plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

from datetime import datetime
from matplotlib.backends.backend_agg import FigureCanvasAgg
import io
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import numpy as np
from matplotlib.backends.backend_pdf import PdfPages
import atexit
import zipfile
import difflib
from concurrent.futures import ThreadPoolExecutor
import difflib
from concurrent.futures import ThreadPoolExecutor
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
import os
import json
import time
import threading
from datetime import datetime
# import matplotlib
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image
# from io import BytesIO
# import matplotlib
# matplotlib.use('Agg')
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from io import BytesIO
import re
# import datetime from datetime
# from tkinter import scrolledtext as ScrolledText
import webbrowser
from tkinter import scrolledtext
import tempfile
from tkhtmlview import HTMLLabel
import socket
import platform
import psutil          # 硬件信息

#
import inspect
class DeviceManagerApp:
    def __init__(self, root):
        self.root = root
        self.ver = "v7.8.5"
        self.root.title(f"网络大剑仙{self.ver} By:WB")
        # 设置主窗口大小为固定值，例如 1024x768
        self.root.geometry("950x820")
        # 禁止用户通过拖拽改变窗口大小
        self.root.resizable(False, False)
        self.device_file = ""
        self.output_directory = os.getcwd()
        self.command_file_path = None
        self.server_chan_key = ""
        self.ping_timeout = 3
        self.ssh_telnet_timeout = 20
        self.cmd_read_interval = 3
        self.enable_notification = tk.BooleanVar(value=False)
        self.log_file = None
        self.cancel_event = threading.Event()
        self.bark_device_key = tk.StringVar()
        self.notification_type = tk.StringVar(value="Server酱")

        # 默认设置为 ICMP 测试
        self.ping_type_var = tk.StringVar(value="ICMP")

        self.ping_stop_event = threading.Event()

        # 初始化分组功能
        self.current_group_var = tk.StringVar(value="默认局点")  # 初始化 current_group_var

        # self.html_content = ""

        # 用于存储每个设备最后执行的命令和输出
        self.last_error_command = {}  # 用于存储每个设备最后执行的命令
        self.last_error_output = {}  # 用于存储每个设备最后执行的输出

        # 动态保存路径，初始为 None
        self.selected_save_directory = None
        # 定义config和device-log目录路径
        self.config_dir = os.path.join(os.getcwd(), "config")
        self.default_device_log_dir = os.path.join(os.getcwd(), "device-log")  # 默认日志目录
        self.log_file_path = os.path.join(self.config_dir, "操作日志.log")

        # 确保目录存在
        # os.makedirs(self.config_dir, exist_ok=True)
        # os.makedirs(self.device_log_dir, exist_ok=True)

        # 确保目录存在
        os.makedirs(self.config_dir, exist_ok=True)
        os.makedirs(self.default_device_log_dir, exist_ok=True)

        # self.create_status_light() # 小圆点呼吸灯

        # 初始化 template_listbox
        self.template_listbox = tk.Listbox(self.root, width=30, height=10)
        # self.template_listbox.pack(pady=10)


        # 初始化命令模板
        self.command_templates = {}
        if os.path.exists("command_templates.json"):
            with open("command_templates.json", "r") as f:
                self.command_templates = json.load(f)
                self.log(f"[提示]：命令模板加载成功")

        # 初始化命令历史和审计日志
        # self.initialize_command_history()
        # self.initialize_audit_log()

        # 初始化日志文件
        self.init_log_file()

        self.check_agreement()
        self.create_widgets()  # 创建GUI界面
        self.log("[提示]：创建GUI界面")
        self.scheduler = BackgroundScheduler(jobstores={'default': MemoryJobStore()},
                                             executors={'default': APSPThreadPoolExecutor(max_workers=20)})
        self.scheduler.start()
        self.jobs = {}
        self.task_statuses = {}

        self.initialize_audit_log()
        self.log("[提示]：初始化审计日志")

        # self.load_scheduled_tasks()
        self.load_saved_tasks()
        self.log("[提示]：加载定时任务")


        self.load_saved_keys()
        self.original_passwords = {}  # 用于存储原始密码
        self.super_passwords = {}  # 用于存储原始super密码
        # self.log("[提示]：开始加载设备信息")
        self.load_device_info_from_json()  # 在初始化时加载设备信息

        self.last_error_command = {}  # 用于存储每个设备最后执行的命令

        self.real_time_log_window = None  # 实时日志窗口
        self.create_real_time_log_menu()  # 实时日志窗口
        # self.create_status_light()  # 小圆点呼吸灯

        self.log("[提示]：初始化命令模板")
        self.initialize_command_history()
        self.log("[提示]：初始化命令历史记录")

        # self.create_widgets()  # 创建GUI界面

        self.create_status_light()
        self.log("[提示]：创建程序呼吸灯")
        # self.create_real_time_log_menu()
        # self.log("[提示]：创建实时日志菜单")

        # 检查并创建默认的命令模板文件
        self.check_and_create_default_templates()
        self.create_command_template_management()
        # self.log("[提示]：创建命令模板管理菜单")

        # 设备信息存储加解密
        self.key = Fernet.generate_key()
        self.cipher_suite = Fernet(self.key)
        # self.log("[成功]：网络大剑仙v5.6.1 By:WB")


        # 创建命令模板管理菜单
        # self.create_command_template_management()
        self.output_text.tag_config("success", foreground="green")  # 设置成功日志的文本颜色

        # 初始化读秒功能
        self.create_countdown_label()

        # 注册清理函数
        atexit.register(self.cleanup)

        # 新增对比功能
        self.compare_menu = None
        self.compare_window = None
        self.compare_tree = None
        self.compare_text1 = None
        self.compare_text2 = None
        self.compare_results_text = None
        self.compare_progress_bar = None
        self.compare_status_label = None
        self.compare_thread_pool = ThreadPoolExecutor(max_workers=8)

        #进度条闪烁功能
        # self.progress_bar = ttk.Progressbar(root, mode="determinate")
        # self.progress_bar.pack(fill=tk.X, expand=True)
        # 初始化静止检测相关变量
        self.progress_last_update = time.time()
        self.static_progress_timer = None
        self.breathing_effect_active = False
        self.breathing_effect_thread = None
        self.flash_count = 0  # 闪烁计数
        # 启动静止检测计时器
        self.reset_static_timer()

        # # 创建进度条
        # self.progress_bar = ttk.Progressbar(root, mode="determinate")
        # self.progress_bar.grid(row=4, column=0, columnspan=8, padx=10, pady=10, sticky="ew")
        #
        # # 创建状态标签
        # self.status_label = tk.Label(root, text="执行状态：等待操作")
        # self.status_label.grid(row=4, column=0, padx=10, pady=10, sticky="w")
        #
        # # 配置网格权重
        # self.root.grid_rowconfigure(4, weight=0)
        # self.root.grid_columnconfigure(0, weight=1)
        #初始化脚本执行错误的不再提供参数
        self.mute = tk.BooleanVar(value=False)
        self.continue_execution = tk.BooleanVar(value=False)

        #错误检测
        self.retry_count = 60  # 默认命令执行超时时间
        self.error_chars = ["Unrecognized","Ambiguous","Too many parameters"]  # 默认错误字符规则
        # self.input_chars = ["Unrecognized","Ambiguous"]  # 默认错误字符规则
        self.prompt_patterns = [r'>', r'#', r']', r'}', r'\$']
        self.execution_mode = "send_command_timing"  # 默认模式为 send_command_timing

        #Super密码字符
        self.load_super_password_chars()

        self.load_command_templates()
        self.command_templates.keys()

        #删除操作日志标记
        self.del_program_log = False
        self.del_device_log = False

        #脚本执行暂停功能
        self.is_executing = False
        self.is_paused = False
        # self.cancel_event = threading.Event()

        # 密码显示控制相关
        self.password_visible = True  # 初始密码可见状态
        self.prompt_text = ""
        #初始化分组局点
        self.update_group_menu()
        self.log(
            f'''[成功]：程序初始化完成
                 [版本]：网络大剑仙{self.ver}
                 [作者]：By:WB'''
        )
        self.log("[提示]：等待用户操作中…")
        self.log_audit("[完成]：初始化完成", "程序初始化完成，等待中")

        #启动通知
        # self.send_server_sauce_notification()
        # 启动 30 分钟周期系统信息推送
        self.send_system_info()
        # self.scheduler.add_job(
        #     self.send_system_info,
        #     trigger=CronTrigger(minute='*/30'),
        #     id='sys_info_push',
        #     name='系统信息循环推送'
        # )
        # # 立即推送一次
        # self.root.after(5000, self.send_system_info)  # 5 秒后先跑一轮
        self.html_content = """
            <!DOCTYPE html>
            <html lang="zh-CN">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>网络大剑仙软件用户使用指导手册</title>
                <style>
                    body {
                        font-family: 'Microsoft YaHei', '微软雅黑', Arial, sans-serif;
                        line-height: 1.6;
                        margin: 0;
                        padding: 20px;
                        color: #333;
                        background-color: #f9f9f9;
                    }
                    .container {
                        max-width: 1200px;
                        margin: 0 auto;
                        background-color: #fff;
                        padding: 20px;
                        border-radius: 5px;
                        box-shadow: 0 0 10px rgba(0, 0, 0, 0.1);
                    }
                    h1, h2, h3, h4 {
                        color: #2c3e50;
                        margin-bottom: 15px;
                    }
                    h1 {
                        font-size: 28px;
                        border-bottom: 2px solid #3498db;
                        padding-bottom: 10px;
                    }
                    h2 {
                        font-size: 24px;
                        margin-top: 30px;
                    }
                    h3 {
                        font-size: 20px;
                    }
                    p {
                        margin-bottom: 15px;
                    }
                    ul, ol {
                        margin-bottom: 15px;
                    }
                    li {
                        margin-bottom: 8px;
                    }
                    li ul, li ol {
                        margin-bottom: 0;
                    }
                    code {
                        background-color: #f1f1f1;
                        padding: 2px 5px;
                        border-radius: 3px;
                        font-family: Consolas, Monaco, 'Courier New', monospace;
                    }
                    .chapter {
                        margin-bottom: 40px;
                    }
                    .footer {
                        margin-top: 40px;
                        padding-top: 20px;
                        border-top: 1px solid #eee;
                        text-align: center;
                        font-size: 14px;
                        color: #777;
                    }
                    .button {
                        display: inline-block;
                        padding: 8px 16px;
                        background-color: #3498db;
                        color: white;
                        text-decoration: none;
                        border-radius: 4px;
                        margin-top: 10px;
                    }
                    .button:hover {
                        background-color: #2980b9;
                    }
                    .code-block {
                        background-color: #f8f8f8;
                        border: 1px solid #ddd;
                        border-radius: 4px;
                        padding: 10px;
                        font-family: Consolas, Monaco, 'Courier New', monospace;
                        overflow-x: auto;
                        margin-bottom: 15px;
                    }
                    .field-name {
                        font-weight: bold;
                    }
                </style>
            </head>
            <body>
                <div class="container">
                    <h1>网络大剑仙用户使用指导手册</h1>

                    <div class="chapter">
                        <h2>一、前言</h2>
                        <p>欢迎使用网络大剑仙！这是一款功能强大的网络设备管理与操作工具，专为网络工程师和管理员设计。本手册将详细指导您如何使用本软件的各项功能，帮助您高效地管理网络设备、执行命令脚本、分析测试结果，并利用定时任务和通知功能提高工作效率。</p>
                    </div>

                    <div class="chapter">
                        <h2>二、软件概述</h2>
                        <h3>（一）功能特点</h3>
                        <ol>
                            <li>多设备管理 ：支持批量导入和管理多种类型的网络设备，包括华三、华为、Cisco、Linux 等。</li>
                            <li>右键菜单 ：右键菜单集成了强大的功能，核心功能在右键中均可找到。</li>
                            <li>快捷键 ：支持复制、粘贴、删除、双击修改等快捷操作。</li>
                            <li>命令执行 ：可对选中的设备执行批量命令，支持自定义命令脚本，提高操作效率。</li>
                            <li>规则设备 ：设置菜单可设置多种规则，灵活的便于对设备帐号密码、执行结果正确性校验。</li>
                            <li>局点管理 ：可创建多个局点，不同局点导入不同的设备信息，并且可实时切换局点。</li>
                            <li>审计功能 ：提供命令执行记录和用户操作记录的直观查看。</li>
                            <li>测试功能 ：提供 Login 测试、Ping 测试（支持 ICMP、TCP、UDP 模式），帮助您快速检查设备连通性和响应状态。</li>
                            <li>定时任务 ：支持创建定时任务，可定期执行设备测试或命令脚本，适用于需要周期性维护的网络环境。</li>
                            <li>结果导出 ：执行结果可导出为多种格式（Excel、TXT、HTML、ZIP），方便存档和分享。</li>
                            <li>命令模板 ：可创建命令模板，并支持导出、导入、应用到设备。</li>
                            <li>日志记录 ：详细记录操作日志和审计日志，便于问题追踪和操作回顾。</li>
                            <li>通知功能 ：集成 Server 酱、Bark 等通知服务，在任务完成或出现错误时及时通知您。</li>
                            <li>一键操作 ：提供一键备份、一键还原、一键清理功能，简化日常维护工作。</li>
                        </ol>

                        <h3>（二）系统要求</h3>
                        <ol>
                            <li>操作系统：Windows 7 及以上版本，推荐 Windows 10 或更高版本。</li>                    
                            <li>硬件要求：至少 2GB 内存，推荐 4GB。</li>
                        </ol>
                    </div>

                    <div class="chapter">
                        <h2>三、安装与启动</h2>
                        <h3>（一）安装步骤</h3>
                        <ol>
                            <li>下载软件安装包：从官方网站或授权渠道下载网络大剑仙的安装文件。</li>
                            <li>运行安装程序：双击下载的安装文件，启动程序。</li>
                        </ol>

                        <h3>（二）启动方式</h3>
                        <ol>
                            <li>桌面快捷方式：双击桌面上的 “网络大剑仙” 快捷方式图标启动软件。</li>
                            <li>开始菜单：点击 Windows 开始菜单，在 “所有程序” 中找到 “网络大剑仙” 并点击启动。</li>
                            <li>快捷键：安装完成后可设置自定义快捷键启动软件（默认无快捷键，可在软件设置中配置）。</li>
                        </ol>
                    </div>

                    <div class="chapter">
                        <h2>四、用户界面介绍</h2>
                        <h3>（一）主界面布局</h3>
                        <p>启动软件后，您将看到以下主要界面元素：</p>
                        <ol>
                            <li>菜单栏 ：位于窗口顶部，提供文件操作、设备管理、测试功能、任务管理、命令模板、审计、设置等菜单选项。</li>
                            <li>工具栏 ：位于菜单栏下方，包含常用功能的快捷按钮，如选择设备文件、下载模板、选择保存目录、选择命令脚本、Login 测试、执行脚本、Ping 测试、停止操作、显示实时日志等。</li>
                            <li>设备信息表格 ：占据主界面大部分区域，以表格形式展示已加载的设备信息，包括主机名、用户名、密码、协议、设备类型、路径、端口、测试结果、执行结果、超级密码等列。支持多列排序和右键菜单操作。</li>
                            <li>命令脚本区域 ：位于主界面右侧，用于显示已加载的命令脚本内容，并提供执行脚本的按钮。</li>
                            <li>操作日志区域 ：位于主界面底部，实时显示软件操作日志，记录用户操作和系统事件。支持导出日志功能。</li>
                            <li>状态栏 ：位于窗口最底部，显示当前操作状态、进度信息和时间显示。</li>
                        </ol>

                        <h3>（二）界面自定义</h3>
                        <ol>
                            <li>调整窗口大小 ：您可以拖动窗口边框调整主界面大小，软件会记住您最后一次调整的尺寸。</li>
                            <li>列宽调整 ：在设备信息表格中，将鼠标指针移到列标题之间的边界上，当指针变为双箭头时，左右拖动可调整列宽。</li>
                            <li>主题切换 ：软件支持亮色和暗色主题（可通过设置菜单切换），满足不同用户的视觉需求。</li>
                            <li>字体设置 ：可在设置菜单中调整界面字体大小和样式，改善阅读体验。</li>
                        </ol>
                    </div>

                    <div class="chapter">
                        <h2>五、快速入门指南</h2>
                        <h3>（一）加载设备信息</h3>
                        <ol>
                            <li>选择设备文件 ：点击菜单栏的 “文件” -> “选择设备文件” 或工具栏上的 “选择设备文件” 按钮。</li>
                            <li>下载模板 ：如果您还没有设备信息文件，可点击 “文件” -> “下载模板” 或工具栏上的 “下载模板” 按钮，下载设备信息模板文件。</li>
                            <li>填写设备信息 ：使用 Excel 打开模板文件，填写设备信息，包括主机名、用户名、密码、协议类型、设备类型、命令脚本路径、端口号等。</li>
                            <li>导入设备信息 ：保存填写好的 Excel 文件后，通过 “选择设备文件” 功能加载到软件中。</li>
                        </ol>

                        <h3>（二）执行基本测试</h3>
                        <ol>
                            <li>Ping 测试 ：选中设备信息表格中的设备，点击菜单栏的 “设备测试” -> “Ping 测试” 或工具栏上的 “Ping 测试” 按钮。软件将对选中的设备进行 ICMP Ping 测试，并在设备信息表格的 “test_result” 列显示结果。</li>
                            <li>更改 Ping 类型 ：您可以通过工具栏上的 “Ping 类型” 下拉菜单选择 TCP 或 UDP 模式，并在 “端口” 输入框中指定目标端口号。</li>
                            <li>Login 测试 ：选中设备后，点击 “设备测试” -> “Login 测试” 或工具栏上的 “Login 测试” 按钮。软件将尝试使用填写的账号信息登录设备，并在 “test_result” 列显示登录测试结果。</li>
                        </ol>

                        <h3>（三）执行命令脚本</h3>
                        <ol>
                            <li>选择命令脚本 ：点击 “文件” -> “选择命令脚本” 或工具栏上的 “选择命令脚本” 按钮，选择包含命令的文本文件。</li>
                            <li>执行脚本 ：选中设备后，点击 “设备测试” -> “执行脚本” 或工具栏上的 “执行脚本” 按钮。软件将依次在选中的设备上执行脚本中的命令，并将结果保存到指定目录。</li>
                            <li>查看执行结果 ：执行完成后，您可以在 “文件” -> “导出脚本执行结果 ZIP” 导出包含执行结果的压缩文件，或在设备信息表格的 “execute_result” 列查看执行状态。</li>
                        </ol>

                        <h3>（四）定时任务设置</h3>
                        <ol>
                            <li>添加定时任务 ：点击 “任务管理” -> “添加定时任务”。</li>
                            <li>设置任务参数 ：输入任务执行时间（HH:MM 格式），选择操作类型（如 Login 测试、执行脚本、Ping 测试），点击 “添加任务” 按钮。</li>
                            <li>管理定时任务 ：在任务列表中选中任务，可通过 “任务管理” 菜单或右键菜单进行删除、暂停、恢复等操作。</li>
                        </ol>

                        <h3>（四）导出执行结果</h3>
                        <ol>
                            <li>脚本执行完成后。</li>
                            <li>在下程序最下方功能区域，可导出对应的执行执行，方便快捷</li>
                        </ol>
                    </div>

                    <div class="chapter">
                        <h2>六、详细功能说明</h2>
                        <h3>（一）设备信息管理</h3>
                        <ol>
                            <li>加载设备信息文件
                                <ul>
                                    <li>支持的文件格式 ：软件支持加载 Excel 文件（.xlsx、.xls 格式），文件应包含设备信息的必要列，如主机名、用户名、密码、协议、设备类型等。</li>
                                    <li>文件拖拽加载 ：您还可以直接将设备信息文件拖拽到设备信息表格区域进行加载。</li>
                                </ul>
                            </li>
                            <li>编辑设备信息
                                <ul>
                                    <li>直接编辑 ：在设备信息表格中双击要修改的单元格，直接输入新值后按回车键确认。</li>
                                    <li>右键菜单编辑 ：选中设备后，右键点击并选择 “编辑”，可批量修改设备信息。</li>
                                    <li>字段说明 ：
                                        <ul>
                                            <li><span class="field-name">主机名（hostname）</span> ：设备的 IP 地址或主机名，用于标识设备并与之建立连接。</li>
                                            <li><span class="field-name">用户名（username）</span> ：登录设备的用户名。</li>
                                            <li><span class="field-name">密码（password）</span> ：登录设备的密码。在表格中以掩码形式（***）显示，实际密码存储在加密文件中。</li>
                                            <li><span class="field-name">协议（protocol）</span> ：与设备通信的协议，可选值包括 telnet、ssh 等。</li>
                                            <li><span class="field-name">设备类型（device_type）</span> ：设备的品牌或类型，如 h3c、huawei、cisco、linux 等，影响命令执行方式和提示符匹配。</li>
                                            <li><span class="field-name">路径（path）</span> ：命令脚本文件的路径，可选填。若填写，执行脚本时将优先使用此路径下的脚本。</li>
                                            <li><span class="field-name">端口（port）</span> ：设备服务的端口号，默认 telnet 为 23，ssh 为 22。留空时软件会自动填充默认端口。</li>
                                            <li><span class="field-name">测试结果（test_result）</span> ：显示设备测试（如 Login 测试、Ping 测试）的结果。</li>
                                            <li><span class="field-name">执行结果（execute_result）</span> ：显示命令脚本执行的结果状态。</li>
                                            <li><span class="field-name">超级密码（super_password）</span> ：某些设备的超级用户密码或特权模式密码。</li>
                                        </ul>
                                    </li>
                                </ul>
                            </li>
                            <li>导出设备信息
                                <ul>
                                    <li>导出为 Excel 文件 ：点击 “文件” -> “导出选中设备”，选择保存路径和文件名，将选中的设备信息导出为 Excel 文件，方便备份和分享。</li>
                                    <li>导出为其他格式 ：软件支持导出为多种格式，您可在相关菜单选项中找到相应功能。</li>
                                </ul>
                            </li>
                        </ol>

                        <h3>（二）命令脚本管理</h3>
                        <ol>
                            <li>加载命令脚本
                                <ul>
                                    <li>选择脚本文件 ：点击 “文件” -> “选择命令脚本” 或工具栏上的 “选择命令脚本” 按钮，选择包含命令的文本文件。脚本文件应为纯文本格式，每行一个命令。</li>
                                    <li>脚本内容显示 ：加载后，命令脚本内容将在命令脚本区域显示，您可在此预览脚本内容。</li>
                                </ul>
                            </li>
                            <li>执行命令脚本
                                <ul>
                                    <li>选中设备 ：在设备信息表格中选中要执行命令的设备。可通过 Ctrl 键多选或 Shift 键选中连续设备。</li>
                                    <li>执行脚本 ：点击 “设备测试” -> “执行脚本” 或工具栏上的 “执行脚本” 按钮。软件将依次在选中的设备上执行脚本中的命令，并将输出结果保存到指定目录。</li>
                                    <li>进度与状态 ：执行过程中，状态栏将显示执行进度和当前操作状态。您可在设备信息表格的 “execute_result” 列查看每个设备的执行结果状态。</li>
                                </ul>
                            </li>
                            <li>脚本执行结果导出
                                <ul>
                                    <li>导出为 ZIP 文件 ：执行完成后，点击 “文件” -> “导出脚本执行结果 ZIP”，选择保存路径和文件名。导出的 ZIP 文件包含每个设备的执行结果，文件格式包括 Excel、TXT、HTML 等。</li>
                                    <li>导出为 HTML 报告 ：点击 “文件” -> “导出脚本执行报告”，生成包含所有设备执行结果的 HTML 报告，方便在浏览器中查看和分享。</li>
                                </ul>
                            </li>
                        </ol>

                        <h3>（三）测试功能详解</h3>
                        <ol>
                            <li>Ping 测试
                                <ul>
                                    <li>ICMP Ping 测试 ：选中设备后，点击 “设备测试” -> “Ping 测试” 或工具栏上的 “Ping 测试” 按钮。软件将对选中的设备进行 ICMP Ping 测试，默认发送 4 个数据包，您可在设置中修改数据包数量和超时时间。</li>
                                    <li>TCP Ping 测试 ：在工具栏的 “Ping 类型” 下拉菜单中选择 “TCP” 模式，在 “端口” 输入框中输入目标端口号（如 80、443 等），然后点击 “Ping 测试” 按钮。软件将尝试与设备的指定端口建立 TCP 连接，检测端口是否开放。</li>
                                    <li>UDP Ping 测试 ：在 “Ping 类型” 下拉菜单中选择 “UDP” 模式，输入目标端口号后点击 “Ping 测试” 按钮。软件将发送 UDP 数据包到指定端口，并等待响应（UDP 测试通常用于检测端口是否可到达，但不像 TCP 那样可靠）。</li>
                                    <li>持续 Ping 测试 ：选中设备后，点击 “设备测试” -> “持续 Ping 测试”，软件将持续对设备进行 Ping 操作，直到您手动停止。可通过 “停止操作” 按钮终止测试。</li>
                                    <li>指定次数 Ping 测试 ：点击 “设备测试” -> “指定次数 Ping 测试”，输入要发送的数据包数量后点击 “开始 Ping”，软件将按指定次数进行 Ping 测试。</li>
                                </ul>
                            </li>
                            <li>Login 测试
                                <ul>
                                    <li>执行测试 ：选中设备后，点击 “设备测试” -> “Login 测试” 或工具栏上的 “Login 测试” 按钮。软件将使用设备信息中的账号和密码尝试登录设备，并在 “test_result” 列显示登录结果。</li>
                                    <li>结果解读 ：测试结果可能包括 “成功”（登录成功）、“失败”（登录失败，可能由于账号错误、网络问题等）、“错误”（其他异常情况）等状态。</li>
                                </ul>
                            </li>
                        </ol>

                        <h3>（四）定时任务管理</h3>
                        <ol>
                            <li>创建定时任务
                                <ul>
                                    <li>添加任务 ：点击 “任务管理” -> “添加定时任务”。</li>
                                    <li>设置任务参数 ：
                                        <ul>
                                            <li>时间设置 ：输入任务执行时间，格式为 HH:MM（小时：分钟）。例如，“23:59” 表示每天晚上 11 点 59 分执行任务。</li>
                                            <li>操作类型 ：从下拉菜单中选择任务类型，包括 Login 测试、执行脚本、Ping 测试。</li>
                                            <li>任务描述（可选） ：您可输入任务描述，便于识别和管理任务。</li>
                                        </ul>
                                    </li>
                                    <li>确认添加 ：点击 “添加任务” 按钮，任务将被添加到任务列表中。</li>
                                </ul>
                            </li>
                            <li>管理定时任务
                                <ul>
                                    <li>任务列表查看 ：在主界面的定时任务区域（通常位于右侧或底部，具体布局取决于软件配置），您可看到所有已创建的定时任务，包括任务名称、时间、状态等信息。</li>
                                    <li>删除任务 ：选中任务后，点击 “任务管理” -> “删除定时任务” 或右键点击任务选择 “删除”。</li>
                                    <li>暂停 / 恢复任务 ：选中任务后，点击 “任务管理” -> “暂停定时任务” 或 “恢复定时任务”，也可通过右键菜单进行操作。暂停后的任务将不会按计划执行，恢复后将继续按原计划执行。</li>
                                </ul>
                            </li>
                        </ol>

                        <h3>（五）结果导出与报告生成</h3>
                        <ol>
                            <li>导出脚本执行结果
                                <ul>
                                    <li>导出为 ZIP 文件 ：执行脚本后，点击 “文件” -> “导出脚本执行结果 ZIP”，选择保存路径。导出的 ZIP 文件包含每个设备的执行结果，文件格式包括 Excel、TXT、HTML 等，方便您存档和分享。</li>
                                    <li>导出为 HTML 报告 ：点击 “文件” -> “导出脚本执行报告”，选择保存路径。生成的 HTML 报告可在浏览器中打开，直观展示所有设备的执行结果，包括命令、输出、执行状态等信息。</li>
                                </ul>
                            </li>
                            <li>导出测试结果
                                <ul>
                                    <li>Ping 测试结果导出 ：执行 Ping 测试后，点击 “文件” -> “导出执行错误报告 ZIP”，可导出包含测试结果的 ZIP 文件，特别是对于测试失败的设备，会详细记录错误信息。</li>
                                    <li>Login 测试结果导出 ：Login 测试结果将显示在设备信息表格的 “test_result” 列，您可通过导出设备信息功能保存测试结果。</li>
                                </ul>
                            </li>
                            <li>导出选中设备信息
                                <ul>
                                    <li>导出为 Excel 文件 ：选中设备后，点击 “文件” -> “导出选中设备”，选择保存路径和文件名。软件将选中的设备信息导出为 Excel 文件，方便备份和进一步分析。</li>
                                </ul>
                            </li>
                        </ol>

                        <h3>（六）日志与审计功能</h3>
                        <ol>
                            <li>操作日志
                                <ul>
                                    <li>查看日志 ：点击 “审计” -> “操作日志审计” 或直接查看主界面底部的操作日志区域。日志按时间顺序记录所有用户操作和系统事件，包括设备加载、命令执行、测试操作等。</li>
                                    <li>导出日志 ：点击 “文件” -> “导出日志”，选择保存路径和文件名，可将日志导出为文本文件，便于存档和分析。</li>
                                </ul>
                            </li>
                            <li>命令执行审计
                                <ul>
                                    <li>查看审计记录 ：点击 “审计” -> “命令执行审计”，打开命令执行审计窗口。此处记录了所有命令脚本的执行历史，包括执行时间、设备、命令、输出等详细信息。</li>
                                    <li>导出审计记录 ：在命令执行审计窗口中，可通过相应功能导出审计记录，支持多种格式。</li>
                                </ul>
                            </li>
                        </ol>

                        <h3>（七）设置与个性化</h3>
                        <ol>
                            <li>通知设置
                                <ul>
                                    <li>启用通知 ：在设置菜单中勾选 “启用通知”，并填写 Server 酱密钥或 Bark 设备密钥。当任务完成或出现错误时，软件将通过您选择的通知服务发送消息到您的设备。</li>
                                    <li>测试通知 ：在设置菜单中点击 “测试通知” 按钮，可发送测试消息，验证通知功能是否正常工作。</li>
                                </ul>
                            </li>
                            <li>命令执行模式
                                <ul>
                                    <li>自适应模式与定时模式 ：在设置菜单中，您可切换命令执行模式。自适应模式（send_command_expect）会自动等待命令执行完成，而定时模式（send_command_timing）会按固定时间间隔读取输出。选择合适的模式可提高命令执行的准确性和效率。</li>
                                </ul>
                            </li>
                            <li>界面显示设置
                                <ul>
                                    <li>主题切换 ：在设置菜单中选择亮色或暗色主题，改善在不同环境光线下使用软件的视觉体验。</li>
                                    <li>字体设置 ：调整界面字体大小和样式，确保文本信息清晰可读。</li>
                                    <li>语言设置 ：软件支持多语言界面（当前版本主要为中文），您可在设置菜单中更改语言。</li>
                                </ul>
                            </li>
                        </ol>

                        <h3>（八）一键操作功能</h3>
                        <ol>
                            <li>一键备份
                                <ul>
                                    <li>执行备份 ：点击 “文件” -> “一键备份”，软件将自动将当前配置、设备信息、命令模板、日志等数据备份为 ZIP 文件。备份文件默认保存在软件安装目录或指定的备份目录下。</li>
                                    <li>备份内容 ：备份内容包括设备信息、命令模板、配置文件、操作日志、审计日志等，确保您可完整恢复软件状态。</li>
                                </ul>
                            </li>
                            <li>一键还原
                                <ul>
                                    <li>选择备份文件 ：点击 “文件” -> “一键还原”，选择之前创建的备份文件。</li>
                                    <li>执行还原 ：确认还原操作后，软件将从备份文件中恢复配置和数据，恢复过程会提示您是否覆盖现有数据，请谨慎操作。</li>
                                </ul>
                            </li>
                            <li>一键清理
                                <ul>
                                    <li>打开清理窗口 ：点击 “文件” -> “一键清理”，打开清理对话框。</li>
                                    <li>选择清理内容 ：在对话框中勾选要清理的内容，包括操作日志、审计日志、设备执行结果、命令模板、配置文件等。</li>
                                    <li>执行清理 ：确认选择后点击 “清理” 按钮，软件将删除相应文件和数据，释放存储空间。</li>
                                </ul>
                            </li>
                        </ol>
                    </div>

                    <div class="chapter">
                        <h2>七、高级功能指南</h2>
                        <h3>（一）命令模板管理</h3>
                        <ol>
                            <li>创建命令模板
                                <ul>
                                    <li>打开管理窗口 ：点击 “命令模板” -> “打开命令模板管理”。</li>
                                    <li>新建模板 ：点击 “新建” 按钮，输入模板名称和描述。</li>
                                    <li>编辑命令 ：在模板编辑区域输入命令，每行一个命令。您可使用变量（如 {hostname}）使模板更具通用性。</li>
                                    <li>保存模板 ：点击 “保存” 按钮保存命令模板。</li>
                                </ul>
                            </li>
                            <li>应用命令模板
                                <ul>
                                    <li>选择模板 ：在设备信息表格中选中设备后，点击 “命令模板” -> “应用模板”，选择要应用的模板。</li>
                                    <li>执行模板 ：点击 “执行” 按钮，软件将使用选中的模板对设备执行命令。</li>
                                </ul>
                            </li>
                            <li>导入与导出模板
                                <ul>
                                    <li>导入模板 ：点击 “命令模板” -> “导入模板”，选择模板文件（JSON 格式）进行导入。</li>
                                    <li>导出模板 ：选中模板后，点击 “导出” 按钮，将模板导出为文件，方便在其他设备或与其他用户共享。</li>
                                </ul>
                            </li>
                        </ol>

                        <h3>（二）错误检测与处理</h3>
                        <ol>
                            <li>错误检测设置
                                <ul>
                                    <li>管理错误字符规则 ：点击 “设置” -> “管理错误字符规则”，编辑用于检测命令执行错误的关键字符。例如，将 “Error”、“Invalid” 等字符添加到列表中。</li>
                                    <li>保存设置 ：编辑完成后点击 “保存” 按钮，软件将在命令执行时根据这些规则检测输出内容。</li>
                                </ul>
                            </li>
                            <li>错误处理机制
                                <ul>
                                    <li>错误提醒 ：当命令执行输出中包含匹配的错误字符时，软件将弹出错误提醒窗口，询问您是否继续执行后续命令。</li>
                                    <li>不再提醒选项 ：在错误提醒窗口中，可选择 “不再提醒” 选项，软件将记录您的选择，后续遇到相同错误时不再弹出提醒。</li>
                                </ul>
                            </li>
                        </ol>

                        <h3>（三）超级密码管理</h3>
                        <ol>
                            <li>设置超级密码检测字符
                                <ul>
                                    <li>管理检测字符 ：点击 “设置” -> “管理超级密码检测字符”，编辑用于检测超级密码提示的关键字符。例如，将 “>”、“#”、“]” 等提示符添加到列表中。</li>
                                    <li>保存设置 ：编辑完成后点击 “保存” 按钮，软件将在执行命令时根据这些提示符判断是否需要输入超级密码。</li>
                                </ul>
                            </li>
                            <li>超级密码使用
                                <ul>
                                    <li>自动输入超级密码 ：在设备信息中填写超级密码后，软件在检测到超级密码提示时会自动输入您设置的超级密码，提高命令执行的自动化程度。</li>
                                </ul>
                            </li>
                        </ol>

                        <h3>（四）图表与数据分析</h3>
                        <ol>
                            <li>执行结果统计图表
                                <ul>
                                    <li>查看图表 ：点击 “图表” -> “执行结果统计”，软件将生成柱状图或饼图，直观展示命令执行结果的分布情况（如成功、失败、错误的比例）。</li>
                                    <li>导出图表 ：在图表窗口中，可通过 “导出” 按钮将图表保存为图片文件（支持 PNG、JPG 等格式）。</li>
                                </ul>
                            </li>
                            <li>测试结果趋势分析
                                <ul>
                                    <li>生成趋势图 ：点击 “图表” -> “测试结果趋势”，选择时间范围和测试类型（如 Ping 测试、Login 测试），软件将生成趋势图，展示测试结果随时间的变化情况。</li>
                                    <li>导出分析报告 ：在趋势分析窗口中，可导出包含图表和数据的分析报告，支持 PDF 和 Excel 格式。</li>
                                </ul>
                            </li>
                        </ol>
                    </div>

                    <div class="chapter">
                        <h2>八、最佳实践与案例</h2>
                        <h3>（一）网络设备配置备份</h3>
                        <ol>
                            <li>场景描述 ：定期备份网络设备配置，确保在设备故障或配置更改失误时可快速恢复。</li>
                            <li>操作步骤 ：
                                <ol>
                                    <li>准备命令脚本 ：创建一个包含 show running-config 命令的脚本文件，用于获取设备当前运行配置。</li>
                                    <li>加载设备信息 ：导入包含所有需要备份的网络设备信息的 Excel 文件。</li>
                                    <li>设置定时任务 ：创建一个定时任务，每天凌晨 2 点执行配置备份脚本。</li>
                                    <li>导出结果 ：每次备份完成后，导出执行结果 ZIP 文件，保存到指定的备份存储位置。</li>
                                </ol>
                            </li>
                        </ol>

                        <h3>（二）多设备固件升级</h3>
                        <ol>
                            <li>场景描述 ：对一批网络设备进行固件升级，确保设备运行最新稳定版本的软件。</li>
                            <li>操作步骤 ：
                                <ol>
                                    <li>准备升级脚本 ：编写命令脚本，包含固件升级所需的命令序列，如 copy tftp://firmware.bin flash:、reload 等。</li>
                                    <li>测试脚本 ：在少量设备上手动执行脚本，验证升级过程和命令的正确性。</li>
                                    <li>批量执行 ：加载所有目标设备信息，执行升级脚本。可在非工作时间进行，避免对网络服务造成影响。</li>
                                    <li>监控进度 ：通过实时日志和进度条监控升级进度，及时处理可能出现的错误。</li>
                                </ol>
                            </li>
                        </ol>

                        <h3>（三）网络连通性监测</h3>
                        <ol>
                            <li>场景描述 ：持续监测关键网络设备的连通性，确保网络稳定运行。</li>
                            <li>操作步骤 ：
                                <ol>
                                    <li>设置 Ping 测试定时任务 ：创建定时任务，每 5 分钟对关键设备进行 ICMP Ping 测试。</li>
                                    <li>启用通知功能 ：配置通知服务，在连通性测试失败时及时收到警报。</li>
                                    <li>分析历史数据 ：定期查看测试结果和日志，分析网络连通性趋势，提前发现潜在问题。</li>
                                </ol>
                            </li>
                        </ol>
                    </div>

                    <div class="chapter">
                        <h2>九、常见问题解答</h2>
                        <h3>（一）设备连接问题</h3>
                        <ol>
                            <li>问题描述 ：无法连接到设备，Login 测试结果显示失败。</li>
                            <li>可能原因 ：
                                <ul>
                                    <li>网络问题 ：设备与管理机之间的网络连接不通，可能是路由、交换机配置错误或物理链路故障。</li>
                                    <li>设备配置问题 ：设备的 IP 地址、用户名、密码、协议端口等配置错误。</li>
                                    <li>访问控制限制 ：设备的访问控制列表（ACL）或防火墙规则阻止了管理连接。</li>
                                    <li>设备服务未启用 ：设备上未启用相应的服务（如 Telnet、SSH）。</li>
                                </ul>
                            </li>
                            <li>解决方法 ：
                                <ul>
                                    <li>检查网络连接 ：使用 Ping 命令测试设备 IP 地址的连通性，排查网络故障。</li>
                                    <li>验证设备信息 ：仔细检查设备信息表格中的主机名、用户名、密码、协议、端口等字段，确保填写正确。</li>
                                    <li>检查设备配置 ：登录设备（如果可能），检查服务是否启用，ACL 和防火墙规则是否允许管理访问。</li>
                                    <li>测试不同端口 ：如果默认端口被占用或更改，尝试在设备信息的 “端口” 字段中指定正确的服务端口号。</li>
                                </ul>
                            </li>
                        </ol>

                        <h3>（二）命令执行问题</h3>
                        <ol>
                            <li>问题描述 ：命令执行失败，设备返回错误信息。</li>
                            <li>可能原因 ：
                                <ul>
                                    <li>命令语法错误 ：输入的命令不符合设备的命令行语法。</li>
                                    <li>权限不足 ：当前账号没有足够的权限执行某些命令。</li>
                                    <li>设备模式错误 ：命令需要在特定模式（如全局配置模式）下执行，但未正确进入该模式。</li>
                                    <li>软件版本不兼容 ：命令在设备当前运行的软件版本中不被支持。</li>
                                </ul>
                            </li>
                            <li>解决方法 ：
                                <ul>
                                    <li>检查命令语法 ：参考设备官方文档，确保命令语法正确。可在设备上手动执行命令进行测试。</li>
                                    <li>验证账号权限 ：确认使用的账号具有执行该命令所需的权限级别。如有需要，可在设备上提升账号权限。</li>
                                    <li>调整命令执行顺序 ：在命令脚本中添加必要的命令，确保进入正确的模式后再执行目标命令。</li>
                                    <li>检查设备软件版本 ：确认设备运行的软件版本支持该命令。如不支持，考虑升级设备固件或调整命令以适应当前版本。</li>
                                </ul>
                            </li>
                        </ol>

                        <h3>（三）定时任务未执行问题</h3>
                        <ol>
                            <li>问题描述 ：设置的定时任务未按计划执行。</li>
                            <li>可能原因 ：
                                <ul>
                                    <li>时间设置错误 ：任务时间设置有误，如格式不正确或时间已过。</li>
                                    <li>任务状态异常 ：任务被意外暂停或禁用。</li>
                                    <li>软件未运行 ：软件未启动或在任务执行时间点未处于运行状态。</li>
                                    <li>权限问题 ：软件运行权限不足，导致无法创建定时任务或执行相关操作。</li>
                                </ul>
                            </li>
                            <li>解决方法 ：
                                <ul>
                                    <li>检查任务时间设置 ：确保任务时间格式正确（HH:MM），且设置的时间尚未过期。</li>
                                    <li>查看任务状态 ：在任务列表中检查任务状态，如为 “已暂停” 或 “已禁用”，重新启用任务。</li>
                                    <li>确保软件运行 ：确认软件在任务计划执行时间处于运行状态。如需要长期运行，可将软件添加到系统启动项。</li>
                                    <li>检查软件权限 ：以管理员身份运行软件，确保其具有足够的权限创建和执行定时任务。</li>
                                </ul>
                            </li>
                        </ol>
                    </div>

                    <div class="chapter">
                        <h2>十、故障排除与技术支持</h2>
                        <h3>（一）故障排除步骤</h3>
                        <ol>
                            <li>查看操作日志 ：操作日志详细记录了软件的所有操作和系统事件，是排查问题的第一步。通过日志信息，您可以了解操作的执行过程和可能的错误点。</li>
                            <li>检查设备信息 ：确保设备信息填写正确，包括主机名、用户名、密码、协议、端口等。错误的设备信息是导致连接失败和命令执行错误的常见原因。</li>
                            <li>验证网络环境 ：测试设备与管理机之间的网络连通性，确保无网络故障影响软件功能。</li>
                            <li>重新执行操作 ：在某些情况下，网络抖动或其他临时性问题可能导致操作失败。重新执行操作可能解决问题。</li>
                            <li>检查软件版本 ：确保您使用的是最新版本的软件。部分问题可能已在新版本中得到修复。您可在官方网站查看版本更新信息并下载最新版本。</li>
                        </ol>

                        <h3>（二）技术支持</h3>
                        <ol>
                            <li>官方支持渠道 ：
                                <ul>
                                    <li>电子邮件 ：您可通过官方支持邮箱（qdbihc@163.com或帮助菜单的反馈）提交问题报告。请详细描述问题现象、操作步骤、设备信息和软件版本，附上操作日志文件以便我们更快地为您解决问题。</li>
                                    <li>在线客服 ：访问官方网站的在线客服系统，在工作时间内与技术支持人员实时沟通。</li>
                                    <li>论坛社区 ：加入软件的官方论坛或用户社区，在此处您可以与其他用户交流使用经验，查找常见问题的解决方案，或向版主和技术专家提问。</li>
                                </ul>
                            </li>
                            <li>提交问题报告
                                <ul>
                                    <li>详细描述问题 ：包括问题发生的时间、操作步骤、预期结果和实际结果。</li>
                                    <li>提供环境信息 ：如操作系统版本、软件版本、设备型号和固件版本等。</li>
                                    <li>附上日志文件 ：将操作日志和相关测试结果导出，作为附件提交给技术支持团队。</li>
                                    <li>屏幕截图 ：如有必要，提供软件界面的屏幕截图，突出显示问题所在区域。</li>
                                </ul>
                            </li>
                        </ol>
                    </div>
                    <div class="chapter">
                        <h2>十二、结语</h2>
                        <p>恭喜您完成网络大剑仙的学习旅程！本手册旨在帮助您全面掌握软件的各项功能，从基础的设备管理和命令执行到高级的定时任务、数据分析和自动化操作。通过实践操作和不断探索，您将能够充分利用本软件提升网络管理效率，降低运维成本。</p>
                        <p>在使用过程中，如遇到任何问题或有功能改进建议，欢迎随时与我们联系。您的反馈对我们至关重要，将帮助我们持续改进软件，满足更多用户的需求。</p>
                        <p>祝您在网络管理工作中事半功倍，网络大剑仙始终是您值得信赖的得力助手！</p>
                    </div>

                    <div class="footer">
                        <p>网络大剑仙软件用户使用指导手册 - 初版发布</p>
                        <p>作者：WB</p>
                        <p>日期：2025 年 04 月 01 日</p>
                    </div>
                </div>
            </body>
            </html>
            """

    def create_widgets(self):
        self.create_menu_bar()  # 创建菜单栏

        buttons_frame = tk.Frame(self.root)  # 创建按钮框架
        buttons_frame.grid(row=0, column=0, columnspan=8, padx=10, pady=10, sticky="w")  # 将按钮框架添加到主窗口

        # 设备文件选择区域
        tk.Label(buttons_frame, text="设备文件:").grid(row=0, column=0, padx=5, pady=5)  # 创建设备文件标签
        self.select_file_button = tk.Button(buttons_frame, text="选择", command=self.select_device_file)  # 创建选择设备文件按钮
        self.select_file_button.grid(row=0, column=1, padx=5, pady=5)  # 将选择设备文件按钮添加到按钮框架
        # 为按钮绑定提示效果和背景色变化效果
        # self.select_file_button.bind("<Enter>", lambda e: self.select_file_button.config(bg="lightgray"))  # 鼠标悬停时变色
        # self.select_file_button.bind("<Leave>",
        #                              lambda e: self.select_file_button.config(bg="SystemButtonFace"))  # 鼠标离开时恢复

        self.download_template_button = tk.Button(buttons_frame, text="下载模板",
                                                  command=self.download_template)  # 创建下载模板按钮
        self.download_template_button.grid(row=0, column=2, padx=5, pady=5)  # 将下载模板按钮添加到按钮框架
        # self.download_template_button.bind("<Enter>", lambda e: self.download_template_button.config(bg="lightgray"))
        # self.download_template_button.bind("<Leave>",
        #                                    lambda e: self.download_template_button.config(bg="SystemButtonFace"))

        # 保存目录选择区域
        tk.Label(buttons_frame, text="保存目录:").grid(row=0, column=3, padx=5, pady=5)  # 创建保存目录标签
        self.select_directory_button = tk.Button(buttons_frame, text="选择",
                                                 command=self.select_output_directory)  # 创建选择保存目录按钮
        self.select_directory_button.grid(row=0, column=4, padx=5, pady=5)  # 将选择保存目录按钮添加到按钮框架
        # self.select_directory_button.bind("<Enter>", lambda e: self.select_directory_button.config(bg="lightgray"))
        # self.select_directory_button.bind("<Leave>",
        #                                   lambda e: self.select_directory_button.config(bg="SystemButtonFace"))

        # 命令脚本选择区域
        tk.Label(buttons_frame, text="命令脚本:").grid(row=0, column=5, padx=5, pady=5)  # 创建命令脚本标签
        self.select_command_file_button = tk.Button(buttons_frame, text="选择",
                                                    command=self.select_command_file)  # 创建选择命令脚本按钮
        self.select_command_file_button.grid(row=0, column=6, padx=5, pady=5)  # 将选择命令脚本按钮添加到按钮框架
        # self.select_command_file_button.bind("<Enter>",
        #                                      lambda e: self.select_command_file_button.config(bg="lightgray"))
        # self.select_command_file_button.bind("<Leave>",
        #                                      lambda e: self.select_command_file_button.config(bg="SystemButtonFace"))

        # 测试按钮区域
        self.test_button = tk.Button(buttons_frame, text="Login 测试",
                                     command=self.start_test_connections)  # 创建Login测试按钮
        self.test_button.grid(row=0, column=8, padx=5, pady=5)  # 将Login测试按钮添加到按钮框架
        # self.test_button.bind("<Enter>", lambda e: self.test_button.config(bg="lightgray"))
        # self.test_button.bind("<Leave>", lambda e: self.test_button.config(bg="SystemButtonFace"))

        # self.execute_button = tk.Button(buttons_frame, text="执行脚本",
        #                                 command=self.start_command_execution)  # 创建执行脚本按钮
        # self.execute_button.grid(row=0, column=9, padx=5, pady=5)  # 将执行脚本按钮添加到按钮框架

        # 执行脚本按钮-改为暂停函数
        self.execute_button = tk.Button(buttons_frame, text="执行脚本", command=self.start_command_execution)
        self.execute_button.grid(row=0, column=9, padx=5, pady=5)

        # self.execute_button.bind("<Enter>", lambda e: self.execute_button.config(bg="lightgray"))
        # self.execute_button.bind("<Leave>", lambda e: self.execute_button.config(bg="SystemButtonFace"))

        self.ping_test_button = tk.Button(buttons_frame, text="Ping 测试", command=self.start_ping_test)  # 创建Ping测试按钮
        self.ping_test_button.grid(row=0, column=7, padx=5, pady=5)  # 将Ping测试按钮添加到按钮框架
        # self.ping_test_button.bind("<Enter>", lambda e: self.ping_test_button.config(bg="lightgray"))
        # self.ping_test_button.bind("<Leave>", lambda e: self.ping_test_button.config(bg="SystemButtonFace"))

        # 取消操作按钮
        self.cancel_button = tk.Button(buttons_frame, text="取消操作", command=self.cancel_operation,
                                       state=tk.DISABLED)  # 创建取消操作按钮
        self.cancel_button.grid(row=0, column=10, padx=5, pady=5)  # 将取消操作按钮添加到按钮框架
        # self.cancel_button.bind("<Enter>", lambda e: self.cancel_button.config(bg="lightgray"))
        # self.cancel_button.bind("<Leave>", lambda e: self.cancel_button.config(bg="SystemButtonFace"))



        # Server酱密钥输入区域
        server_frame = tk.Frame(self.root)  # 创建Server酱框架
        server_frame.grid(row=1, column=0, columnspan=8, padx=10, pady=10, sticky="w")  # 将Server酱框架添加到主窗口

        tk.Label(server_frame, text="Server酱密钥:").grid(row=0, column=0, padx=5, pady=5)  # 创建Server酱密钥标签
        self.server_chan_key_entry = tk.Entry(server_frame, width=30)  # 创建Server酱密钥输入框
        self.server_chan_key_entry.grid(row=0, column=1, padx=5, pady=5)  # 将Server酱密钥输入框添加到Server酱框架

        self.server_chan_key_entry.bind("<FocusOut>", lambda event: self.save_keys_to_json())  # 保存密钥
        # self.help_button = tk.Button(server_frame, text="使用帮助", command=self.show_help)  # 创建使用帮助按钮
        # self.help_button.grid(row=0, column=15, padx=5, pady=5)  # 将使用帮助按钮添加到Server酱框架

        self.notification_checkbox = tk.Checkbutton(server_frame, text="启用微信通知",
                                                    variable=self.enable_notification)  # 创建启用微信通知复选框


        # 设置默认值
        self.server_chan_key_entry.insert(0, "请输入API密钥")  # 设置默认值
        self.server_chan_key_entry.config(fg="gray")  # 设置文本颜色为灰色

        # 分组选择区域
        group_frame = tk.Frame(self.root)
        group_frame.grid(row=1, column=7, padx=1, pady=10, sticky="w")

        tk.Label(group_frame, text="命令间隔:").grid(row=0, column=3, padx=0, pady=5, sticky="w")
        self.cmd_interval_entry = tk.Entry(group_frame, width=2)
        self.cmd_interval_entry.grid(row=0, column=4, padx=5, pady=5, sticky="w")
        self.cmd_interval_entry.insert(0, "1")
        self.cmd_interval_entry.config(fg="gray")

        tk.Label(group_frame, text="超时设置:").grid(row=0, column=0, padx=0, pady=1, sticky="w")
        self.timeout_entry = tk.Entry(group_frame, width=2)
        self.timeout_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.timeout_entry.insert(0, "3")
        self.timeout_entry.config(fg="gray")

        tk.Label(group_frame, text="局点:").grid(row=0, column=5, padx=0, pady=5, sticky="e")
        self.current_group_var = tk.StringVar(value="默认局点")
        self.group_menu = ttk.Combobox(group_frame, textvariable=self.current_group_var, width=8)
        self.group_menu.grid(row=0, column=6, padx=5, pady=5, sticky="e")
        self.group_menu.bind("<<ComboboxSelected>>", self.on_group_change)

        # 新增“新建局点”按钮
        self.create_group_button = tk.Button(group_frame, text="新建", command=self.create_group, font=("微软雅黑", 7))
        self.create_group_button.grid(row=0, column=7, padx=5, pady=5, sticky="w")
        self.create_group_button.bind("<Enter>", lambda e: self.create_group_button.config(bg="lightgray"))
        self.create_group_button.bind("<Leave>", lambda e: self.create_group_button.config(bg="SystemButtonFace"))


        help_button_frame = tk.Frame(self.root)  # 创建使用说明框架
        help_button_frame.grid(row=1, column=7, padx=1, pady=1, sticky="e")  # 将使用说明框架添加到主窗口

        self.help_button = tk.Button(help_button_frame, text="使用说明", command=self.show_help, font=("微软雅黑", 9),
                                     fg="blue", bg="white")
        self.help_button.grid(row=1, column=7, padx=5, pady=5)  # 将使用帮助按钮添加到Server酱框架
        self.help_button.bind("<Enter>", lambda e: self.help_button.config(bg="lightgray"))
        self.help_button.bind("<Leave>", lambda e: self.help_button.config(bg="white"))
        # 创建“使用说明”按钮并添加闪烁效果
        self.help_button.bind("<Enter>", lambda e: self.help_button.config(bg="lightgray"))
        self.help_button.bind("<Leave>", lambda e: self.help_button.config(bg="white"))
        self.start_help_button_flash()

        # # 设备信息窗口，支持滚动和自适应宽度
        # device_frame = tk.LabelFrame(self.root, text="设备信息")
        # device_frame.grid(row=2, column=0, columnspan=8, padx=10, pady=10, sticky="ew")  # 调整为"ew"
        # device_frame.grid_columnconfigure(0, weight=1)  # 设置设备信息框架的列权重

        # 创建设备信息面板
        self.device_frame = tk.LabelFrame(self.root, text=f"设备&局点信息 - {self.current_group_var.get()}")
        self.device_frame.grid(row=2, column=0, columnspan=8, padx=10, pady=10, sticky="ew")
        self.device_frame.grid_columnconfigure(0, weight=1)

        # 创建一个包含Treeview和垂直滚动条的Frame
        tree_frame = tk.Frame(self.device_frame)
        tree_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)  # 调整为填满父框架

        # 创建Treeview
        columns = ("hostname", "username", "password", "protocol", "device_type", "path", "port",
                   "test_result", "execute_result", "super_password")  # 定义Treeview的列
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='headings', selectmode="extended")  # 创建Treeview，支持多选
        self.column_sort_direction = {col: False for col in columns}  # 跟踪每列的排序方向

        # 设置列宽和伸缩性
        self.tree.column("hostname", width=100, anchor="w")
        self.tree.column("username", width=70, anchor="w")
        self.tree.column("password", width=70, anchor="w")
        self.tree.column("protocol", width=60, anchor="w")
        self.tree.column("device_type", width=80, anchor="w")
        self.tree.column("path", width=110, anchor="w")
        self.tree.column("port", width=40, anchor="w")
        self.tree.column("test_result", width=110, anchor="w")
        self.tree.column("execute_result", width=130, anchor="w")
        self.tree.column("super_password", width=60, anchor="w")
        # 配置标签样式
        self.tree.tag_configure("success", background="#D6EEEE")  # 淡蓝色背景
        self.tree.tag_configure("failure", background="#FFFAD6")  # 淡黄色背景

        # 启动跑马灯效果
        self.marquee_title = self.device_frame.cget("text")
        self.marquee_offset = 0
        self.update_marquee_title()

        for col in columns:
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_treeview_column(c))  # 调整排序函数调用
            self.tree.column(col, stretch=True)  # 启用自适应宽度

        self.tree.pack(fill=tk.BOTH, expand=True)

        # 绑定拖拽事件到 Treeview
        windnd.hook_dropfiles(self.tree.winfo_id(), self.on_drop_files)

        # 绑定快捷键到 Treeview
        self.tree.bind("<Control-z>", self.undo)  # 撤销
        self.tree.bind("<Control-Z>", self.undo)  # 撤销（大写Z）
        self.tree.bind("<Control-y>", self.redo)  # 重做
        self.tree.bind("<Control-Y>", self.redo)  # 重做（大写Y）
        self.tree.bind("<Control-c>", self.copy_selected)  # 复制
        self.tree.bind("<Control-C>", self.copy_selected)  # 复制（大写C）
        self.tree.bind("<Control-v>", self.paste_selected)  # 粘贴
        self.tree.bind("<Control-V>", self.paste_selected)  # 粘贴（大写V）
        self.tree.bind("<Control-x>", self.cut_selected)  # 剪切
        self.tree.bind("<Control-X>", self.cut_selected)  # 剪切（大写X）
        # self.tree.bind("<Delete>", self.delete_selected)  # 删除
        # self.tree.bind("<Control-a>", self.select_all)  # 全选
        # self.tree.bind("<Control-A>", self.select_all)  # 全选（大写A）

        # 存储原始背景色
        self.original_bg_color = "SystemButtonFace"  # 默认背景色

        # 绑定鼠标事件到 Treeview 以模拟拖拽效果
        self.tree.bind("<B1-Motion>", self.on_drag_enter)  # 按住鼠标左键拖动
        self.tree.bind("<ButtonRelease-1>", self.on_drag_leave)  # 鼠标左键释放

        # 添加垂直滚动条
        tree_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)  # 创建垂直滚动条
        self.tree.configure(yscrollcommand=tree_scrollbar.set)  # 将Treeview与滚动条关联

        # 使用grid布局Treeview和滚动条
        self.tree.grid(row=0, column=0, sticky='nsew')  # 将Treeview添加到Treeview框架
        tree_scrollbar.grid(row=0, column=1, sticky='ns')  # 将滚动条添加到Treeview框架

        # 配置Grid的权重
        tree_frame.grid_rowconfigure(0, weight=1)  # 设置Treeview框架的行权重
        tree_frame.grid_columnconfigure(0, weight=1)  # 设置Treeview框架的列权重

        display_frame = tk.Frame(self.root)  # 创建显示框架
        display_frame.grid(row=3, column=0, columnspan=8, padx=10, pady=10, sticky="nsew")  # 将显示框架添加到主窗口

        output_frame = tk.LabelFrame(display_frame, text="操作日志", width=700, height=200)  # 创建操作日志框架
        output_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")  # 将操作日志框架添加到显示框架
        output_frame.grid_propagate(False)  # 禁止操作日志框架自动调整大小

        self.output_text = ScrolledText(output_frame, height=15, width=60)  # 创建带滚动条的文本框
        self.output_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)  # 将文本框添加到操作日志框架

        scheduler_frame = tk.LabelFrame(display_frame, text="设置定时任务", width=400, height=200)  # 创建定时任务框架
        scheduler_frame.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")  # 将定时任务框架添加到显示框架
        scheduler_frame.grid_propagate(False)  # 禁止定时任务框架自动调整大小

        # 调整列宽权重，使元素分布更合理
        scheduler_frame.columnconfigure(0, weight=1)
        scheduler_frame.columnconfigure(1, weight=1)
        scheduler_frame.columnconfigure(2, weight=1)
        scheduler_frame.columnconfigure(3, weight=1)

        tk.Label(scheduler_frame, text="时间(HH:MM)").grid(row=0, column=0, padx=2, pady=2)  # 创建时间标签
        self.task_entry = tk.Entry(scheduler_frame, width=5)  # 创建时间输入框
        self.task_entry.grid(row=0, column=1, padx=2, pady=2)  # 将时间输入框添加到定时任务框架

        # 减小下拉框的宽度
        self.task_menu = ttk.Combobox(scheduler_frame, values=["Login 测试", "执行脚本", "Ping 测试"],
                                      width=10)  # 创建操作类型下拉框
        self.task_menu.grid(row=0, column=2, padx=2, pady=2)  # 将操作类型下拉框添加到定时任务框架
        self.task_menu.set("选择操作")  # 设置操作类型下拉框的默认值

        self.placeholder = "00:00"  # 设置时间输入框的占位符
        self.task_entry.insert(0, self.placeholder)  # 将占位符插入时间输入框
        self.task_entry.config(fg="gray")  # 设置时间输入框文本颜色为灰色

        self.add_task_button = tk.Button(scheduler_frame, text="添加任务", command=self.add_task)  # 创建添加任务按钮
        self.add_task_button.grid(row=0, column=3, padx=2, pady=2)  # 将添加任务按钮添加到定时任务框架
        self.add_task_button.bind("<Enter>", lambda e: self.add_task_button.config(bg="lightgray"))
        self.add_task_button.bind("<Leave>", lambda e: self.add_task_button.config(bg="SystemButtonFace"))

        self.task_listbox = tk.Listbox(scheduler_frame, height=7, width=60)  # 创建任务列表
        self.task_listbox.grid(row=1, column=0, columnspan=10, padx=5, pady=5)  # 将任务列表添加到定时任务框架

        # 任务栏绑定快捷菜单功能
        self.task_listbox.bind("<Double-1>", self.edit_task)
        self.task_listbox.bind("<Control-a>", self.select_all_tasks)
        self.task_listbox.bind("<Control-1>", self.ctrl_click_select)

        self.remove_task_button = tk.Button(scheduler_frame, text="删除任务", command=self.remove_task)  # 创建取消任务按钮
        self.remove_task_button.grid(row=2, column=0, columnspan=1, padx=5, pady=5)  # 将取消任务按钮添加到定时任务框架
        self.remove_task_button.bind("<Enter>", lambda e: self.remove_task_button.config(bg="lightgray"))
        self.remove_task_button.bind("<Leave>", lambda e: self.remove_task_button.config(bg="SystemButtonFace"))

        self.pause_task_button = tk.Button(scheduler_frame, text="暂停任务", command=self.pause_task)  # 创建暂停任务按钮
        self.pause_task_button.grid(row=2, column=1, columnspan=1, padx=5, pady=5)  # 将暂停任务按钮添加到定时任务框架
        self.pause_task_button.bind("<Enter>", lambda e: self.pause_task_button.config(bg="lightgray"))
        self.pause_task_button.bind("<Leave>", lambda e: self.pause_task_button.config(bg="SystemButtonFace"))

        self.resume_task_button = tk.Button(scheduler_frame, text="恢复任务", command=self.resume_task)  # 创建恢复任务按钮
        self.resume_task_button.grid(row=2, column=2, columnspan=1, padx=5, pady=5)  # 将恢复任务按钮添加到定时任务框架
        self.resume_task_button.bind("<Enter>", lambda e: self.resume_task_button.config(bg="lightgray"))
        self.resume_task_button.bind("<Leave>", lambda e: self.resume_task_button.config(bg="SystemButtonFace"))

        # self.resume_task_button = tk.Button(scheduler_frame, text="恢复任务", command=self.resume_task)  # 创建恢复任务按钮
        # self.resume_task_button.grid(row=2, column=2, columnspan=4, padx=5, pady=5)  # 将恢复任务按钮添加到定时任务框架
        # self.resume_task_button.bind("<Enter>", lambda e: self.resume_task_button.config(bg="lightgray"))
        # self.resume_task_button.bind("<Leave>", lambda e: self.resume_task_button.config(bg="SystemButtonFace"))

        # 创建任务栏右键菜单
        self.create_context_menu()

        # 创建进度条框架
        progress_frame = tk.Frame(self.root)
        progress_frame.grid(row=4, column=0, columnspan=8, padx=10, pady=10, sticky="nsew")
        progress_frame.grid_columnconfigure(1, weight=1)

        # 定义进度条样式
        style = ttk.Style()
        style.theme_use('default')  # 使用默认主题
        style.configure("green.Horizontal.TProgressbar", background="#4CAF50")  # 绿色
        style.configure("yellow.Horizontal.TProgressbar", background="#FFC107")  # 黄色
        style.configure("red.Horizontal.TProgressbar", background="#F44336")  # 红色

        self.progress_bar = ttk.Progressbar(progress_frame, style="green.Horizontal.TProgressbar", orient="horizontal",
                                            mode="determinate")
        self.progress_bar.grid(row=0, column=1, columnspan=7, padx=5, pady=5, sticky="ew")

        # # 创建进度条
        # self.progress_bar = ttk.Progressbar(root, mode="determinate")
        # self.progress_bar.grid(row=4, column=0, columnspan=8, padx=10, pady=10, sticky="ew")

        self.status_label = tk.Label(progress_frame, text="执行状态：等待操作")  # 创建状态标签
        self.status_label.grid(row=0, column=0, padx=5, pady=5, sticky="ew")  # 将状态标签添加到进度条框架

        self.output_text.tag_config("normal", foreground="black")  # 设置正常日志的文本颜色
        self.output_text.tag_config("warning", foreground="darkorange")  # 设置警告日志的文本颜色
        self.output_text.tag_config("error", foreground="darkred")  # 设置错误日志的文本颜色

        self.output_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)  # 将文本框添加到操作日志框架

        # 密码显示控制相关
        # self.password_visible = True  # 初始密码可见状态
        # self.original_passwords = {}  # 用于存储原始密码

        # 导出报告功能
        export_frame = tk.Frame(self.root)  # 创建导出框架
        export_frame.grid(row=5, column=0, columnspan=8, padx=10, pady=10, sticky="w")  # 将导出框架添加到主窗口

        self.export_log_button = tk.Button(export_frame, text="导出日志", command=self.export_log)  # 创建导出日志按钮
        self.export_log_button.pack(side=tk.LEFT, padx=5, pady=5)  # 将导出日志按钮添加到导出框架
        self.export_log_button.bind("<Enter>", lambda e: self.export_log_button.config(bg="lightgray"))
        self.export_log_button.bind("<Leave>", lambda e: self.export_log_button.config(bg="SystemButtonFace"))

        self.export_result_button = tk.Button(export_frame, text="导出脚本执行结果汇总",
                                              command=self.export_command_result_zip)  # 创建导出脚本执行结果ZIP按钮
        self.export_result_button.pack(side=tk.LEFT, padx=5, pady=5)  # 将导出脚本执行结果ZIP按钮添加到导出框架
        self.export_result_button.bind("<Enter>", lambda e: self.export_result_button.config(bg="lightgray"))
        self.export_result_button.bind("<Leave>", lambda e: self.export_result_button.config(bg="SystemButtonFace"))

        export_report_button = tk.Button(export_frame, text="导出脚本执行报告",
                                         command=self.export_script_report)  # 创建导出脚本执行报告按钮
        export_report_button.pack(side=tk.LEFT, padx=5, pady=5)  # 将导出脚本执行报告按钮添加到导出框架
        export_report_button.bind("<Enter>", lambda e: export_report_button.config(bg="lightgray"))
        export_report_button.bind("<Leave>", lambda e: export_report_button.config(bg="SystemButtonFace"))

        export_error_report_button = tk.Button(export_frame, text="导出执行错误报告ZIP",
                                               command=self.export_error_report)
        export_error_report_button.pack(side=tk.LEFT, padx=5, pady=5)
        export_error_report_button.bind("<Enter>", lambda e: export_error_report_button.config(bg="lightgray"))
        export_error_report_button.bind("<Leave>", lambda e: export_error_report_button.config(bg="SystemButtonFace"))

        # 增加日志清理按钮
        # self.clear_logs_button = tk.Button(buttons_frame, text="清理日志", command=self.clear_logs)
        self.clear_logs_button = tk.Button(export_frame, text="清理日志", command=self.clear_logs)
        # self.clear_logs_button.grid(row=0, column=12, padx=5, pady=5)
        self.clear_logs_button.pack(side=tk.LEFT, padx=5, pady=5)
        self.clear_logs_button.bind("<Enter>", lambda e: self.clear_logs_button.config(bg="lightgray"))
        self.clear_logs_button.bind("<Leave>", lambda e: self.clear_logs_button.config(bg="SystemButtonFace"))

        # 增加一键清理按钮
        self.clean_all_button = tk.Button(export_frame, text="一键清理", command=self.show_clean_all_dialog)
        self.clean_all_button.pack(side=tk.LEFT, padx=5, pady=5)
        self.clean_all_button.bind("<Enter>", lambda e: self.clean_all_button.config(bg="lightgray"))
        self.clean_all_button.bind("<Leave>", lambda e: self.clean_all_button.config(bg="SystemButtonFace"))

        # # 增加一键备份和一键还原按钮
        # backup_frame = tk.Frame(self.root)
        # backup_frame.grid(row=5, column=0, columnspan=8, padx=10, pady=10, sticky="w")

        self.backup_button = tk.Button(export_frame, text="一键备份", command=self.create_backup)
        self.backup_button.pack(side=tk.LEFT, padx=5, pady=5)
        self.backup_button.bind("<Enter>", lambda e: self.backup_button.config(bg="lightgray"))
        self.backup_button.bind("<Leave>", lambda e: self.backup_button.config(bg="SystemButtonFace"))

        self.restore_button = tk.Button(export_frame, text="一键还原", command=self.restore_backup)
        self.restore_button.pack(side=tk.LEFT, padx=5, pady=5)
        self.restore_button.bind("<Enter>", lambda e: self.restore_button.config(bg="lightgray"))
        self.restore_button.bind("<Leave>", lambda e: self.restore_button.config(bg="SystemButtonFace"))


        server_frame = tk.Frame(self.root)  # 创建通知框架
        server_frame.grid(row=1, column=0, columnspan=8, padx=10, pady=10, sticky="w")  # 将通知框架添加到主窗口

        tk.Label(server_frame, text="通知类型:").grid(row=0, column=0, padx=1, pady=5)  # 创建通知类型标签
        self.notification_type_menu = ttk.Combobox(server_frame, textvariable=self.notification_type,
                                                   values=["Server酱", "Bark-Server", "Contoso"], width=10)  # 创建通知类型下拉框

        # self.notification_type_menu = tk.Combobox(server_frame, width=10)  # 创建通知下拉输入框
        self.notification_type_menu.grid(row=0, column=1, padx=1, pady=5)  # 将通知类型下拉框添加到通知框架

        tk.Label(server_frame, text="通知密钥:").grid(row=0, column=2, padx=1, pady=5)  # 创建通知密钥标签
        self.notification_key_entry = tk.Entry(server_frame, width=15)  # 创建通知密钥输入框
        self.server_chan_key_entry.bind("<FocusOut>", lambda event: self.save_keys_to_json())  # 保存密钥
        self.notification_key_entry.bind("<FocusOut>", lambda event: self.save_keys_to_json())  # 保存密钥
        self.notification_key_entry.grid(row=0, column=3, padx=1, pady=5)  # 将通知密钥输入框添加到通知框架
        self.notification_checkbox = tk.Checkbutton(server_frame, text=":启用通知",
                                                    variable=self.enable_notification)  # 创建启用通知复选框
        self.notification_checkbox.grid(row=0, column=4, padx=5, pady=5)  # 将启用通知复选框添加到通知框架

        # self.server_chan_key_entry.bind("<FocusOut>", lambda event: self.save_keys_to_json())
        # self.notification_key_entry.bind("<FocusOut>", lambda event: self.save_keys_to_json())

        # 以下为新增代码
        # self.placeholder = "请输入密钥"  # 设置通知密钥输入框的占位符
        # self.notification_key_entry.insert(0, self.placeholder)  # 将占位符插入通知密钥输入框
        # self.notification_key_entry.config(fg="gray")  # 设置通知密钥输入框文本颜色为灰色

        # 设置默认值
        self.notification_key_entry.insert(0, "请输入API密钥")  # 设置默认值
        self.notification_key_entry.config(fg="gray")  # 设置文本颜色为灰色

        # self.server_chan_key_entry.bind("<FocusOut>", lambda event: self.save_keys_to_json())
        # self.notification_key_entry.bind("<FocusOut>", lambda event: self.save_keys_to_json())

        # 状态配置
        self.notification_type.trace("w", self.update_notification_config)  # 绑定通知类型变化事件

        self.tree.bind("<Double-1>", self.on_tree_double_click)

        # self.load_device_info_from_json()

        # 创建呼吸灯效果的同心圆图标
        # self.create_breathing_light()

        # 创建右键菜单，设备信息编辑
        self.create_right_click_menu()

        # 为按钮添加提示词
        self.button_tooltips = {
            "选择设备文件": "选择包含设备信息的Excel文件，支持拖入设备信息文件到设备信息栏加载",
            "下载模板": "下载设备信息模板文件",
            "选择保存目录": "选择命令执行结果的保存目录",
            "选择命令脚本": "选择包含命令的脚本文件",
            "Login 测试": "测试设备的连接状态",
            "执行脚本": "执行选定设备的命令脚本",
            "Ping 测试": "测试设备的网络连通性",
            "取消操作": "取消当前正在进行的操作",
            "使用说明": "软件使用帮助",
            "导出日志": "导出程序产生的所有日志记录文件",
            "导出脚本执行结果ZIP": "导出HTML、EXCEL、TXT、汇总结果）",
            "导出脚本执行报告": "导出所有设备的脚本执行汇总情况",
            # "删除定时任务": "删除选择的定时任务",
            # "导出执行错误报告ZIP": "导出所有设备脚本执行错误的汇报信息，支持HTML、EXCEL格式"
            "清理日志": "清除程序产生的操作日志和审计日志",
            "一键清理" : "一键删除程序产生的所有文件",
        }

        # 为按钮绑定提示词显示事件
        self.select_file_button.bind("<Enter>",
                                     lambda event: self.show_tooltip(event, self.button_tooltips["选择设备文件"]))
        self.select_file_button.bind("<Motion>", self.update_tooltip_position)
        self.select_file_button.bind("<Leave>", self.hide_tooltip)

        self.download_template_button.bind("<Enter>",
                                           lambda event: self.show_tooltip(event, self.button_tooltips["下载模板"]))
        self.download_template_button.bind("<Motion>", self.update_tooltip_position)
        self.download_template_button.bind("<Leave>", self.hide_tooltip)

        self.select_directory_button.bind("<Enter>",
                                          lambda event: self.show_tooltip(event, self.button_tooltips["选择保存目录"]))
        self.select_directory_button.bind("<Motion>", self.update_tooltip_position)
        self.select_directory_button.bind("<Leave>", self.hide_tooltip)

        self.select_command_file_button.bind("<Enter>", lambda event: self.show_tooltip(event, self.button_tooltips[
            "选择命令脚本"]))
        self.select_command_file_button.bind("<Motion>", self.update_tooltip_position)
        self.select_command_file_button.bind("<Leave>", self.hide_tooltip)

        self.test_button.bind("<Enter>", lambda event: self.show_tooltip(event, self.button_tooltips["Login 测试"]))
        self.test_button.bind("<Motion>", self.update_tooltip_position)
        self.test_button.bind("<Leave>", self.hide_tooltip)

        self.execute_button.bind("<Enter>", lambda event: self.show_tooltip(event, self.button_tooltips["执行脚本"]))
        self.execute_button.bind("<Motion>", self.update_tooltip_position)
        self.execute_button.bind("<Leave>", self.hide_tooltip)

        self.ping_test_button.bind("<Enter>", lambda event: self.show_tooltip(event, self.button_tooltips["Ping 测试"]))
        self.ping_test_button.bind("<Motion>", self.update_tooltip_position)
        self.ping_test_button.bind("<Leave>", self.hide_tooltip)

        self.cancel_button.bind("<Enter>", lambda event: self.show_tooltip(event, self.button_tooltips["取消操作"]))
        self.cancel_button.bind("<Motion>", self.update_tooltip_position)
        self.cancel_button.bind("<Leave>", self.hide_tooltip)

        # self.help_button.bind("<Enter>", lambda event: self.show_tooltip(event, self.button_tooltips["使用说明"]))
        # self.help_button.bind("<Motion>", self.update_tooltip_position)
        # self.help_button.bind("<Leave>", self.hide_tooltip)

        self.export_log_button.bind("<Enter>", lambda event: self.show_tooltip(event, self.button_tooltips["导出日志"]))
        self.export_log_button.bind("<Motion>", self.update_tooltip_position)
        self.export_log_button.bind("<Leave>", self.hide_tooltip)

        self.export_result_button.bind("<Enter>",lambda event: self.show_tooltip(event, self.button_tooltips["导出脚本执行结果ZIP"]))
        self.export_result_button.bind("<Motion>", self.update_tooltip_position)
        self.export_result_button.bind("<Leave>", self.hide_tooltip)

        # self.export_report_button.bind("<Enter>", lambda event: self.show_tooltip(event, self.button_tooltips["导出脚本执行报告"]))
        # self.export_report_button.bind("<Motion>", self.update_tooltip_position)
        # self.export_report_button.bind("<Leave>", self.hide_tooltip)

    # def show_right_click_menu(self, event):
    #     self.right_click_menu.post(event.x_root, event.y_root)

        # # 创建对比功能菜单
        # self.create_compare_menu()

    # server 酱通知函数
    def send_server_sauce_notification(self, title, content, test_type):
        # title = "程序启动通知"
        # content = "启动成功"
        tags = test_type
        key = "sctp5121tx5qrgap3xf2mlrzemn16po"

        try:
            response = None  # 初始化 response
            url = f"https://sctapi.ftqq.com/{key}.send"
            data = {"title": title, "desp": content, "tags": tags}
            response = requests.post(url, data=data)
            print(f"[成功]：发送通知成功")

        except Exception as e:
            print(f"[错误]：发送通知失败: {str(e)}")

    def show_notification(self, message, duration=2):
        """
        显示桌面通知
        :param title: 通知标题
        :param message: 通知内容
        :param duration: 通知显示时长（秒）
        """
        title = "网络大剑仙"
        try:
            notification.notify(
                # app_name="网络大剑仙",
                message=message,
                title="网络大剑仙",
                timeout=duration  # 通知显示时长
            )
            print(f"通知已发送: {title} - {message}")
            self.log(f"[提示]：桌面通知已发送：{title} - {message}")

        except Exception as e:
            print(f"发送通知时出现错误: {e}")

    def create_countdown_label(self):
        # 创建一个Canvas来绘制圆形背景
        self.countdown_canvas = tk.Canvas(
            self.root,
            width=80,
            height=80,
            highlightthickness=0
        )
        self.countdown_canvas.place(relx=0.9, rely=0.05, anchor="center")

        # 绘制圆形背景
        self.circle_id = self.countdown_canvas.create_oval(
            10, 5, 70, 65,
            fill="#f0f0f0",  # 浅灰色背景
            outline="#cccccc",  # 浅灰色边框
            width=2
        )

        # 绘制秒针（初始指向12点方向）
        self.second_hand_id = self.countdown_canvas.create_line(
            40, 35, 40, 15,  # 初始位置指向12点方向
            fill="red",  # 红色秒针
            width=1
        )

        # 创建读秒标签
        self.countdown_label = self.countdown_canvas.create_text(
            40, 35,
            text="00:00",
            font=("Arial", 12),
            fill="black"
        )

        # 启动读秒
        self.start_countdown()

    def start_countdown(self):
        # 启动读秒
        self.start_time = time.time()
        self.update_countdown()

    def update_countdown(self):
        # 更新读秒
        elapsed_time = time.time() - self.start_time
        minutes = int(elapsed_time // 60)
        seconds = int(elapsed_time % 60)
        time_str = f"{minutes:02d}:{seconds:02d}"
        self.countdown_canvas.itemconfig(self.countdown_label, text=time_str)

        # 更新秒针位置（顺时针旋转）
        angle = math.radians(90 - (seconds * 6))  # 每秒顺时针旋转6度
        x = 40 + 25 * math.cos(angle)
        y = 35 + 25 * math.sin(angle)
        self.countdown_canvas.coords(self.second_hand_id, 40, 35, x, y)

        # 每秒更新一次
        self.root.after(1000, self.update_countdown)

    def create_gradient_background(self):
        # 创建一个渐变色背景
        gradient = Image.new("RGBA", (100, 30), (0, 255, 0, 128))  # 初始颜色
        for x in range(100):
            r = int(0 + (255 - 0) * (x / 100))
            g = int(255 - (255 - 0) * (x / 100))
            b = 0
            alpha = 128
            gradient.putpixel((x, 0), (r, g, b, alpha))
            for y in range(1, 30):
                gradient.putpixel((x, y), (r, g, b, alpha))

        # 将渐变色背景转换为Tkinter可识别的格式
        self.gradient_image = ImageTk.PhotoImage(gradient)
        self.countdown_label.config(image=self.gradient_image, compound=tk.CENTER)

    def show_tooltip(self, event, text):
        # 如果提示窗口已经存在，先销毁它
        if hasattr(self, 'tooltip_window') and self.tooltip_window:
            self.tooltip_window.destroy()

        # 创建新的提示窗口
        self.tooltip_window = tk.Toplevel(self.root)
        self.tooltip_window.overrideredirect(True)  # 去掉窗口边框和标题栏
        self.tooltip_window.geometry(f"+{event.x_root + 10}+{event.y_root + 10}")  # 设置窗口位置在鼠标附近
        self.tooltip_window.wm_attributes("-topmost", True)  # 确保窗口始终在最前面

        # 创建提示标签
        label = tk.Label(
            self.tooltip_window,
            text=text,
            background="#ffffe0",
            relief="solid",
            borderwidth=1,
            font=("Arial", 10),
            padx=5,
            pady=3
        )
        label.pack()

    def update_tooltip_position(self, event):
        # 更新提示窗口的位置
        if hasattr(self, 'tooltip_window') and self.tooltip_window:
            self.tooltip_window.geometry(f"+{event.x_root + 10}+{event.y_root + 10}")

    def hide_tooltip(self, event):
        # 隐藏提示窗口
        if hasattr(self, 'tooltip_window') and self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None

    def on_drag_enter(self, event):
        # 拖拽文件进入窗口时，显示视觉反馈（蓝色背景）
        self.root.configure(background='lightblue')
        # self.log(f"[提示]：支持拖入设备文件到此窗口加载设备")

    def on_drag_motion(self, event):
        # 拖拽文件在窗口内移动时，保持视觉反馈
        pass

    def on_drag_leave(self, event):
        # 拖拽文件离开窗口时，恢复原始背景色
        self.root.configure(background=self.original_bg_color)



    def create_status_light(self):
        # 创建一个Canvas用于绘制状态灯
        self.status_light_canvas = tk.Canvas(self.root, width=25, height=25, highlightthickness=0)
        self.status_light_canvas.grid(row=0, column=7, padx=5, pady=5, sticky="ne")  # 放置在右上角

        # 创建一个圆形（内圈）
        self.status_light_inner = self.status_light_canvas.create_oval(5, 5, 20, 20, fill='green', outline='')

        # 创建一个稍大的圆形（外圈）
        self.status_light_outer = self.status_light_canvas.create_oval(3, 3, 22, 22, fill='', outline='gray', width=1)

        # 启动闪烁效果
        self.flash_green_light()

    def flash_green_light(self):
        # 获取当前填充颜色
        current_color = self.status_light_canvas.itemcget(self.status_light_inner, 'fill')

        # 切换颜色（绿色和透明）
        if current_color == 'green':
            self.status_light_canvas.itemconfig(self.status_light_inner, fill='')
        else:
            self.status_light_canvas.itemconfig(self.status_light_inner, fill='green')

        # 使用 after 方法定时切换颜色
        self.root.after(1000, self.flash_green_light)

    def flash_red_light(self):
        # 切换到红色
        self.status_light_canvas.itemconfig(self.status_light_inner, fill='red')
        # 延迟后恢复绿色闪烁
        self.root.after(1000, self.flash_green_light)

    # 修改 log 方法，添加错误时闪烁红灯的逻辑
    def log(self, message):
        timestamp = datetime.now().strftime("%m/%d %H:%M:%S")  # 获取当前时间戳
        log_message = f"[{timestamp}] {message}\n"  # 格式化日志消息
        # self.log_file.write(log_message)  # 将日志消息写入日志文件
        # self.log_file.flush()  # 刷新日志文件缓冲区
        # print(log_message.rstrip())  # 打印日志消息到控制台

        if self.log_file:
            self.log_file.write(log_message)
            self.log_file.flush()
        print(log_message.rstrip())

        # 更新操作日志窗口
        self.root.after(1000, self.update_output_text, log_message.rstrip())

        # 更新实时日志窗口
        if hasattr(self,
                   'real_time_log_window') and self.real_time_log_window is not None and self.real_time_log_window.winfo_exists():
            self.root.after(1000, self.update_real_time_log_text, log_message.rstrip())

        # 如果是错误消息，闪烁红灯
        if "失败" in message or "错误" in message:
            self.flash_red_light()

    def update_output_text(self, message):
        self.output_text.configure(state='normal')  # 设置文本框为可编辑状态

        # 分析日志类型并设置颜色
        if "失败" in message or "错误" in message:  # 如果日志消息包含“失败”或“错误”
            self.output_text.insert(tk.END, message + '\n', "error")  # 插入红色日志消息
        elif "警告" in message:
            self.output_text.insert(tk.END, message + '\n', "warning")  # 插入橙色日志消息
        elif "成功" in message:
            self.output_text.insert(tk.END, message + '\n', "success")  # 插入绿色日志消息
        else:
            self.output_text.insert(tk.END, message + '\n', "normal")  # 插入黑色日志消息

        self.output_text.configure(state='disabled')  # 设置文本框为不可编辑状态
        self.output_text.yview(tk.END)  # 滚动到文本框底部

    def check_agreement(self):
        config_file = os.path.join(self.config_dir, "config.json")  # 配置文件路径
        show_agreement = True  # 是否显示用户协议

        if os.path.exists(config_file):  # 如果配置文件存在
            with open(config_file, "r") as f:
                config = json.load(f)  # 读取配置文件
                show_agreement = config.get("show_agreement", True)  # 获取是否显示协议的配置
        else:
            self.save_agreement(False)  # 如果配置文件不存在，保存默认配置

        if not show_agreement:  # 如果不显示协议，直接返回
            return

        self.agreement_window = tk.Toplevel(self.root)  # 创建协议窗口
        self.agreement_window.title("软件协议")  # 设置协议窗口标题
        self.agreement_window.geometry("500x280")  # 设置协议窗口大小
        self.center_window(self.agreement_window)  # 居中显示协议窗口
        self.agreement_window.transient(self.root)  # 设置协议窗口为根窗口的子窗口
        self.agreement_window.grab_set()  # 设置协议窗口为模态窗口

        agreement_label = tk.Label(self.agreement_window,
                                   text="此软件仅供测试，请勿在生产环境中使用本软件。\n否则由此产生的后果，本作者不承担任何责任。\n",
                                   font=("Arial", 13), justify="left", wraplength=450)  # 创建协议内容标签
        agreement_label.pack(pady=15)  # 将协议内容标签添加到协议窗口

        self.agreement_check_var = tk.IntVar()  # 创建协议复选框变量
        agreement_check = tk.Checkbutton(self.agreement_window,
                                         text="我已经阅读并同意以上协议内容，并承诺仅用于测试环境",
                                         variable=self.agreement_check_var)  # 创建协议复选框
        agreement_check.pack(pady=10)  # 将协议复选框添加到协议窗口

        accept_button = tk.Button(self.agreement_window, text="接受", command=self.accept_agreement)  # 创建接受按钮
        accept_button.pack(pady=10, side=tk.LEFT, padx=30)  # 将接受按钮添加到协议窗口

        deny_button = tk.Button(self.agreement_window, text="拒绝", command=self.deny_agreement)  # 创建拒绝按钮
        deny_button.pack(pady=10, side=tk.RIGHT, padx=30)  # 将拒绝按钮添加到协议窗口

        self.agreement_window.protocol("WM_DELETE_WINDOW", self.deny_agreement)  # 设置关闭协议窗口时的回调函数

    def save_agreement(self, show_agreement):
        config = {"show_agreement": show_agreement}  # 创建配置字典
        config_file = os.path.join(self.config_dir, "config.json")  # 配置文件路径
        with open(config_file, "w") as f:
            json.dump(config, f)  # 将配置保存到JSON文件


    def center_window(self, window):
        window.update_idletasks()  # 更新窗口布局
        width = window.winfo_width()  # 获取窗口宽度
        height = window.winfo_height()  # 获取窗口高度
        screen_width = window.winfo_screenwidth()  # 获取屏幕宽度
        screen_height = window.winfo_screenheight()  # 获取屏幕高度
        x = (screen_width - width) // 2  # 计算窗口居中时的X坐标
        y = (screen_height - height) // 2  # 计算窗口居中时的Y坐标
        window.geometry(f"+{x}+{y}")  # 设置窗口位置

    def accept_agreement(self):
        if self.agreement_check_var.get() == 1:  # 如果用户接受了协议
            self.save_agreement(False)  # 保存配置，不再显示协议
            self.agreement_window.destroy()  # 关闭协议窗口
        else:
            messagebox.showwarning("警告", "请先阅读并同意协议内容！")  # 提示用户未接受协议

    def deny_agreement(self):
        self.save_agreement(True)  # 保存配置，继续显示协议
        self.root.destroy()  # 关闭主窗口

    def log_audit(self, action, details=""):
        entry = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "action": action,
            "details": details,
            "user": os.getlogin()  # 获取当前用户名
        }
        self.audit_log.append(entry)

        # 确保 config 目录存在
        os.makedirs(self.config_dir, exist_ok=True)

        # 将审计日志保存到 config 目录
        audit_log_path = os.path.join(self.config_dir, "audit_log.json")
        with open(audit_log_path, "w") as f:
            json.dump(self.audit_log, f, indent=4)


    def view_audit_log(self):
        audit_window = tk.Toplevel(self.root)
        audit_window.title("操作审计日志")
        audit_window.geometry("1000x600")
        self.center_window(audit_window)

        # 创建一个包含表格和滚动条的框架
        log_frame = tk.Frame(audit_window)
        log_frame.pack(fill=tk.BOTH, expand=True)

        # 创建表格显示审计日志
        columns = ("timestamp", "action", "details")
        self.audit_tree = ttk.Treeview(log_frame, columns=columns, show='headings')
        for col in columns:
            self.audit_tree.heading(col, text=col)
            self.audit_tree.column(col, width=200)
        self.audit_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 添加垂直滚动条
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.audit_tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.audit_tree.configure(yscrollcommand=scrollbar.set)

        # 加载审计日志
        valid_entries = 0
        for entry in self.audit_log:
            if isinstance(entry, dict) and "timestamp" in entry and "action" in entry and "details" in entry:
                self.audit_tree.insert("", tk.END, values=(
                    entry["timestamp"],
                    entry["action"],
                    entry["details"]
                ))
                valid_entries += 1
            else:
                self.log("[警告]：已跳过一条格式不正确的审计日志记录")

        if valid_entries == 0:
            self.audit_tree.insert("", tk.END, values=("无有效日志", "无有效日志", "无有效日志"))

        # 添加按钮框架
        button_frame = tk.Frame(audit_window)
        button_frame.pack(fill=tk.X, pady=5)

        # 添加刷新和关闭按钮，并居中对齐
        refresh_button = tk.Button(button_frame, text="刷新", command=self.refresh_audit_log)
        close_button = tk.Button(button_frame, text="关闭", command=audit_window.destroy)

        refresh_button.pack(side=tk.LEFT, padx=5, pady=5)
        close_button.pack(side=tk.RIGHT, padx=5, pady=5)

        # 设置按钮框架的列权重，使按钮居中
        button_frame.columnconfigure(0, weight=1)
        button_frame.columnconfigure(1, weight=1)

    def load_audit_log_to_tree(self):
        # 清空表格
        for item in self.audit_tree.get_children():
            self.audit_tree.delete(item)

        # 加载审计日志
        for entry in self.audit_log:
            timestamp = entry.get("timestamp", "未知时间")
            user = entry.get("user", "未知用户")
            action = entry.get("action", "未知操作")
            details = entry.get("details", "无详情")
            self.audit_tree.insert("", tk.END, values=(timestamp, user, action, details))

    def refresh_audit_log(self):
        self.load_audit_log_to_tree()

    def init_log_file(self):
        # 确保日志文件所在的目录存在
        os.makedirs(os.path.dirname(self.log_file_path), exist_ok=True)
        # 确保日志文件正确关闭
        if hasattr(self, 'log_file') and self.log_file:
            self.log_file.close()

        # try:
        #     self.log_file = open(self.log_file_path, "a", encoding="utf-8")
        #     self.log("[提示]：日志文件已初始化")

        try:
            self.log_file = open(self.log_file_path, "a", encoding="utf-8")
            self.log_file.write("\n" + "=" * 50 + "\n")
            self.log_file.write(f"日志开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            self.log_file.write("=" * 50 + "\n")
            self.log_file.flush()


        except Exception as e:
            # 如果日志文件初始化失败，记录错误信息到控制台
            self.log(f"[错误]：日志文件初始化失败: {str(e)}")
            # 设置一个默认的日志文件对象，避免后续操作报错
            self.log_file = open(os.path.join(os.getcwd(), "操作日志.log"), "a", encoding="utf-8")

    def update_real_time_log_text(self, message):
        # 检查实时日志窗口是否存在
        if hasattr(self,
                   'real_time_log_window') and self.real_time_log_window is not None and self.real_time_log_window.winfo_exists():
            self.real_time_log_text.configure(state='normal')  # 设置文本框为可编辑状态

            # 分析日志类型并设置颜色
            if "失败" in message or "错误" in message:
                self.real_time_log_text.insert(tk.END, message + '\n', "error")
            elif "警告" in message:
                self.real_time_log_text.insert(tk.END, message + '\n', "warning")
            elif "成功" in message:
                self.real_time_log_text.insert(tk.END, message + '\n', "success")
            else:
                self.real_time_log_text.insert(tk.END, message + '\n', "normal")

            self.real_time_log_text.configure(state='disabled')  # 设置文本框为不可编辑状态
            self.real_time_log_text.yview(tk.END)  # 滚动到文本框底部
        else:
            # 如果实时日志窗口不存在，则不执行任何操作
            pass




    def sort_treeview_column(self, col):
        # 获取当前列的数据
        data = [(self.tree.set(item, col), item) for item in self.tree.get_children("")]

        # 检查是否是IP地址列
        if col == "hostname":
            # 按照IP地址格式进行排序
            data.sort(key=lambda x: self.parse_ip_for_sorting(x[0]), reverse=self.column_sort_direction[col])
        else:
            # 根据数据类型选择排序方式
            try:
                # 尝试将列数据转换为浮点数进行排序（适用于数值列）
                data.sort(key=lambda x: float(x[0]) if x[0].replace('.', '', 1).isdigit() else x[0],
                          reverse=self.column_sort_direction[col])
            except:
                # 按字符串进行排序（适用于文本列）
                data.sort(key=lambda x: x[0], reverse=self.column_sort_direction[col])

        # 重新排列行
        for index, (_, item) in enumerate(data):
            self.tree.move(item, '', index)

        # 切换排序方向
        self.column_sort_direction[col] = not self.column_sort_direction[col]

        # 设置标题样式以指示排序方向
        dir_str = " ↓" if self.column_sort_direction[col] else " ↑"
        self.tree.heading(col, text=f"{col}{dir_str}")

    def parse_ip_for_sorting(self, ip_str):
        # 将IP地址解析为一个元组，用于排序
        try:
            parts = list(map(int, ip_str.split('.')))
            if len(parts) != 4:
                raise ValueError
            return tuple(parts)
        except ValueError:
            # 如果不是有效的IP地址格式，返回一个很大的值，使其排在后面
            return (255, 255, 255, 255)

    def create_hover_effect(self, widget, original_bg, hover_bg):
        def on_enter(event):
            widget.config(bg=hover_bg)

        def on_leave(event):
            widget.config(bg=original_bg)

        widget.bind("<Enter>", on_enter)
        widget.bind("<Leave>", on_leave)



    def compare_command_results_diff(self):
        # 弹出文件选择对话框，让用户选择两个历史执行结果文件
        file_paths = filedialog.askopenfilenames(
            title="选择两个历史执行结果文件",
            filetypes=[("Excel files", "*.xlsx"), ("Text files", "*.txt"), ("HTML files", "*.html")]
        )

        if len(file_paths) != 2:
            messagebox.showinfo("提示", "请选择两个文件进行对比")
            return

        # 创建对比结果窗口
        compare_window = tk.Toplevel(self.root)
        compare_window.title("命令执行结果对比")
        compare_window.geometry("1200x800")
        self.center_window(compare_window)

        # 创建标签页控件
        notebook = ttk.Notebook(compare_window)
        notebook.pack(fill=tk.BOTH, expand=True)

        # 读取并显示第一个文件的内容
        frame1 = ttk.Frame(notebook)
        notebook.add(frame1, text="文件1: " + os.path.basename(file_paths[0]))
        self.display_file_content_with_commands_diff(frame1, file_paths[0])

        # 读取并显示第二个文件的内容
        frame2 = ttk.Frame(notebook)
        notebook.add(frame2, text="文件2: " + os.path.basename(file_paths[1]))
        self.display_file_content_with_commands_diff(frame2, file_paths[1])

        # 创建对比结果标签页
        result_frame = ttk.Frame(notebook)
        notebook.add(result_frame, text="对比结果")

        # 找出命令差异
        differences = self.find_command_differences_diff(file_paths[0], file_paths[1])

        # 显示详细差异
        self.display_detailed_differences_diff(result_frame, differences)

        # 添加差异统计
        self.add_difference_statistics_diff(result_frame, differences)

        # 添加导出按钮
        self.export_comparison_results_diff(result_frame, differences)

        # 添加关闭按钮
        tk.Button(compare_window, text="关闭", command=compare_window.destroy).pack(pady=5)

    def display_file_content_with_commands_diff(self, parent_frame, file_path):
        # 创建一个 ScrolledText 组件来显示文件内容
        text_widget = ScrolledText(parent_frame, wrap=tk.WORD)
        text_widget.pack(fill=tk.BOTH, expand=True)

        try:
            # 根据文件类型读取内容
            if file_path.endswith(".xlsx"):
                df = pd.read_excel(file_path)
                content = self.format_dataframe_content_diff(df)
            elif file_path.endswith(".txt"):
                with open(file_path, "r", encoding="utf-8") as f:
                    content = f.read()
            elif file_path.endswith(".html"):
                with open(file_path, "r", encoding="utf-8") as f:
                    content = f.read()
            else:
                content = "不支持的文件类型"

            text_widget.insert(tk.END, content)
            text_widget.configure(state="disabled")
        except Exception as e:
            text_widget.insert(tk.END, f"读取文件失败: {str(e)}")
            text_widget.configure(state="disabled")

    def format_dataframe_content_diff(self, df):
        content = ""
        for _, row in df.iterrows():
            # 确保命令和输出都是字符串类型
            command = str(row['命令']).strip()
            output = str(row['输出']).strip() if not pd.isna(row['输出']) else ""
            content += f"命令: {command}\n"
            content += f"输出:\n{output}\n\n"
        return content

    def find_command_differences_diff(self, file1_path, file2_path):
        differences = []

        try:
            # 读取两个文件的内容
            content1 = self.get_file_content_diff(file1_path)
            content2 = self.get_file_content_diff(file2_path)

            # 将内容按命令分割
            commands1 = self.split_content_by_command_diff(content1)
            commands2 = self.split_content_by_command_diff(content2)

            # 创建一个集合来跟踪已比较的命令
            compared_commands = set()

            # 比较两个文件中的命令
            for command in commands1:
                if command not in commands2:
                    differences.append({
                        "命令": command,
                        "文件1输出": commands1[command]["输出"],
                        "文件2输出": "命令不存在"
                    })
                    compared_commands.add(command)
                else:
                    # 比较命令输出
                    output1 = commands1[command]["输出"]
                    output2 = commands2[command]["输出"]

                    # 使用difflib进行智能对比
                    differ = difflib.SequenceMatcher(None, output1, output2)
                    if differ.ratio() < 0.95:  # 如果相似度低于95%，认为有差异
                        differences.append({
                            "命令": command,
                            "文件1输出": output1,
                            "文件2输出": output2
                        })
                    compared_commands.add(command)

            # 检查文件2中是否有文件1中不存在的命令
            for command in commands2:
                if command not in compared_commands:
                    differences.append({
                        "命令": command,
                        "文件1输出": "命令不存在",
                        "文件2输出": commands2[command]["输出"]
                    })

        except Exception as e:
            differences.append(f"对比失败: {str(e)}")

        return differences

    def on_drop_files(self, files):
        for file_path in files:
            # 尝试多种编码方式解码文件路径
            try:
                # 尝试使用utf-8解码
                file_path = file_path.decode('utf-8')
            except UnicodeDecodeError:
                try:
                    # 尝试使用系统默认编码（例如gbk）
                    file_path = file_path.decode('mbcs')
                except UnicodeDecodeError:
                    try:
                        # 尝试使用latin1编码（兼容性较强）
                        file_path = file_path.decode('latin1')
                    except UnicodeDecodeError:
                        self.log(f"[错误]：无法解码文件路径：{file_path}")
                        messagebox.showerror("错误", f"无法解码文件路径")
                        continue

            if file_path.endswith(('.xlsx', '.xls')):
                self.select_device_file(file_path)
            else:
                messagebox.showerror("错误", f"不支持的文件格式：{file_path}")
                self.log(f"[错误]：不支持的文件格式：{file_path}")

    def select_device_file(self, file_path=None):
        if file_path is None:
            file_path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx;*.xls")])
        if file_path:
            try:
                # 确保文件存在
                if not os.path.exists(file_path):
                    self.log(f"[错误]：文件不存在：{file_path}")
                    messagebox.showerror("错误", f"文件不存在：{file_path}")
                    return

                # 尝试读取Excel文件
                try:
                    df = pd.read_excel(file_path)
                    self.load_device_info_from_dataframe(df, file_path)
                    self.log(f"[成功]：成功加载设备信息文件：{file_path}")
                    self.auto_save_device_info()
                except Exception as e:
                    self.log(f"[错误]：读取Excel文件失败：{str(e)}")
                    messagebox.showerror("错误", f"读取Excel文件失败：{str(e)}")

            except Exception as e:
                self.log(f"[错误]：加载设备信息文件失败：{str(e)}")
                messagebox.showerror("错误", f"加载设备信息文件失败：{str(e)}")


    def validate_ip(self, ip_str):
        """验证IPv4地址格式"""
        ip_pattern = r"^((25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$"
        if not re.match(ip_pattern, ip_str):
            messagebox.showerror("IP格式错误",
                                 f"无效的IP地址：{ip_str}\n"
                                 "正确格式示例：192.168.1.1\n"
                                 "每个数字范围：0-255")
            self.log(f"[错误]：无效的IP地址：{ip_str}")
            return False
        return True


    def load_device_info_from_dataframe(self, df, file_path):
        try:
            required_columns = ["hostname", "username", "password", "protocol", "device_type", "path"]
            if not all(col in df.columns for col in required_columns):
                self.log(f"[错误]：Excel文件缺少必要的列：{required_columns}")
                messagebox.showerror("错误", f"Excel文件缺少必要的列：{required_columns}")
                return

            self.tree.delete(*self.tree.get_children())
            self.original_passwords = {}  # 清空原始密码字典
            self.super_passwords = {}  # 清空Super密码字典


            for _, row in df.iterrows():
                hostname = str(row["hostname"]).strip()
                # username = str(row["username"]).strip() if "username" in row else ""
                username = str(row["username"]).strip() if pd.notnull(row["username"]) else ""  # 如果为空，则保持为空
                password = str(row["password"]).strip() if "password" in row else ""
                # connection_protocol = str(row["protocol"]).strip() if "protocol" in row else None
                connection_protocol = str(row["protocol"]).strip().lower() if "protocol" in row else None
                device_type = str(row["device_type"]).strip() if "device_type" in row else None
                path = str(row["path"]).strip() if "path" in row else None
                # port = str(row["port"]).strip() if "port" in row else None
                # port = str(row["port"]).strip()
                port = row.get("port", None)  # 获取端口值，如果列不存在则为None
                super_password = str(row["super_password"]).strip() if "super_password" in row else ""  # 加载Super密码
                # print(f"从excel中加载Super密码{super_password}")
                # 检查端口是否为空或无效
                if pd.isna(port) or str(port).strip() == "" or port == "nan":
                    port = None  # 确保空值或特殊值被识别为None

                # 自动填充默认端口
                if port is None : # 端口为空时
                    # self.log("端口为空，进行默认填充")
                    if connection_protocol.lower() == "telnet":
                        port = 23
                        self.log(f"[提示]：设备 {hostname} Telnet端口为空，自动填充 23")
                    elif connection_protocol.lower() == "ssh":
                        port = 22
                        self.log(f"[提示]：设备 {hostname} SSH端口为空，自动填充 22")
                    else:
                        port = None  # 其他协议默认端口设置为None，允许后续处理
                else:
                    try:
                        port = int(float(port))  # 支持浮点数格式的端口
                    except (ValueError, TypeError):
                        self.log(f"[错误]：设备 {hostname}:{port} 端口无效")
                        continue  # 跳过当前设备，继续处理下一个设备

                # 校验IP地址格式
                if not self.validate_ip(hostname):
                    self.log(f"[错误]: 设备 {hostname} IP无效")
                    continue  # 跳过当前设备，继续处理下一个设备

                try:
                    if pd.notnull(port):
                        port = int(port)
                        if port < 1 or port > 65535:
                            raise ValueError
                    else:
                        port = ""
                except ValueError:
                    self.log(f"[错误]: 设备 {hostname}:{port} 端口无效")
                    continue

                # 插入到Treeview
                masked_password = '*' * len(password) if password else ""
                masked_super_password = '*' * len(super_password) if super_password else ""
                self.tree.insert("", "end", values=(
                    hostname,
                    username,
                    masked_password,
                    connection_protocol,
                    device_type,
                    path,
                    port if isinstance(port, int) else "",  # 确保插入的端口是整数或空字符串
                    "",
                    "",
                    masked_super_password # Super密码列
                ))

                # ## 更新原始密码和Super密码字典
                # if password:
                #     self.original_passwords[hostname] = password
                # if super_password:
                #     # print(f"Super密码加载{super_password}")
                #     self.super_passwords[hostname] = super_password  # 存储Super密码

                ## 更新原始密码和Super密码字典
                if password:
                    self.original_passwords[hostname] = password
                if super_password:  # 如果Super密码字段不为空字符串
                    self.super_passwords[hostname] = super_password if super_password != "nan" else ""  # 修正这里的逻辑判断
                else:
                    self.super_passwords[hostname] = ""  # 如果Super密码字段为空字符串，也设置为空

            # 显示成功消息（即使部分设备加载失败）
            loaded_count = len(self.tree.get_children())
            self.log(f"[成功]：设备信息加载成功, 共计{loaded_count}条信息")
            messagebox.showinfo("成功", f"设备信息加载成功, 共计{loaded_count}条信息")

        except Exception as e:
            self.log(f"[错误]：加载设备信息失败: {str(e)}")
            messagebox.showerror("错误", f"加载设备信息失败: {str(e)}")

    def validate_port(self, port_str):
        """验证端口号格式"""
        try:
            port = int(float(port_str))
            if 1 <= port <= 65535:
                return True
            else:
                messagebox.showerror("端口错误",
                                     f"无效端口：{port}\n"
                                     "有效范围：1-65535")
                self.log(f"[错误]：无效端口：{port}")
                return False
        except ValueError:
            messagebox.showerror("端口错误",
                                 f"非数字端口：{port_str}\n"
                                 "必须为整数")
            self.log(f"[错误]：非数字端口：{port_str}")
            return False

    def display_detailed_differences_diff(self, parent_frame, differences):
        # 创建一个框架来包含表格和滚动条
        container_frame = ttk.Frame(parent_frame)
        container_frame.pack(fill=tk.BOTH, expand=True)

        # 创建横向和纵向滚动条
        v_scrollbar = ttk.Scrollbar(container_frame, orient=tk.VERTICAL)
        h_scrollbar = ttk.Scrollbar(container_frame, orient=tk.HORIZONTAL)

        # 创建表格显示差异
        columns = ("命令", "行号", "文件1输出", "文件2输出")
        diff_tree = ttk.Treeview(
            container_frame,
            columns=columns,
            show='headings',
            yscrollcommand=v_scrollbar.set,
            xscrollcommand=h_scrollbar.set
        )

        # 配置滚动条
        v_scrollbar.config(command=diff_tree.yview)
        h_scrollbar.config(command=diff_tree.xview)

        # 布局
        diff_tree.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")

        # 设置行列权重，使元素分布更合理
        container_frame.grid_rowconfigure(0, weight=1)
        container_frame.grid_columnconfigure(0, weight=1)

        # 设置列标题和宽度
        for col in columns:
            diff_tree.heading(col, text=col)
            diff_tree.column(col, width=200, minwidth=100, stretch=tk.YES)

        # 调整行高
        diff_tree.configure(height=20)

        # 插入差异数据
        for diff in differences:
            if isinstance(diff, dict):
                command = diff["命令"]
                file1_output = diff["文件1输出"].split("\n")
                file2_output = diff["文件2输出"].split("\n")

                # 使用difflib比较输出内容
                differ = difflib.Differ()
                diff_lines = list(differ.compare(file1_output, file2_output))

                # 插入差异行
                for i, line in enumerate(diff_lines):
                    if line.startswith("- ") or line.startswith("+ "):
                        # 获取行号
                        line_number = i + 1  # 行号从1开始

                        # 提取内容
                        content = line[2:].strip()

                        # 确定是文件1还是文件2的差异
                        if line.startswith("- "):
                            file1_content = content
                            file2_content = ""
                        else:
                            file1_content = ""
                            file2_content = content

                        # 插入表格
                        diff_tree.insert("", tk.END, values=(
                            command,
                            line_number,
                            file1_content,
                            file2_content
                        ), tags=("diff",))
                    elif line.startswith("  "):
                        # 相同行，也插入表格以便查看
                        line_number = i + 1
                        content = line[2:].strip()
                        diff_tree.insert("", tk.END, values=(
                            command,
                            line_number,
                            content,
                            content
                        ), tags=("same",))

        # 设置差异行的高亮样式
        diff_tree.tag_configure("diff", background="#ffcccc")
        diff_tree.tag_configure("same", background="#e6ffe6")

        # 添加按钮框架
        button_frame = ttk.Frame(parent_frame)
        button_frame.pack(fill=tk.X, pady=5)

        # 定义按钮命令函数
        def show_diff_diff():
            for child in diff_tree.get_children():
                tags = diff_tree.item(child, "tags")
                if "diff" in tags:
                    diff_tree.item(child, values=diff_tree.item(child, "values"))
                else:
                    diff_tree.item(child, values=("", "", "", ""))

        def show_same_diff():
            for child in diff_tree.get_children():
                tags = diff_tree.item(child, "tags")
                if "same" in tags:
                    diff_tree.item(child, values=diff_tree.item(child, "values"))
                else:
                    diff_tree.item(child, values=("", "", "", ""))

        def reset_view_diff():
            for child in diff_tree.get_children():
                values = diff_tree.item(child, "values")
                diff_tree.item(child, values=values)

        # 添加一键显示不同按钮
        show_diff_diff_button = ttk.Button(button_frame, text="一键显示不同", command=show_diff_diff)
        show_diff_diff_button.pack(side=tk.LEFT, padx=5)

        # 添加一键显示相同按钮
        show_same_diff_button = ttk.Button(button_frame, text="一键显示相同", command=show_same_diff)
        show_same_diff_button.pack(side=tk.LEFT, padx=5)

        # 添加重置显示按钮
        reset_button = ttk.Button(button_frame, text="重置显示", command=reset_view_diff)
        reset_button.pack(side=tk.LEFT, padx=5)

    def add_difference_statistics_diff(self, parent_frame, differences):
        # 创建统计框架
        stats_frame = ttk.LabelFrame(parent_frame, text="差异统计")
        stats_frame.pack(fill=tk.X, padx=5, pady=5)

        # 统计差异类型
        command_only_in_file1 = 0
        command_only_in_file2 = 0
        output_difference = 0

        for diff in differences:
            if isinstance(diff, dict):
                if diff["文件1输出"] == "命令不存在":
                    command_only_in_file2 += 1
                elif diff["文件2输出"] == "命令不存在":
                    command_only_in_file1 += 1
                else:
                    output_difference += 1

        # 显示统计结果
        ttk.Label(stats_frame, text=f"仅在文件1中的命令: {command_only_in_file1}").grid(row=0, column=0, padx=5, pady=5)
        ttk.Label(stats_frame, text=f"仅在文件2中的命令: {command_only_in_file2}").grid(row=0, column=1, padx=5, pady=5)
        ttk.Label(stats_frame, text=f"输出不同的命令: {output_difference}").grid(row=1, column=0, padx=5, pady=5)

    def export_comparison_results_diff(self, parent_frame, differences):
        # 创建导出按钮
        export_button = ttk.Button(parent_frame, text="导出对比结果", command=lambda: self.export_differences_diff(differences))
        export_button.pack(pady=5)

    def export_differences_diff(self, differences):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            try:
                # 创建 DataFrame
                data = []
                for diff in differences:
                    if isinstance(diff, dict):
                        data.append([
                            diff.get("命令", ""),
                            diff.get("文件1输出", ""),
                            diff.get("文件2输出", ""),
                            diff.get("差异描述", "")
                        ])
                    else:
                        data.append([diff, "", "", ""])

                df = pd.DataFrame(data, columns=["命令", "文件1输出", "文件2输出", "差异描述"])
                df.to_excel(file_path, index=False)
                messagebox.showinfo("成功", "对比结果已导出")
                self.log(f"[成功]：对比结果已导出至: {file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"导出对比结果失败: {str(e)}")
                self.log(f"[错误]：导出对比结果失败: {str(e)}")

    def split_content_by_command_diff(self, content):
        commands = []
        lines = content.split("\n")
        current_command = None
        current_output = []

        for line in lines:
            if line.startswith("命令: "):
                if current_command is not None:
                    commands.append({
                        "命令": current_command,
                        "输出": "\n".join(current_output).strip()
                    })
                current_command = line.replace("命令: ", "").strip()
                current_output = []
            elif current_command is not None:
                current_output.append(line.strip())

        # 添加最后一个命令
        if current_command is not None:
            commands.append({
                "命令": current_command,
                "输出": "\n".join(current_output).strip()
            })

        # 将命令转换为字典，方便查找
        command_dict = {}
        for cmd in commands:
            command_dict[cmd["命令"]] = cmd

        return command_dict

    def get_file_content_diff(self, file_path):
        content = ""

        if file_path.endswith(".xlsx"):
            df = pd.read_excel(file_path)
            content = self.format_dataframe_content_diff(df)
        elif file_path.endswith(".txt"):
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()
        elif file_path.endswith(".html"):
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()

        return content

    class BreathingLight:
        def __init__(self, root):
            self.root = root
            self.canvas = tk.Canvas(root, width=100, height=100, highlightthickness=0)
            self.canvas.pack()
            self.outer_circle = self.canvas.create_oval(10, 10, 90, 90, outline='green', width=2)
            self.inner_circle = self.canvas.create_oval(20, 20, 80, 80, outline='green', width=2)
            self.current_alpha = 0.0
            self.direction = 1  # 1 for increasing, -1 for decreasing
            self.update_breathing_light()

        def update_breathing_light(self):
            # Update the alpha value
            self.current_alpha += 0.02 * self.direction
            if self.current_alpha >= 1.0:
                self.current_alpha = 1.0
                self.direction = -1
            elif self.current_alpha <= 0.0:
                self.current_alpha = 0.0
                self.direction = 1

            # Update the color with the current alpha
            alpha = int(self.current_alpha * 255)
            color = "#{:02x}{:02x}{:02x}".format(0, 255, 0)  # Green color
            self.canvas.itemconfig(self.outer_circle, outline=color)
            self.canvas.itemconfig(self.inner_circle, outline=color)
            self.canvas.itemconfig(self.outer_circle, stipple='gray%d' % (alpha))
            self.canvas.itemconfig(self.inner_circle, stipple='gray%d' % (alpha))

            # Schedule the next update
            self.root.after(50, self.update_breathing_light)

    def breathing_light(self):
        # 呼吸灯效果的颜色和透明度变化
        colors = ['green', 'yellow', 'red']
        alpha_values = [0.2, 0.4, 0.6, 0.8, 1.0, 0.8, 0.6, 0.4, 0.2]

        # 循环改变颜色和透明度
        for color in colors:
            for alpha in alpha_values:
                # 设置外圆颜色和透明度
                self.light_canvas.itemconfig(self.outer_circle, outline=color)
                self.light_canvas.itemconfig(self.outer_circle, fill=color)
                self.light_canvas.itemconfig(self.outer_circle, stipple='gray%d' % (alpha * 100))

                # 设置内圆颜色和透明度
                self.light_canvas.itemconfig(self.inner_circle, outline=color)
                self.light_canvas.itemconfig(self.inner_circle, fill=color)
                self.light_canvas.itemconfig(self.inner_circle, stipple='gray%d' % (alpha * 100))

                # 更新UI
                self.root.update()
                time.sleep(0.1)

        # 递归调用自身，持续显示呼吸灯效果
        self.root.after(100, self.breathing_light)

    def create_real_time_log_menu(self):
        # 在菜单栏中添加“实时日志”菜单
        real_time_log_menu = tk.Menu(self.menu_bar, tearoff=0)
        real_time_log_menu.add_command(label="打开实时日志独立大窗口", command=self.toggle_real_time_log_window)
        self.menu_bar.add_cascade(label="实时日志窗口", menu=real_time_log_menu)
        # self.log_audit("[操作]：实时日志窗口", "用户已点击打开实时日志窗口")
        # self.log_audit("[操作]：执行 ping 测试", "用户点击 ping 测试按钮")

    # 切换实时日志窗口
    def toggle_real_time_log_window(self):
        if not hasattr(self,
                       'real_time_log_window') or self.real_time_log_window is None or not self.real_time_log_window.winfo_exists():
            self.create_real_time_log_window()
            self.log(f"[提示]：已打开实时日志大窗口")
            self.log_audit("[操作]：实时日志窗口", "用户已打开实时日志窗口")
        else:
            self.real_time_log_window.destroy()
            self.real_time_log_window = None
            self.log(f"[提示]：已关闭实时日志大窗口")
            self.log_audit("[操作]：实时日志窗口", "用户已关闭实时日志窗口")

    # 创建实时日志窗口
    def create_real_time_log_window(self):
        # 创建一个可调整大小的窗口
        self.real_time_log_window = tk.Toplevel(self.root)
        self.real_time_log_window.title("实时日志")
        self.real_time_log_window.geometry("800x600")
        self.center_window(self.real_time_log_window)
        # self.log(f"[提示]：用户点击实时日志大窗口")
        self.log_audit("[操作]：实时日志窗口", "用户已点击打开实时日志窗口")

        # 创建筛选框架
        filter_frame = tk.Frame(self.real_time_log_window)
        filter_frame.pack(fill=tk.X, padx=5, pady=5)

        # 日志级别筛选
        tk.Label(filter_frame, text="日志级别:").pack(side=tk.LEFT, padx=5)
        self.log_level_var = tk.StringVar(value="所有")
        log_level_menu = ttk.Combobox(filter_frame, textvariable=self.log_level_var,
                                      values=["所有", "正常", "警告", "错误", "提示", "成功", "失败"], width=10)
        log_level_menu.pack(side=tk.LEFT, padx=5)
        log_level_menu.bind("<<ComboboxSelected>>", self.filter_real_time_logs)

        # 设备信息筛选
        tk.Label(filter_frame, text="设备信息:").pack(side=tk.LEFT, padx=5)
        self.device_var = tk.StringVar(value="所有")
        self.device_menu = ttk.Combobox(filter_frame, textvariable=self.device_var, values=["所有"], width=15)
        self.device_menu.pack(side=tk.LEFT, padx=5)
        self.device_menu.bind("<<ComboboxSelected>>", self.filter_real_time_logs)
        self.update_device_menu()

        # 关键字搜索
        tk.Label(filter_frame, text="关键字搜索:").pack(side=tk.LEFT, padx=5)
        self.search_var = tk.StringVar()
        search_entry = tk.Entry(filter_frame, textvariable=self.search_var, width=15)
        search_entry.pack(side=tk.LEFT, padx=5)
        search_entry.bind("<KeyRelease>", self.filter_real_time_logs)

        # 创建一个带滚动条的文本框
        self.real_time_log_text = ScrolledText(self.real_time_log_window, height=40, width=60)
        self.real_time_log_text.pack(fill=tk.BOTH, expand=True)

        # 配置文本框为只读
        self.real_time_log_text.configure(state='disabled')

        # 定义标签样式
        self.real_time_log_text.tag_config("normal", foreground="black")
        self.real_time_log_text.tag_config("warning", foreground="darkorange")
        self.real_time_log_text.tag_config("error", foreground="darkred")

        # 添加筛选按钮
        filter_button = tk.Button(filter_frame, text="筛选", command=self.filter_real_time_logs)
        filter_button.pack(side=tk.LEFT, padx=5)

        reset_filter_button = tk.Button(filter_frame, text="重置筛选", command=self.reset_real_time_log_filters)
        reset_filter_button.pack(side=tk.LEFT, padx=5)

    # 处理调整大小的事件
    def on_resize(self, event):
        # 调整实时日志窗口的大小
        self.real_time_log_text.configure(width=event.width)

    # 修改 log 方法，将日志同时写入实时日志窗口

    def update_device_menu(self):
        # 更新设备信息下拉菜单
        devices = set()
        for item_id in self.tree.get_children():
            values = self.tree.item(item_id, "values")
            hostname = values[0]
            devices.add(hostname)
        devices = sorted(devices)
        devices.insert(0, "所有")
        self.device_menu['values'] = devices

    def filter_real_time_logs(self, event=None):
        # 获取筛选条件
        log_level = self.log_level_var.get()
        device = self.device_var.get()
        search_keyword = self.search_var.get()

        # 清空当前显示
        self.real_time_log_text.configure(state='normal')
        self.real_time_log_text.delete('1.0', tk.END)

        # 遍历日志文件并筛选
        with open(self.log_file_path, 'r', encoding='utf-8') as file:
            for line in file:
                # 检查日志级别
                if log_level != "所有" and f"[{log_level}]" not in line:
                    continue
                # 检查设备信息
                if device != "所有" and device not in line:
                    continue
                # 检查关键字
                if search_keyword and search_keyword not in line:
                    continue

                # 插入日志并设置颜色
                if "[失败]" in line or "[错误]" in line:
                    self.real_time_log_text.insert(tk.END, line, "error")
                elif "[警告]" in line:
                    self.real_time_log_text.insert(tk.END, line, "warning")
                elif "[提示]" in line:
                    self.real_time_log_text.insert(tk.END, line, "normal")
                elif "[成功]" in line:
                    self.real_time_log_text.insert(tk.END, line, "normal")
                else:
                    self.real_time_log_text.insert(tk.END, line, "normal")

        self.real_time_log_text.configure(state='disabled')
        self.real_time_log_text.yview(tk.END)

    def reset_real_time_log_filters(self):
        # 重置筛选条件
        self.log_level_var.set("所有")
        self.device_var.set("所有")
        self.search_var.set("")
        self.filter_real_time_logs()

    def show_clean_all_dialog(self):
        # 创建对话框
        clean_dialog = tk.Toplevel(self.root)
        clean_dialog.title("一键清理")
        clean_dialog.geometry("400x350")
        self.center_window(clean_dialog)

        # 创建变量
        self.clean_vars = {
            "操作日志": tk.BooleanVar(value=True),
            "审计日志": tk.BooleanVar(value=True),
            "设备执行结果": tk.BooleanVar(value=True),
            "命令模板": tk.BooleanVar(value=True),
            "配置文件": tk.BooleanVar(value=True),
            # "设备信息": tk.BooleanVar(value=True),
            "其他文件": tk.BooleanVar(value=True)
        }

        # 创建复选框
        for i, (name, var) in enumerate(self.clean_vars.items()):
            tk.Checkbutton(clean_dialog, text=name, variable=var).pack(anchor="w", padx=20, pady=5)

        # 创建按钮框架
        button_frame = tk.Frame(clean_dialog)
        button_frame.pack(fill=tk.X, padx=20, pady=20)

        # 创建按钮
        tk.Button(button_frame, text="确认清理", command=lambda: self.clean_all(clean_dialog)).pack(side=tk.LEFT,
                                                                                                    padx=5)
        tk.Button(button_frame, text="取消", command=clean_dialog.destroy).pack(side=tk.RIGHT, padx=5)

    def clean_all(self, dialog):
        selected = {name: var.get() for name, var in self.clean_vars.items()}
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # 要写入的内容
        content_to_write = f"时间 - 清除日志（{current_time}）\n"
        if not any(selected.values()):
            messagebox.showinfo("提示", "请至少选择一项要清理的内容")
            return

        try:
            if selected["设备执行结果"]:
                device_log_dir = os.path.join(os.getcwd(), "device-log")
                print(f"设备执行结果目录路径: {device_log_dir}")  # 打印路径
                if os.path.exists(device_log_dir):
                    # 尝试删除目录中的所有文件和子目录
                    for root, dirs, files in os.walk(device_log_dir, topdown=False):
                        for name in files:
                            file_path = os.path.join(root, name)
                            try:
                                # 尝试关闭文件句柄
                                os.close(os.open(file_path, os.O_RDONLY))
                            except Exception as e:
                                self.log(f"[警告]：关闭文件句柄失败: {str(e)}")
                            try:
                                os.remove(file_path)
                                self.log(f"[成功]：已删除文件: {file_path}")
                            except PermissionError:
                                self.log(f"[警告]：文件 {file_path} 正在被使用，无法删除")
                            except Exception as e:
                                self.log(f"[错误]：删除文件 {file_path} 失败: {str(e)}")
                        for name in dirs:
                            dir_path = os.path.join(root, name)
                            try:
                                os.rmdir(dir_path)
                                self.log(f"[成功]：已删除目录: {dir_path}")
                            except Exception as e:
                                self.log(f"[错误]：删除目录 {dir_path} 失败: {str(e)}")
                    # 删除空的device-log目录
                    try:
                        os.rmdir(device_log_dir)
                        self.log(f"[成功]：已删除目录: {device_log_dir}")
                    except Exception as e:
                        self.log(f"[错误]：删除目录 {device_log_dir} 失败: {str(e)}")

                    # 清理操作日志
                    # 清理操作日志
            if selected["操作日志"]:
                        try:
                            # 清空的日志文件路径
                            # log_file_path = "your_log_file.log"

                            # 获取当前日期和时间
                            # current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                            # 要写入的内容
                            content_to_write = f"时间 - 清除日志（{current_time}）\n"
                            log_file_path = os.path.join(self.config_dir, "操作日志.log")
                            if hasattr(self, 'log_file') and not self.log_file.closed:
                                if os.path.exists(log_file_path):
                                    with open(log_file_path, "w", encoding="utf-8") as log_file:
                                        log_file.write(content_to_write)
                                        self.log(f"[成功]: 操作日志已清理")
                                else:
                                    self.log("[失败]：操作日志文件不存在")
                            else:
                                self.log("[失败]：操作日志文件不存在")
                        except Exception as e:
                            self.log(f"[错误]：清理操作日志失败: {str(e)}")

            if selected["审计日志"]:
                audit_log_path = os.path.join(self.config_dir, "audit_log.json")
                if os.path.exists(audit_log_path):
                    try:
                        with open(audit_log_path, "w") as f:
                            json.dump([], f)
                        self.log("[成功]：审计日志已清空")
                    except Exception as e:
                        self.log(f"[错误]：清空审计日志失败: {str(e)}")

            if selected["命令模板"]:
                template_path = os.path.join(self.config_dir, "command_templates.json")
                if os.path.exists(template_path):
                    try:
                        with open(template_path, "w") as f:
                            json.dump({}, f)
                        self.log("[成功]：命令模板已清空")
                    except Exception as e:
                        self.log(f"[错误]：清空命令模板失败: {str(e)}")

            # 清理其他文件
            if selected["其他文件"]:
                    try:
                        config_dir = os.path.join(os.getcwd(), "config")
                        if os.path.exists(config_dir):
                            for filename in os.listdir(config_dir):
                                if filename not in ["command_templates.json", "Key.json", "prompt_patterns.json",
                                                    "tasks.json","audit_log.json", "config.json"]:
                                    file_path = os.path.join(config_dir, filename)
                                    try:
                                        os.remove(file_path)
                                    except PermissionError:
                                        self.log(f"[警告]：文件 {file_path} 正在被使用，无法删除")
                                    except Exception as e:
                                        self.log(f"[错误]：删除文件 {file_path} 时出错: {str(e)}")
                            self.log("[成功]：其他文件已清理")
                        else:
                            self.log("[提示]：配置目录不存在")
                    except Exception as e:
                        self.log(f"[错误]：清理其他文件失败: {str(e)}")

            if selected["配置文件"]:
                config_files = {
                    "Key.json": {},
                    "prompt_patterns.json": {},
                    "tasks.json": [],
                    "device.json": [],
                    "config.json": {}
                }
                for filename, default_content in config_files.items():
                    config_path = os.path.join(self.config_dir, filename)
                    if os.path.exists(config_path):
                        try:
                            with open(config_path, "w") as f:
                                if isinstance(default_content, dict):
                                    json.dump(default_content, f)
                                elif isinstance(default_content, list):
                                    json.dump(default_content, f)
                            # self.load_default_device_info()
                            self.log(f"[成功]：文件 {filename} 已重置")
                        except Exception as e:
                            self.log(f"[错误]：重置文件 {filename} 失败: {str(e)}")

            messagebox.showinfo("清理完成", "所选数据已清理完成")
            dialog.destroy()

        except Exception as e:
            self.log(f"[错误]：一键清理失败: {str(e)}")
            messagebox.showerror("清理失败", f"一键清理失败: {str(e)}")
            dialog.destroy()


    def clear_logs(self):
        # 弹出确认对话框
        confirm = messagebox.askyesno("确认清理", "确定要清理日志吗？", default=messagebox.NO)
        if not confirm:
            return

        # 清理操作日志
        try:
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # 要写入的内容
            content_to_write = f"时间 - 清除日志（{current_time}）\n"
            log_file_path = os.path.join(self.config_dir, "操作日志.log")
            if hasattr(self, 'log_file') and not self.log_file.closed:
                if os.path.exists(log_file_path):
                    with open(log_file_path, "w",encoding="utf-8") as log_file:
                        log_file.write(content_to_write)
                        self.log(f"[成功]: 操作日志已清理")

                else:
                    self.log("[失败]：操作日志文件不存在")
            else:
                self.log("[失败]：操作日志文件不存在")
        except Exception as e:
            self.log(f"[错误]：清理操作日志失败: {str(e)}")

        # 清理审计日志
        try:
            audit_log_path = os.path.join(self.config_dir, "audit_log.json")
            if os.path.exists(audit_log_path):
                with open(audit_log_path, "w") as f:
                    json.dump([], f)
                self.log("[成功]：审计日志已清理")
        except Exception as e:
            self.log(f"[错误]：清理审计日志失败: {str(e)}")

    def export_error_report(self):
        # 收集所有执行失败的命令信息
        self.log_audit("[操作]：导出报告", "用户点击导出错误报告按钮")
        error_data = []
        for item_id in self.tree.get_children():
            values = self.tree.item(item_id, "values")
            hostname = values[0]
            execute_result = values[8] if len(values) > 8 else ""
            if "[失败]" in execute_result or "[错误]" in execute_result:
                error_command = self.last_error_command.get(hostname, "")
                error_output = self.last_error_output.get(hostname, "")
                error_data.append({
                    "设备IP": hostname,
                    "错误命令": error_command,
                    "错误输出": error_output,
                    "错误提示": execute_result
                })

        if not error_data:
            messagebox.showinfo("提示", "没有发现执行错误的命令")
            return

        # 创建Excel报告
        excel_report_path = os.path.join(self.output_directory, "error_report.xlsx")
        df = pd.DataFrame(error_data)
        df.to_excel(excel_report_path, index=False)

        # 创建HTML报告
        html_report_path = os.path.join(self.output_directory, "error_report.html")
        html_content = """
        <html>
        <head>
            <title>命令执行错误报告</title>
            <style>
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
            </style>
        </head>
        <body>
            <h1>命令执行错误报告</h1>
            <table>
                <tr>
                    <th>设备IP</th>
                    <th>错误命令</th>
                    <th>错误输出</th>
                    <th>错误提示</th>
                </tr>
                {}
            </table>
        </body>
        </html>
        """.format("\n".join(
            [
                f"<tr><td>{row['设备IP']}</td><td>{row['错误命令']}</td><td>{row['错误输出']}</td><td>{row['错误提示']}</td></tr>"
                for _, row in
                df.iterrows()]))

        with open(html_report_path, "w", encoding="utf-8") as f:
            f.write(html_content)

        # 创建ZIP文件并添加报告
        zip_file_path = filedialog.asksaveasfilename(defaultextension=".zip", filetypes=[("Zip files", "*.zip")])
        if zip_file_path:
            try:
                with zipfile.ZipFile(zip_file_path, 'w') as zipf:
                    zipf.write(excel_report_path, os.path.basename(excel_report_path))
                    zipf.write(html_report_path, os.path.basename(html_report_path))
                messagebox.showinfo("成功", "错误报告已导出为ZIP文件")
                self.log(f"[成功]：错误报告已导出至: {zip_file_path}")
                self.log_audit("[操作]：导出报告", "用户导出错误报告成功")
            except Exception as e:
                messagebox.showerror("错误", f"导出错误报告失败: {str(e)}")
                self.log(f"[错误]：导出错误报告失败: {str(e)}")
                self.log_audit("[操作]：导出报告", "用户导出错误报告失败")

        # 删除临时文件
        if os.path.exists(excel_report_path):
            os.remove(excel_report_path)
        if os.path.exists(html_report_path):
            os.remove(html_report_path)

    def export_script_report(self):
        report_data = []
        columns = ["hostname", "username", "port", "execute_result"]  # 定义报告数据的列
        self.log_audit("[操作]：导出报告", "用户点击导出脚本执行报告按钮")

        # 遍历 Treeview 获取数据
        for item_id in self.tree.get_children():
            values = self.tree.item(item_id, "values")  # 获取Treeview中每一行的数据
            hostname = values[0]  # 获取主机名
            username = values[1]  # 获取用户名
            port = values[6] if len(values) > 6 else ""  # 获取端口号
            execute_result = values[8] if len(values) > 8 else ""  # 获取执行结果
            report_data.append([hostname, username, port, execute_result])  # 将数据添加到报告数据列表

        # 保存为 Excel 文件
        excel_report_path = os.path.join(self.output_directory, "script_execution_report.xlsx")
        df = pd.DataFrame(report_data, columns=columns)
        df.to_excel(excel_report_path, index=False)

        # 新增HTML报告生成逻辑
        html_report_path = os.path.join(self.output_directory, "script_execution_report.html")
        html_content = """
        <html>
        <head>
            <title>脚本执行报告</title>
            <style>
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
            </style>
        </head>
        <body>
            <h1>脚本执行报告</h1>
            <table>
                <tr>
                    <th>设备IP</th>
                    <th>用户名</th>
                    <th>端口</th>
                    <th>执行结果</th>
                </tr>
                {}
            </table>
        </body>
        </html>
        """.format("\n".join(
            [
                f"<tr><td>{row['hostname']}</td><td>{row['username']}</td><td>{row['port']}</td><td>{row['execute_result']}</td></tr>"
                for _, row in
                df.iterrows()]))

        with open(html_report_path, "w", encoding="utf-8") as f:
            f.write(html_content)

        # 创建ZIP文件并添加报告
        zip_file_path = filedialog.asksaveasfilename(defaultextension=".zip", filetypes=[("Zip files", "*.zip")])
        if zip_file_path:
            try:
                with zipfile.ZipFile(zip_file_path, 'w') as zipf:
                    zipf.write(excel_report_path, os.path.basename(excel_report_path))
                    zipf.write(html_report_path, os.path.basename(html_report_path))
                messagebox.showinfo("成功", "脚本执行报告已导出为ZIP文件")
                self.log(f"[成功]：脚本执行报告已导出至: {zip_file_path}")
                self.log_audit("[操作]：导出报告", "用户导出脚本执行报告成功")
            except Exception as e:
                messagebox.showerror("错误", f"导出脚本执行报告失败: {str(e)}")
                self.log(f"[错误]：导出脚本执行报告失败: {str(e)}")
                self.log_audit("[操作]：导出报告", "用户导出脚本执行报告失败")

        # 删除临时文件
        if os.path.exists(excel_report_path):
            os.remove(excel_report_path)
        if os.path.exists(html_report_path):
            os.remove(html_report_path)

    def toggle_password_visibility(self, show_password):
        # 遍历 Treeview 的每一行
        for item_id in self.tree.get_children():
            values = list(self.tree.item(item_id, "values"))  # 将元组转换为列表
            hostname = values[0]  # 获取主机名
            # 获取原始密码
            if hostname in self.original_passwords:
                original_password = self.original_passwords[hostname]  # 获取原始密码
                original_super_password = self.super_passwords.get(hostname, "")
            else:
                original_password = values[2]  # 获取密码
                original_super_password =  values[9]

            # 更新密码显示
            if show_password:
                # 显示密码
                values[2] = original_password  # 设置密码为原始密码
                values[9] = original_super_password

            else:
                # 隐藏密码
                masked_password = '*' * len(original_password)  # 生成掩码密码
                masked_super_password = '*' * len(original_super_password)
                values[2] = masked_password  # 设置密码为掩码密码
                values[9] = masked_super_password

            self.tree.item(item_id, values=values)  # 更新Treeview中的值
        self.password_visible = show_password  # 更新密码可见状态

    def create_menu_bar(self):
        self.menu_bar = tk.Menu(self.root)  # 创建菜单栏
        self.root.config(menu=self.menu_bar)  # 将菜单栏添加到主窗口

        # 创建文件菜单
        file_menu = tk.Menu(self.menu_bar, tearoff=0)
        file_menu.add_command(label="选择设备文件", command=self.select_device_file)
        file_menu.add_command(label="下载模板", command=self.download_template)
        file_menu.add_command(label="选择保存目录", command=self.select_output_directory)
        file_menu.add_command(label="选择命令脚本", command=self.select_command_file)
        file_menu.add_command(label="导出日志", command=self.export_log)
        file_menu.add_command(label="导出脚本执行结果ZIP", command=self.export_command_result_zip)
        file_menu.add_command(label="导出执行错误报告ZIP", command=self.export_error_report)
        file_menu.add_command(label="导出脚本执行报告", command=self.export_script_report)
        file_menu.add_separator()
        file_menu.add_command(label="导出选中设备", command=self.export_selected_devices)
        file_menu.add_command(label="导入设备信息", command=self.import_devices)
        file_menu.add_separator()
        file_menu.add_command(label="复制", command=self.copy_selected)
        file_menu.add_command(label="粘贴", command=self.paste_selected)
        file_menu.add_command(label="剪切", command=self.cut_selected)
        file_menu.add_command(label="删除", command=self.delete_selected)
        file_menu.add_separator()
        file_menu.add_command(label="隐藏密码", command=lambda: self.toggle_password_visibility(False))
        file_menu.add_command(label="显示密码", command=lambda: self.toggle_password_visibility(True))
        file_menu.add_separator()
        file_menu.add_command(label="设置超时时间", command=lambda: self.timeout_entry.focus())
        file_menu.add_command(label="设置命令执行间隔", command=lambda: self.cmd_interval_entry.focus())
        # file_menu.add_command(label="设置检测次数", command=lambda: self.command=self.set_retry_count())
        file_menu.add_command(label="实时日志窗口", command=self.toggle_real_time_log_window)
        file_menu.add_separator()
        file_menu.add_command(label="退出", command=self.root.quit)

        # 创建设备测试菜单
        device_test_menu = tk.Menu(self.menu_bar, tearoff=0)
        device_test_menu.add_command(label="Ping-持续", command=self.start_continuous_ping)
        device_test_menu.add_command(label="Ping-指定N次", command=self.start_specified_ping)
        device_test_menu.add_command(label="Ping-1次", command=self.start_ping_test)
        device_test_menu.add_command(label="Login 测试", command=self.start_test_connections)
        # devcie_test.add_command(label="执行脚本", command=self.start_command_execution)
        # 添加停止Ping选项
        device_test_menu.add_command(label="停止Ping进程", command=self.stop_ping)

        # 创建设置菜单
        settings_menu = tk.Menu(self.menu_bar, tearoff=0)
        settings_menu.add_command(label="隐藏密码", command=lambda: self.toggle_password_visibility(False))
        settings_menu.add_command(label="显示密码", command=lambda: self.toggle_password_visibility(True))
        settings_menu.add_separator()
        settings_menu.add_command(label="Ping超时", command=lambda: self.timeout_entry.focus())
        settings_menu.add_command(label="命令间隔", command=lambda: self.cmd_interval_entry.focus())
        settings_menu.add_command(label="命令执行等待超时", command=self.set_retry_count)  # 新增重试次数设置
        # 在设置菜单中添加管理提示符选项
        settings_menu.add_command(label="管理输入提示符", command=self.manage_prompt_patterns)
        settings_menu.add_command(label="管理执行结果错误判断字符", command=self.manage_error_chars)  # 新增错误字符管理
        settings_menu.add_command(label="管理Super密码检测字符",
                                  command=self.manage_super_password_chars)  # 新增Super密码检测字符管理
        settings_menu.add_command(label="切换命令执行模式", command=self.show_command_mode_dialog)  # 新增模式选择

        # 创建帮助菜单
        help_menu = tk.Menu(self.menu_bar, tearoff=0)
        help_menu.add_command(label="关于", command=self.show_about)
        help_menu.add_command(label="使用帮助", command=self.show_help)
        help_menu.add_command(label="下载用户手册", command=self.download_user_manual)
        help_menu.add_command(label="用户反馈", command=self.show_feedback)  # 添加用户反馈菜单项

        # 创建审计和历史记录菜单
        audit_menu = tk.Menu(self.menu_bar, tearoff=0)
        audit_menu.add_command(label="命令执行审计", command=self.view_command_history)
        audit_menu.add_command(label="操作日志审计", command=self.view_audit_log)

        # 创建测试类型切换菜单
        test_type_menu = tk.Menu(self.menu_bar, tearoff=0)
        test_type_menu.add_radiobutton(label="ICMP 测试", variable=self.ping_type_var, value="ICMP")
        test_type_menu.add_radiobutton(label="TCP 测试", variable=self.ping_type_var, value="TCP")
        test_type_menu.add_radiobutton(label="UDP 测试", variable=self.ping_type_var, value="UDP")
        test_type_menu.add_radiobutton(label="SNMP 测试", variable=self.ping_type_var, value="SNMP")

        # 将菜单添加到菜单栏
        self.menu_bar.add_cascade(label="文件", menu=file_menu)
        self.menu_bar.add_cascade(label="设备测试", menu=device_test_menu)  # 新增设备测试菜单
        self.menu_bar.add_cascade(label="测试切换", menu=test_type_menu)
        self.menu_bar.add_cascade(label="参数管理", menu=settings_menu)
        self.menu_bar.add_cascade(label="日志审计", menu=audit_menu)
        # self.menu_bar.add_cascade(label="帮助", menu=help_menu)

        # 创建结果对比菜单
        compare_menu = tk.Menu(self.menu_bar, tearoff=0)
        compare_menu.add_command(label="文件对比", command=self.compare_command_results_diff)
        compare_menu.add_command(label="文本对比增强", command=self.open_compare_window)
        self.menu_bar.add_cascade(label="对比", menu=compare_menu)

        # 创建定时任务管理菜单
        schedule_menu = tk.Menu(self.menu_bar, tearoff=0)
        schedule_menu.add_command(label="添加定时任务", command=self.add_task)
        schedule_menu.add_command(label="删除定时任务", command=self.remove_task)
        schedule_menu.add_command(label="暂停定时任务", command=self.pause_task)
        schedule_menu.add_command(label="恢复定时任务", command=self.resume_task)
        self.menu_bar.add_cascade(label="任务管理", menu=schedule_menu)

        self.menu_bar.add_cascade(label="帮助和反馈", menu=help_menu)

        device_group_menu = tk.Menu(self.menu_bar, tearoff=0)
        device_group_menu.add_command(label="创建新局点", command=self.create_group)
        device_group_menu.add_command(label="删除当前局点", command=self.delete_group)
        # device_group_menu.add_command(label="切换局点", command=self.switch_group)
        self.menu_bar.add_cascade(label="局点管理", menu=device_group_menu)



        # # 绑定全局快捷键
        self.root.bind("<Control-o>", lambda event: self.select_device_file())
        self.root.bind("<Control-s>", lambda event: self.export_log())
        self.root.bind("<Control-q>", lambda event: self.root.quit())
        # self.root.bind("<Control-c>", lambda event: self.copy_selected())
        # self.root.bind("<Control-v>", lambda event: self.paste_selected())
        # self.root.bind("<Control-x>", lambda event: self.cut_selected())
        self.root.bind("<Delete>", lambda event: self.delete_selected())
        self.root.bind("<Control-a>", lambda event: self.select_all())
        # self.root.bind("<Control-z>", lambda event: self.undo())
        # self.root.bind("<Control-y>", lambda event: self.redo())


    def create_right_click_menu(self):
        # 创建右键菜单
        self.right_click_menu = tk.Menu(self.root, tearoff=0)
        # 设备操作
        # 直接在右键菜单中添加常用功能
        self.right_click_menu.add_command(label="全选", command=self.select_all)
        self.right_click_menu.add_command(label="编辑", command=self.edit_selected_devices)
        self.right_click_menu.add_command(label="全选", command=self.edit_selected_devices)
        self.right_click_menu.add_command(label="复制", command=self.copy_selected)
        self.right_click_menu.add_command(label="粘贴", command=self.paste_selected)
        self.right_click_menu.add_command(label="剪切", command=self.cut_selected)
        self.right_click_menu.add_command(label="删除", command=self.delete_selected)

        self.right_click_menu.add_separator()
        # 在 create_right_click_menu 方法中调用
        self.add_preview_script_to_context_menu()
        self.right_click_menu.add_command(label="导入设备信息", command=self.import_devices)
        self.right_click_menu.add_command(label="导出选中设备", command=self.export_selected_devices)

        self.right_click_menu.add_separator()

        self.right_click_menu.add_command(label="连接设备", command=self.connect_to_device)
        self.right_click_menu.add_command(label="停止Ping", command=self.stop_ping)

        self.right_click_menu.add_command(label="TcpPing 测试", command=self.tcp_ping_test)
        self.right_click_menu.add_command(label="Login 测试", command=self.start_test_connections)
        self.right_click_menu.add_command(label="Ping 测试", command=self.start_ping_test)

        self.right_click_menu.add_separator()
        # 执行脚本命令
        self.right_click_menu.add_command(label="执行脚本", command=self.start_command_execution)

        # 设备文件操作
        self.right_click_menu.add_command(label="导入设备", command=self.select_device_file)
        self.right_click_menu.add_command(label="下载模板", command=self.download_template)

        # 命令脚本操作
        self.right_click_menu.add_command(label="选择脚本", command=self.select_command_file)

        # 保存目录操作
        self.right_click_menu.add_command(label="保存目录", command=self.select_output_directory)

        self.right_click_menu.add_separator()

        devcie_test = tk.Menu(self.right_click_menu, tearoff=0)
        # devcie_test.add_command(label="设备测试", command=self.compare_command_results)
        # 添加 TCP Ping 测试选项

        # 添加 Ping 相关选项
        devcie_test.add_separator()
        devcie_test.add_command(label="Login 测试", command=self.start_test_connections)
        # devcie_test.add_command(label="执行脚本", command=self.start_command_execution)
        devcie_test.add_command(label="Ping-1次", command=self.start_ping_test)
        devcie_test.add_command(label="Ping-持续", command=self.start_continuous_ping)
        devcie_test.add_command(label="Ping-指定N次", command=self.start_specified_ping)
        # 添加停止Ping选项
        devcie_test.add_command(label="停止Ping进程", command=self.stop_ping)

        # 文件操作子菜单
        file_menu = tk.Menu(self.right_click_menu, tearoff=0)
        file_menu.add_command(label="选择设备文件", command=self.select_device_file)
        file_menu.add_command(label="下载模板", command=self.download_template)
        file_menu.add_command(label="选择保存目录", command=self.select_output_directory)
        file_menu.add_command(label="选择命令脚本", command=self.select_command_file)
        file_menu.add_separator()
        file_menu.add_command(label="导出日志", command=self.export_log)
        file_menu.add_command(label="导出脚本执行结果ZIP", command=self.export_command_result_zip)
        file_menu.add_command(label="导出执行错误报告ZIP", command=self.export_error_report)
        file_menu.add_command(label="导出脚本执行报告", command=self.export_script_report)
        self.right_click_menu.add_cascade(label="文件", menu=file_menu)

        # 设备操作子菜单
        device_menu = tk.Menu(self.right_click_menu, tearoff=0)
        device_menu.add_command(label="全选", command=self.select_all)
        device_menu.add_command(label="复制", command=self.copy_selected)
        device_menu.add_command(label="粘贴", command=self.paste_selected)
        device_menu.add_command(label="剪切", command=self.cut_selected)
        device_menu.add_command(label="删除", command=self.delete_selected)
        self.right_click_menu.add_cascade(label="设备", menu=device_menu)

        # 测试操作子菜单
        test_menu = tk.Menu(self.right_click_menu, tearoff=0)
        test_menu.add_command(label="Ping-持续", command=self.start_continuous_ping)
        test_menu.add_command(label="Ping-指定N次", command=self.start_specified_ping)
        test_menu.add_command(label="Ping-1次", command=self.start_ping_test)
        test_menu.add_command(label="Login 测试", command=self.start_test_connections)
        test_menu.add_command(label="TcpPing 测试", command=self.tcp_ping_test)
        self.right_click_menu.add_cascade(label="测试", menu=test_menu)

        # 对比功能子菜单
        compare_menu = tk.Menu(self.right_click_menu, tearoff=0)
        compare_menu.add_command(label="文件对比", command=self.compare_command_results_diff)
        compare_menu.add_command(label="文本对比增强", command=self.open_compare_window)
        self.right_click_menu.add_cascade(label="对比", menu=compare_menu)

        # 导出功能子菜单
        export_menu = tk.Menu(self.right_click_menu, tearoff=0)
        export_menu.add_command(label="导出日志", command=self.export_log)
        export_menu.add_command(label="导出脚本执行结果ZIP", command=self.export_command_result_zip)
        export_menu.add_command(label="导出执行错误报告ZIP", command=self.export_error_report)
        export_menu.add_command(label="导出脚本执行报告", command=self.export_script_report)
        self.right_click_menu.add_cascade(label="导出", menu=export_menu)

        # 设置子菜单
        settings_menu = tk.Menu(self.right_click_menu, tearoff=0)
        settings_menu.add_command(label="设置超时时间", command=lambda: self.timeout_entry.focus())
        settings_menu.add_command(label="设置命令执行间隔", command=lambda: self.cmd_interval_entry.focus())
        settings_menu.add_command(label="显示密码", command=lambda: self.toggle_password_visibility(True))
        settings_menu.add_command(label="隐藏密码", command=lambda: self.toggle_password_visibility(False))
        settings_menu.add_checkbutton(label="启用通知", variable=self.enable_notification)
        settings_menu.add_command(label="管理提示符", command=self.manage_prompt_patterns)
        settings_menu.add_command(label="管理执行结果错误判断字符", command=self.manage_error_chars)
        self.right_click_menu.add_cascade(label="设置", menu=settings_menu)

        # 审计和历史记录子菜单
        audit_menu = tk.Menu(self.right_click_menu, tearoff=0)
        audit_menu.add_command(label="命令执行审计", command=self.view_command_history)
        audit_menu.add_command(label="操作日志审计", command=self.view_audit_log)
        self.right_click_menu.add_cascade(label="审计", menu=audit_menu)

        # 定时任务管理子菜单
        schedule_menu = tk.Menu(self.right_click_menu, tearoff=0)
        schedule_menu.add_command(label="添加定时任务", command=self.add_task)
        schedule_menu.add_command(label="删除定时任务", command=self.remove_task)
        schedule_menu.add_command(label="暂停定时任务", command=self.pause_task)
        schedule_menu.add_command(label="恢复定时任务", command=self.resume_task)
        self.right_click_menu.add_cascade(label="任务管理", menu=schedule_menu)

        # 帮助和支持子菜单
        help_menu = tk.Menu(self.right_click_menu, tearoff=0)
        help_menu.add_command(label="关于", command=self.show_about)
        help_menu.add_command(label="使用帮助", command=self.show_help)
        help_menu.add_command(label="用户反馈", command=self.show_feedback)
        help_menu.add_command(label="下载用户手册", command=self.download_user_manual)
        self.right_click_menu.add_cascade(label="帮助", menu=help_menu)


        # 绑定右键事件
        self.tree.bind("<Button-3>", self.show_right_click_menu)

    def edit_selected_devices(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showinfo("提示", "请先选择要编辑的设备")
            return

        # 弹出对话框，让用户选择要编辑的字段和新值
        edit_window = tk.Toplevel(self.root)
        edit_window.title("编辑设备信息")
        edit_window.geometry("400x150")
        self.center_window(edit_window)

        # 创建字段选择下拉框
        tk.Label(edit_window, text="选择要编辑的字段:").grid(row=0, column=0, padx=10, pady=10)
        field_var = tk.StringVar()
        field_menu = ttk.Combobox(edit_window, textvariable=field_var,
                                  values=["hostname", "username", "password", "protocol", "device_type", "path",
                                          "port"])
        field_menu.grid(row=0, column=1, padx=10, pady=10)
        field_menu.set("hostname")  # 默认选择主机名

        # 创建新值输入框
        tk.Label(edit_window, text="输入新值:").grid(row=1, column=0, padx=10, pady=10)
        new_value_entry = tk.Entry(edit_window, width=24)
        new_value_entry.grid(row=1, column=1, padx=10, pady=10)

        # 确认按钮
        def confirm_edit():
            field = field_var.get()
            new_value = new_value_entry.get().strip()

            if not field or not new_value:
                messagebox.showerror("错误", "请选择字段并输入新值")
                return

            # 遍历选中的设备并更新字段
            for item_id in selected_items:
                values = list(self.tree.item(item_id, "values"))
                col_index = self.tree["columns"].index(field)
                hostname = values[0]
                edit_value = values[col_index]
                self.log(f"[成功]：编辑设备成功，主机{hostname}：{edit_value}改为新值{new_value} ")

                self.log_audit("[操作]：批量编辑设备", f"编辑设备成功，主机{hostname}：{edit_value}改为新值{new_value} ")

                # 校验输入内容
                if field == "hostname":
                    if not self.validate_ip(new_value):
                        messagebox.showerror("错误", f"无效的IP地址：{new_value}")
                        return
                    values[col_index] = new_value
                elif field == "port":
                    if not self.validate_port(new_value):
                        messagebox.showerror("错误", f"无效的端口号：{new_value}")
                        return
                    values[col_index] = new_value
                elif field == "password":
                    # 更新密码并保存到原始密码字典
                    hostname = values[0]
                    self.original_passwords[hostname] = new_value
                    values[col_index] = '*' * len(new_value)
                else:
                    values[col_index] = new_value

                # 更新表格
                self.tree.item(item_id, values=values)

            # 自动保存设备信息
            self.auto_save_device_info()
            messagebox.showinfo("成功", f"编辑 {len(selected_items)} 台设备成功")
            # self.log(f"[成功]: 编辑 {len(selected_items)} 台设备成功“)
            self.log(f"[成功]：编辑 {len(selected_items)} 台设备成功，输入新值为{new_value},详情看日志")
            self.log_audit("[操作]：批量编辑设备", f"编辑 {len(selected_items)} 台设备成功，输入新值为{new_value}")
            edit_window.destroy()

        tk.Button(edit_window, text="确认", command=confirm_edit).grid(row=2, column=0, columnspan=2, pady=20)

    def show_right_click_menu(self, event):
        self.right_click_menu.post(event.x_root, event.y_root)

    def select_all(self):
        # 全选设备
        for item_id in self.tree.get_children():
            self.tree.selection_add(item_id)
        self.log(f"[提示]：已全选所有设备,总计: {len(self.tree.get_children())}")

    def undo(self, event=None):
        # 撤销操作
        # 这里可以实现撤销功能，例如撤销最近的编辑操作
        self.log("[警告]：撤销操作（无法执行操作）")

    def redo(self, event=None):
        # 重做操作
        # 这里可以实现重做功能，例如重做最近的撤销操作
        self.log("[警告]：重做操作（无法执行操作）")

    def copy_selected(self, event=None):
        # """复制选中设备的信息"""
        selected_items = self.tree.selection()
        if not selected_items:
            # messagebox.showinfo("提示", "请先选择设备")
            self.log(f"[提示]：复制操作未选择设备信息")

            return

        copied_data = []
        for item_id in selected_items:
            values = self.tree.item(item_id, "values")
            copied_data.append(values)

        self.log(f"[提示]：快捷键Ctrl+C: 复制设备信息成功 {copied_data}")
        self.root.clipboard_clear()
        self.root.clipboard_append(str(copied_data))

    def paste_selected(self, event=None):
        # """粘贴设备信息"""
        try:
            clipboard_data = self.root.clipboard_get()
            # 解析剪贴板数据
            if clipboard_data.startswith("["):
                # 如果是列表格式，直接使用
                devices = eval(clipboard_data)
            else:
                # 如果是文本格式，按行分割
                devices = [line.strip().split("\t") for line in clipboard_data.split("\n")]

            for device in devices:
                if len(device) >= 9:
                    # 插入到 Treeview
                    self.tree.insert("", "end", values=device)
                    # 更新原始密码字典
                    hostname = device[0]
                    password = device[2]
                    self.original_passwords[hostname] = password

            self.auto_save_device_info()
            self.log(f"[提示]：快捷键 Ctrl+V 粘贴设备信息成功 {clipboard_data}")
        except Exception as e:
            # messagebox.showerror("错误", f"快捷键 Ctrl+V 粘贴设备信息失败 {str(e)}")
            self.log(f"[错误]：快捷键 Ctrl+V 粘贴设备信息失败: {str(e)}")

    def cut_selected(self, event=None):
        # """剪切选中设备的信息"""
        self.copy_selected()

        self.delete_selected()
        cut_data = []
        # self.log(f"[提示]：快捷键 Ctrl+X：已剪切选中设备信息 {cut_data}")

    def delete_selected(self):
        # """删除选中设备"""
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showinfo("提示", "请先选择设备")
            return

        for item_id in selected_items:
            values = self.tree.item(item_id, "values")
            self.tree.delete(item_id)

        self.auto_save_device_info()
        self.log(f"[提示]：快捷键 Delete：已删除选中设备 {values}")


    def export_selected_devices(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showinfo("提示", "请先选择设备")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            try:
                columns = ["hostname", "username", "password", "protocol", "device_type", "path", "port", "test_result",
                           "execute_result", "super_password"]
                data = []

                for item_id in selected_items:
                    values = self.tree.item(item_id, "values")
                    hostname = values[0]  # 获取主机名

                    # 获取原始密码和超级密码
                    password = self.original_passwords.get(hostname, "")
                    super_password = self.super_passwords.get(hostname, "")

                    # 构建要导出的行数据
                    row_data = [
                        values[0],  # hostname
                        values[1],  # username
                        password,  # 原始密码（从字典获取）
                        values[3],  # protocol
                        values[4],  # device_type
                        values[5],  # path
                        values[6],  # port
                        values[7],  # test_result
                        values[8],  # execute_result
                        super_password  # 原始超级密码（从字典获取）
                    ]
                    data.append(row_data)

                df = pd.DataFrame(data, columns=columns)
                df.to_excel(file_path, index=False)
                messagebox.showinfo("成功", "设备信息导出成功")
                self.log(f"[提示]：设备信息已导出至: {file_path}")
                self.log_audit("[操作]：导出设备", "设备信息导出成功")

            except Exception as e:
                messagebox.showerror("错误", f"导出设备信息失败: {str(e)}")
                self.log(f"[错误]：导出设备信息失败: {str(e)}")
                self.log_audit("[操作]：导出设备", "设备信息导出失败")

    def import_devices(self):
        """导入设备信息"""
        file_path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])
        if file_path:
            try:
                df = pd.read_excel(file_path)
                self.tree.delete(*self.tree.get_children())
                # self.original_passwords = {}
                for _, row in df.iterrows():
                    hostname = str(row["hostname"]).strip()
                    username = str(row["username"]).strip() if "username" in row else None
                    password = str(row["password"]).strip() if "password" in row else None
                    # connection_protocol = str(row["protocol"]).strip() if "protocol" in row else None
                    connection_protocol = str(row["protocol"]).strip().lower() if "protocol" in row else None
                    device_type = str(row["device_type"]).strip() if "device_type" in row else None
                    path = str(row["path"]).strip() if "path" in row else None
                    # port = str(row["port"]).strip() if "port" in row else None
                    # port = str(row["port"]).strip()
                    port = row.get("port", None)  # 获取端口值，如果列不存在则为None
                    super_password = str(row["super_password"]).strip() if "super_password" in row else ""  # 加载Super密码

                    # 检查端口是否为空或无效
                    if pd.isna(port) or str(port).strip() == "" or port == "nan":
                        port = None  # 确保空值或特殊值被识别为None

                    # 自动填充默认端口
                    if port is None:  # 端口为空时
                        # self.log("端口为空，进行默认填充")
                        if connection_protocol.lower() == "telnet":
                            port = 23
                            self.log(f"[提示]：设备 {hostname} Telnet端口为空，自动填充 23")
                        elif connection_protocol.lower() == "ssh":
                            port = 22
                            self.log(f"[提示]：设备 {hostname} SSH端口为空，自动填充 22")
                        else:
                            port = None  # 其他协议默认端口设置为None，允许后续处理
                    else:
                        try:
                            port = int(float(port))  # 支持浮点数格式的端口
                        except (ValueError, TypeError):
                            self.log(f"[错误]：设备 {hostname}:{port} 端口无效")
                            continue  # 跳过当前设备，继续处理下一个设备

                    # 验证 IP 格式
                    if not self.validate_ip(hostname):
                        self.log(f"[错误]：设备 {hostname} IP无效")
                        continue

                    # 强制将端口转换为整数
                    try:
                        if pd.notnull(port):
                            port = int(port)
                            if port < 1 or port > 65535:
                                raise ValueError
                        else:
                            port = ""
                    except ValueError:
                        self.log(f"[错误]：设备 {hostname}:{port} 端口无效")
                        continue

                    # 填充默认值
                    username = username if username else None
                    password = password if password else None
                    super_password = super_password if super_password else None
                    connection_protocol = connection_protocol if connection_protocol else ssh
                    device_type = device_type if device_type else h3c
                    path = path if path else None

                    # 插入到 Treeview
                    masked_password = '*' * len(password) if password else None
                    masked_super_password = '*' * len(super_password) if super_password else None

                    self.tree.insert("", "end", values=(
                        hostname,
                        username,
                        masked_password,
                        connection_protocol,
                        device_type,
                        path,
                        port,
                        "",
                        "",
                        masked_super_password  # Super密码列

                    ))

                    # 更新原始密码和Super密码字典
                    if password:
                        self.original_passwords[hostname] = password
                    if super_password:  # 如果Super密码字段不为空字符串
                        self.super_passwords[hostname] = super_password if super_password != "nan" else ""  # 修正这里的逻辑判断
                    else:
                        self.super_passwords[hostname] = ""  # 如果Super密码字段为空字符串，也设置为空

                messagebox.showinfo("成功", "设备信息导入成功")
                self.log(f"[提示]：设备信息已从: {file_path} 导入")
                self.log_audit("[操作]：导入设备", f"设备信息导入成功:{file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"导入设备信息失败: {str(e)}")
                self.log(f"[错误]：导入设备信息失败: {str(e)}")
                self.log_audit("[操作]：导入设备", f"设备信息导入失败:{file_path}")


    def on_tree_double_click(self, event):
        """处理 Treeview 单元格双击事件"""
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return

        column = self.tree.identify_column(event.x)
        item_id = self.tree.identify_row(event.y)

        # 获取当前列的标题
        column_title = self.tree.heading(column, "text")

        # 获取当前单元格的值
        values = self.tree.item(item_id, "values")
        current_value = values[self.tree["columns"].index(column_title)]

        # 使用 simpledialog 弹出对话框获取新值
        new_value = simpledialog.askstring("编辑", f"请输入新的 {column_title}:", initialvalue=current_value)
        if new_value is not None:
            # 校验输入内容
            if column_title == "hostname":
                if not self.validate_ip(new_value):
                    return
            elif column_title == "port":
                if not self.validate_port(new_value):
                    return

            values = list(self.tree.item(item_id, "values"))
            col_index = self.tree["columns"].index(column_title)
            old_value = values[col_index]
            values[col_index] = new_value
            self.tree.item(item_id, values=values)

            # 如果修改的是密码列，更新原始密码字典
            if col_index == 2:  # 假设密码是第三列（索引为2）
                hostname = values[0]
                self.original_passwords[hostname] = new_value
            # 如果修改的是密码列，更新原始密码字典
            if col_index == 9:  # 假设密码是第三列（索引为2）
                hostname = values[0]
                self.super_passwords [hostname] = new_value


            # 自动保存设备信息
            self.auto_save_device_info()

            # 记录日志
            hostname = values[0]
            self.log(f"[提示]：用户修改了设备信息： {hostname}: {column_title}: {old_value} 改为-> {new_value}")

    def update_tree_cell(self, item_id, column, new_value):
        # 更新表格中的单元格值
        values = list(self.tree.item(item_id, "values"))
        col_index = int(column[1]) - 1
        values[col_index] = new_value
        self.tree.item(item_id, values=values)

        # 如果修改的是密码列，更新原始密码字典
        if col_index == 2:  # 假设密码是第三列（索引为2）
            hostname = values[0]
            self.original_passwords[hostname] = new_value
            # 如果修改的是super密码列，更新原始密码字典
        if col_index == 9:  # 假设密码是第三列（索引为2）
            hostname = values[0]
            self.super_passwords[hostname] = new_value

        # 销毁输入框
        for widget in self.tree.winfo_children():
            if isinstance(widget, tk.Entry):
                widget.destroy()

    def select_output_directory(self):
        # 选择保存目录
        try:
            directory = filedialog.askdirectory()
            if directory:
                self.output_directory = directory
                self.selected_save_directory = directory
                self.status_label.config(text=f"状态：选择保存目录 {directory}")
                self.log(f"[提示]：选择了保存目录：{directory}")
                self.log_audit("[操作]：选择保存目录", f"用户选择了命令执行保存目录：{directory}")
        except Exception as e:
            self.log(f"[错误]：选择保存目录时出错: {str(e)}")
            messagebox.showerror("错误", f"选择保存目录时出错: {str(e)}")

    # def select_command_file(self):
    #     file_path = filedialog.askopenfilename(filetypes=[("文本文件", "*.txt")])
    #     if file_path:
    #         self.command_file_path = file_path
    #         for item_id in self.tree.get_children():
    #             self.tree.set(item_id, "path", file_path)
    #         self.status_label.config(text=f"状态：选择命令脚本 {file_path}")
    #         self.log(f"[提示]：选择了命令脚本：{file_path}")
    #         self.log_audit("[操作]：选择脚本", "用户选择了命令脚本")
    def select_command_file(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showinfo("提示", "请先选择设备")
            return

        file_path = filedialog.askopenfilename(filetypes=[("文本文件", "*.txt")])
        if file_path:
            self.command_file_path = file_path
            # 仅更新选中设备的脚本路径
            for item_id in selected_items:
                values = list(self.tree.item(item_id, "values"))
                values[5] = file_path  # 假设第6列是命令脚本路径
                self.tree.item(item_id, values=values)
            self.log(f"[成功]：已选择命令脚本：{file_path} 并应用到选中的设备")

    def connect_to_device(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showinfo("提示", "请先选择要连接的设备")
            return

        for item_id in selected_items:
            values = self.tree.item(item_id, "values")
            hostname = values[0]
            username = values[1]
            password = self.original_passwords.get(hostname, "")
            connection_protocol = values[3]
            device_type = values[4]
            port = values[6] if len(values) > 6 else None
            # 获取Super密码
            super_password = self.super_passwords.get(hostname, "")  # 获取Super密码
            print(f"检测到Super密码,连接设备前：{self.super_passwords}")

            # 创建一个线程来处理连接，避免主程序假死
            threading.Thread(target=self._connect_to_device_thread,
                             args=(item_id, hostname, username, password, connection_protocol, device_type, port, super_password),
                             daemon=True).start()

    def _connect_to_device_thread(self, item_id, hostname, username, password, connection_protocol, device_type, port,super_password):
        try:
            # connection = None  # 初始化 connection 变量为 None
            connection_protocol = connection_protocol.lower()
            device_type = device_type.lower()
            port = int(float(port))
            # 创建连接
            if connection_protocol == "telnet":
                # 支持无用户名的Telnet连接
                if not username:  # 如果用户名为空
                    connection = {
                        "device_type": "cisco_ios_telnet",
                        "host": hostname,
                        "password": password,
                        "port": port or 23,
                        "timeout": self.ssh_telnet_timeout
                    }
                else:
                    connection = {
                        "device_type": "cisco_ios_telnet",
                        "host": hostname,
                        "username": username,
                        "password": password,
                        "port": port or 23,
                        "timeout": self.ssh_telnet_timeout
                    }
            elif connection_protocol == "ssh":
                if device_type.lower() == "h3c":
                    connection = {
                        "device_type": "hp_comware",
                        "host": hostname,
                        "username": username,
                        "password": password,
                        "port": port or 22,
                        "timeout": self.ssh_telnet_timeout
                    }
                elif device_type.lower() == "huawei":
                    connection = {
                        "device_type": "huawei",
                        "host": hostname,
                        "username": username,
                        "password": password,
                        "port": port or 22,
                        "timeout": self.ssh_telnet_timeout
                    }
                elif device_type.lower() == "cisco":
                    connection = {
                        "device_type": "cisco_ios",
                        "host": hostname,
                        "username": username,
                        "password": password,
                        "port": port or 22,
                        "timeout": self.ssh_telnet_timeout
                    }
                elif device_type.lower() == "linux":
                    connection = {
                        "device_type": "linux_ssh",
                        "host": hostname,
                        "username": username,
                        "password": password,
                        "port": port or 22,
                        "timeout": self.ssh_telnet_timeout,
                        "global_delay_factor": 2
                    }
                else:
                    connection = {
                        "device_type": "generic",
                        "host": hostname,
                        "username": username,
                        "password": password,
                        "port": port or 22,
                        "timeout": self.ssh_telnet_timeout
                    }
            if connection is None:
                raise ValueError("无法确定连接类型")

            conn = ConnectHandler(**connection)
            conn.find_prompt()  # 确保连接已建立并找到提示符
            self.log(f"[成功]：成功连接到设备 {hostname}")

            # 处理Super密码
            if super_password:
                print(f"处理Super密码逻辑")
                if device_type == "cisco":
                    conn.send_command("enable")
                    conn.send_command_timing(super_password)
                    conn.find_prompt()  # 确保Super密码已正确输入并找到新的提示符
                    print(f"检测CISCO设备-到Super密码{super_password},发送Super密码")
                else:
                    conn.send_command_timing("super")
                    conn.send_command_timing(super_password)
                    conn.find_prompt()  # 确保Super密码已正确输入并找到新的提示符
                    print(f"检测H3C-HUAWEI设备-到Super密码{super_password},发送Super密码")

            #
            # 创建交互窗口
            self.root.after(0, lambda: self.create_interactive_window(conn, hostname))

        except (NetMikoTimeoutException, NetMikoAuthenticationException) as e:
            self.log(f"[错误]：连接设备 {hostname} 失败: {str(e)}")
            self.log_audit("[操作]：连接设备", f"连接设备 {hostname} 失败: {str(e)}")
            messagebox.showerror("连接失败", f"连接设备 {hostname} 失败: {str(e)}")
        except Exception as e:
            self.log(f"[错误]：连接设备 {hostname} 时出错: {str(e)}")
            self.log_audit("[操作]：连接设备", f"连接设备 {hostname} 时出错: {str(e)}")
            messagebox.showerror("连接出错", f"连接设备 {hostname} 时出错: {str(e)}")



    def create_interactive_window(self, conn, hostname, is_linux=False):
        # 创建交互窗口
        interactive_window = tk.Toplevel(self.root)
        interactive_window.title(f"设备交互 - {hostname}")
        interactive_window.geometry("800x600")

        # 创建命令输入框和输出框
        output_frame = tk.Frame(interactive_window)
        output_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        output_text = ScrolledText(output_frame, height=20)
        output_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        input_frame = tk.Frame(interactive_window)
        input_frame.pack(fill=tk.X, padx=10, pady=10)

        command_entry = tk.Entry(input_frame)
        command_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        def send_command():
            command = command_entry.get().strip()
            if command.lower() == "exit":
                conn.disconnect()
                interactive_window.destroy()
                return

            if command:
                output_text.insert(tk.END, f"> {command}\n")
                output_text.update_idletasks()
                try:
                    output = conn.send_command_timing(command)
                    # 清理ANSI转义序列
                    ansi_escape = re.compile(r'(\x9B|\x1B\[)[0-?]*[ -/]*[@-~]')
                    cleaned_output = ansi_escape.sub('', output)
                    output_text.insert(tk.END, f"{cleaned_output}\n")
                    output_text.update_idletasks()
                    command_entry.delete(0, tk.END)

                    # 将命令和输出存储到列表中
                    command_outputs.append({"命令": command, "输出": cleaned_output})

                except Exception as e:
                    output_text.insert(tk.END, f"执行命令失败: {str(e)}\n")
                    output_text.update_idletasks()

        send_button = tk.Button(input_frame, text="发送", command=send_command)
        send_button.pack(side=tk.RIGHT, padx=5)

        command_entry.bind("<Return>", lambda event: send_button.invoke())

        # 用于存储命令和输出
        command_outputs = []

        # 定义断开连接超时时间（秒）
        DISCONNECT_TIMEOUT = 2

        def on_closing():
            start_time = time.time()  # 记录关闭开始时间
            # self.log(f"[调试]：开始关闭窗口，耗时计时器已启动")

            def disconnect_and_save():
                nonlocal start_time
                disconnected = False
                try:
                    # 尝试断开设备连接
                    if conn and hasattr(conn, 'disconnect'):
                        # 为Linux设备设置较短的超时时间
                        if is_linux:
                            conn.timeout = DISCONNECT_TIMEOUT
                        conn.disconnect()
                        disconnected = True
                        self.log(f"[提示]：成功断开与设备 {hostname} 的连接")
                except Exception as e:
                    self.log(f"[警告]：断开设备 {hostname} 连接时出错: {str(e)}")
                finally:
                    # 保存交互日志
                    try:
                        self.save_interactive_session_logs(hostname, command_outputs)
                        self.log(f"[提示]：交互日志已保存")
                    except Exception as e:
                        self.log(f"[错误]：保存交互日志失败: {str(e)}")
                    finally:
                        # 销毁窗口
                        self.root.after(0, interactive_window.destroy)
                        # end_time = time.time()  # 记录关闭结束时间
                        # self.log(f"[调试]：关闭窗口耗时: {end_time - start_time:.2f} 秒")

            # 启动线程处理断开连接和保存日志
            thread = threading.Thread(target=disconnect_and_save, daemon=True)
            thread.start()

            # 设置一个计时器，如果线程在超时时间内未完成，则强制销毁窗口
            def force_destroy_window():
                if thread.is_alive():
                    self.log(f"[警告]：断开连接操作超时，强制销毁窗口-{hostname}")
                    self.root.after(0, interactive_window.destroy)

            # 设置一个计时器，例如 5 秒后强制销毁窗口
            self.root.after(DISCONNECT_TIMEOUT * 1000 + 1000, force_destroy_window)

        interactive_window.protocol("WM_DELETE_WINDOW", on_closing)

        # 初始输出
        output_text.insert(tk.END, f"已连接到设备: {hostname}\n")
        output_text.insert(tk.END, "输入命令 (输入'exit'退出):\n")

        return interactive_window



    def save_interactive_session_logs(self, hostname, command_outputs):
        # 清理ANSI转义序列
        ansi_escape = re.compile(r'(\x9B|\x1B\[)[0-?]*[ -/]*[@-~]')

        # 确保 command_outputs 是一个包含字典的列表，每个字典包含 "命令" 和 "输出" 键
        if not all(isinstance(item, dict) and "命令" in item and "输出" in item for item in command_outputs):
            self.log(f"[错误]：command_outputs 格式错误，无法保存交互日志")
            return

        # 清理每个命令的输出
        cleaned_outputs = []
        for item in command_outputs:
            cleaned_command = ansi_escape.sub('', item["命令"]) if isinstance(item["命令"], str) else item["命令"]
            cleaned_output = ansi_escape.sub('', item["输出"]) if isinstance(item["输出"], str) else item["输出"]
            cleaned_outputs.append({"命令": cleaned_command, "输出": cleaned_output})

        # 确保设备目录存在
        directory = os.path.join(self.selected_save_directory or self.default_device_log_dir, hostname)
        os.makedirs(directory, exist_ok=True)

        # 获取当前时间戳
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # 保存为Excel文件
        df = pd.DataFrame(cleaned_outputs)
        excel_filename = f"{hostname}_{timestamp}_interactive.xlsx"
        excel_path = os.path.join(directory, excel_filename)
        df.to_excel(excel_path, index=False)
        self.log(f"[提示]：交互日志已保存为 Excel 文件：{excel_path}")

        # 保存为TXT文件
        txt_filename = f"{hostname}_{timestamp}_interactive.txt"
        txt_path = os.path.join(directory, txt_filename)
        with open(txt_path, "w", encoding="utf-8") as txt_file:
            for entry in cleaned_outputs:
                txt_file.write(f"命令: {entry['命令']}\n")
                txt_file.write(f"输出:\n{entry['输出'].strip()}\n\n")
        self.log(f"[提示]：交互日志已保存为TXT文本文件：{txt_path}")

        # 保存为HTML文件
        html_filename = f"{hostname}_{timestamp}_interactive.html"
        html_path = os.path.join(directory, html_filename)
        html_content = """
        <html>
        <head>
            <title>交互日志 - {hostname}</title>
            <style>
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
            </style>
        </head>
        <body>
            <h1>交互日志 - {hostname}</h1>
            <table>
                <tr>
                    <th>命令</th>
                    <th>输出</th>
                </tr>
                {table_rows}
            </table>
        </body>
        </html>
        """.format(
            hostname=hostname,
            table_rows="\n".join([
                f"<tr><td>{row['命令']}</td><td><pre>{row['输出']}</pre></td></tr>"
                for row in cleaned_outputs
            ])
        )
        with open(html_path, "w", encoding="utf-8") as html_file:
            html_file.write(html_content)
        self.log(f"[提示]：交互日志已保存为 HTML 文件：{html_path}")

    def load_custom_styles(self):
        # 加载用户自定义样式
        config_dir = os.path.join(os.getcwd(), "config")
        os.makedirs(config_dir, exist_ok=True)
        styles_path = os.path.join(config_dir, "styles.json")

        if os.path.exists(styles_path):
            try:
                with open(styles_path, "r") as f:
                    self.custom_styles = json.load(f)
            except Exception as e:
                self.custom_styles = {}
                self.log(f"[错误]：加载自定义样式失败: {str(e)}")
        else:
            self.custom_styles = {}

    def save_custom_styles(self):
        # 保存用户自定义样式
        config_dir = os.path.join(os.getcwd(), "config")
        os.makedirs(config_dir, exist_ok=True)
        styles_path = os.path.join(config_dir, "styles.json")

        with open(styles_path, "w") as f:
            json.dump(self.custom_styles, f, indent=4)

    def start_ping_test(self):
        self.cancel_event.clear()
        self.ping_test_button.config(state=tk.DISABLED)
        self.cancel_button.config(state=tk.NORMAL)
        # self.execute_button.config(state=tk.DISABLED)
        threading.Thread(target=self.ping_test, daemon=True).start()
        self.log("[提示]：用户点击Ping测试")
        self.log_audit("[操作]：执行 ping 测试", "用户点击 ping 测试按钮")
        self.log("[警告]：启动 Ping 测试…")

    def ping_test(self):  # 优化tcp ping通知结果为统一生成
        selected_items = self.tree.selection()
        if not selected_items:
            self.log("[错误]：执行操作请先选择设备")
            self.ping_test_button.config(state=tk.NORMAL)
            self.cancel_button.config(state=tk.DISABLED)
            # self.execute_button.config(state=tk.DISABLED)
            return

        self.cancel_event.clear()
        self.ping_test_button.config(state=tk.DISABLED)
        self.cancel_button.config(state=tk.NORMAL)
        self.progress_bar["maximum"] = len(selected_items)  # 设置进度条最大值
        self.progress_bar["value"] = 0
        self.status_label.config(text="状态：正在进行 Ping 测试...")

        success_count = 0
        failure_count = 0
        ping_results = []  # 用于收集每个设备的测试结果

        devices = {}
        for item_id in selected_items:
            values = self.tree.item(item_id, "values")
            hostname = values[0]
            username = values[1]
            password = self.original_passwords.get(hostname, "")
            connection_protocol = values[3]
            device_type = values[4]
            path = values[5]
            port = values[6] if len(values) > 6 else None
            devices[item_id] = {
                "hostname": hostname,
                "username": username,
                "password": password,
                "connection_protocol": connection_protocol,
                "device_type": device_type,
                "path": path,
                "port": port
            }

        try:
            self.ping_timeout = int(self.timeout_entry.get() or 3)
        except ValueError:
            messagebox.showerror("错误", "Ping超时时间必须为整数")
            self.ping_test_button.config(state=tk.NORMAL)
            self.status_label.config(text="状态：Ping测试停止")
            self.log("[错误]：超时时间必须为整数，Ping测试停止")
            return

        def ping_device(device, item_id):
            nonlocal success_count, failure_count
            if self.cancel_event.is_set():
                return
            try:
                hostname = device["hostname"]
                if self.ping_type_var.get() == "ICMP":
                    #response_time = ping(hostname, timeout=self.ping_timeout / 1000)
                    response_time = ping(hostname, timeout=self.ping_timeout)
                    # if response_time is None:
                        # result_text = "0[失败] 无响应"
                        # failure_count += 1
                    if response_time is None or response_time <= 0:
                    # 如果 response_time 是 None 或者小于等于0，认为是失败
                        result_text = "0[失败]: 无响应"
                        failure_count += 1
                        self.tree.item(item_id, tags=("failure",))
                    else:
                        result_text = f"1[成功]: 延迟 {response_time * 1000:.2f} ms"
                        success_count += 1
                        self.tree.item(item_id, tags=("success",))
                elif self.ping_type_var.get() == "TCP":
                    if self.ping_tcp_port(device):
                        result_text = "1[成功]: TCP 端口连通"
                        success_count += 1
                        self.tree.item(item_id, tags=("success",))
                    else:
                        result_text = "0[失败]: TCP 端口不通"
                        failure_count += 1
                        self.tree.item(item_id, tags=("failure",))
                elif self.ping_type_var.get() == "UDP":
                    # 实现UDP Ping逻辑
                    try:
                        port = 22
                        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
                        sock.settimeout(self.ping_timeout)
                        # 发送UDP数据包
                        port = int(float(port))
                        sock.sendto(b'', (hostname, port))
                        # 等待响应
                        data, addr = sock.recvfrom(1024)
                        result_text = f"1[成功]: UDP 端口连通"
                        success_count += 1
                        self.tree.item(item_id, tags=("success",))
                        # self.tree.item(item_id, tags=("success",))
                    except socket.timeout:
                        result_text = f"0[失败]: UDP 端口无响应"
                        failure_count += 1
                        self.tree.item(item_id, tags=("failure",))
                        # self.tree.item(item_id, tags=("failure",))
                    except Exception as e:
                        result_text = f"0[错误]: {str(e)}"
                        failure_count += 1
                        self.tree.item(item_id, tags=("failure",))
                    finally:
                        sock.close()

                elif self.ping_type_var.get() == "SNMP":
                    try:
                        for (errorIndication, errorStatus, _, _) in getCmd(
                                SnmpEngine(),
                                CommunityData('public', mpModel=1),
                                UdpTransportTarget((hostname, {port}), timeout=2, retries=0),
                                ContextData(),
                                ObjectType(ObjectIdentity('1.3.6.1.2.1.1.1.0'))
                        ):
                            if errorIndication:
                                raise Exception(str(errorIndication))
                            result_text = "1[成功]: SNMP 协议连通"
                            success_count += 1
                            self.tree.item(item_id, tags=("success",))
                    except Exception as e:
                        print(f"[SNMP调试] 异常: {e}")  # ← 打印异常信息
                        result_text = "0[失败]: SNMP 协议不通"
                        failure_count += 1
                        self.tree.item(item_id, tags=("failure",))
                else:
                    result_text = f"0[错误]: 未知协议类型: {protocol_type}"
                    failure_count += 1
                    self.tree.item(item_id, tags=("failure",))

                ping_results.append(f"{hostname.ljust(20)} | {result_text}")
                # self.root.after(0, self.update_tree_item, item_id, result_text)
                # self.root.after(0, lambda: self.update_progress(success_count, failure_count, len(selected_items)))
                self.log(f"[成功]：Ping 测试结果-{hostname}:{result_text}")
            except Exception as e:
                result_text = f"[错误] {str(e)}"
                failure_count += 1
                # self.tree.item(item_id, tags=("failure",))
                self.tree.item(item_id, tags=("failure",))
                # self.root.after(0, self.update_tree_item, item_id, result_text)
                # self.root.after(0, lambda: self.update_progress(success_count, failure_count, len(selected_items)))
                self.log(f"[错误]：Ping 测试结果-{hostname}: {result_text}")

            finally:
                self.root.after(0, self.update_tree_item, item_id, result_text)
                self.root.after(0, lambda: self.update_progress(success_count, failure_count, len(selected_items)))

        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = []
            for item_id in selected_items:
                values = self.tree.item(item_id, "values")
                hostname = values[0]
                username = values[1]
                password = self.original_passwords.get(hostname, "")
                connection_protocol = values[3]
                device_type = values[4]
                path = values[5]
                port = values[6] if len(values) > 6 else None
                device = {
                    "hostname": hostname,
                    "username": username,
                    "password": password,
                    "connection_protocol": connection_protocol,
                    "device_type": device_type,
                    "path": path,
                    "port": port
                }
                futures.append(executor.submit(ping_device, device, item_id))

            for future in futures:
                if self.cancel_event.is_set():
                    break
                future.result()

        # 生成汇总结果
        if ping_results:
            title = "Ping测试结果汇总"
            content = "\n".join(["主机名      | 测试结果\n"] + ping_results)
            # messagebox.showinfo(title, content)
            self.log(f"[提示]：Ping 测试汇总结果:\n {content}")
            self.send_server_chan_notification(title, content, test_type="Ping测试")

        self.status_label.config(text="状态：Ping 测试完成")
        self.ping_test_button.config(state=tk.NORMAL)
        self.cancel_button.config(state=tk.DISABLED)
        # self.execute_button.config(state=tk.NORMAL)
        self.log("[成功]：Ping 测试完成")

    def tcp_ping_test(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showinfo("提示", "请先选择设备")
            self.test_button.config(state=tk.DISABLED)
            self.cancel_button.config(state=tk.DISABLED)
            self.execute_button.config(state=tk.DISABLED)
            return

        # 进度条更新
        self.progress_bar["maximum"] = len(selected_items)
        self.progress_bar["value"] = 0
        self.status_label.config(text="状态：正在测试连接...")

        # 禁用右键菜单中的TCP Ping测试选项，防止重复点击
        # self.right_click_menu.entryconfig("TCP Ping 测试", state="disabled")

        success_count = 0
        failure_count = 0
        tcp_ping_results = []  # 用于收集每个设备的测试结果

        devices = {}
        for item_id in selected_items:
            values = self.tree.item(item_id, "values")
            hostname = values[0]
            port = values[6] if len(values) > 6 else None
            devices[item_id] = {
                "hostname": hostname,
                "port": port
            }

        try:
            self.ping_timeout = int(self.timeout_entry.get() or 3)
        except ValueError:
            messagebox.showerror("错误", "Ping超时时间必须为整数")
            self.ping_test_button.config(state=tk.NORMAL)
            self.status_label.config(text="状态：Ping测试停止")
            self.log("[错误]：超时时间必须为整数，Ping测试停止")
            return

        # 启动TCP Ping测试线程
        def run_tcp_ping():
            nonlocal success_count, failure_count

            def ping_device(device, item_id):
                nonlocal success_count, failure_count
                try:
                    hostname = device["hostname"]
                    port = device.get("port", None)
                    port = int(float(port))
                    if port is None:
                        result_text = "0[失败] 未配置端口号"
                        failure_count += 1
                        self.tree.item(item_id, tags=("failure",))
                    else:
                        if self.ping_tcp_port(device):
                            result_text = "1[成功] TCP 端口连通"
                            success_count += 1
                            self.tree.item(item_id, tags=("success",))
                        else:
                            result_text = "0[失败] TCP 端口不通"
                            failure_count += 1
                            self.tree.item(item_id, tags=("failure",))
                    tcp_ping_results.append(f"{hostname.ljust(20)} | {result_text}")
                    # # 更新Treeview
                    # self.root.after(0, self.update_tree_item, item_id, None, result_text)
                    # 更新 Treeview 的 test_result 列
                    self.root.after(0, self.update_tree_item, item_id, result_text, None)

                except Exception as e:
                    result_text = f"[错误] {str(e)}"
                    failure_count += 1
                    self.tree.item(item_id, tags=("failure",))
                    self.root.after(0, self.update_tree_item, item_id, None, result_text)
                finally:
                    self.root.after(0, lambda: self.update_progress(success_count, failure_count, len(selected_items)))

            with ThreadPoolExecutor(max_workers=10) as executor:
                futures = []
                for item_id in selected_items:
                    device = devices[item_id]
                    futures.append(executor.submit(ping_device, device, item_id))

                for future in futures:
                    future.result()

            # 生成汇总结果
            if tcp_ping_results:
                title = "TCP Ping 测试结果汇总"
                content = "\n".join(["主机名      | 测试结果\n"] + tcp_ping_results)
                self.log(f"[提示]：TCP Ping 测试汇总结果:\n {content}")
                self.send_server_chan_notification(title, content, test_type="TCP Ping测试")

            self.status_label.config(text="状态：TCP Ping 测试完成")
            self.log("[成功]：TCP Ping 测试完成")

            # 恢复右键菜单中的TCP Ping测试选项
            # self.root.after(0, lambda: self.right_click_menu.entryconfig("TCP Ping 测试", state="normal"))

            # self.root.after(0, self.update_tree_item, item_id, result_text)

        # 在单独的线程中执行TCP Ping测试
        threading.Thread(target=run_tcp_ping, daemon=True).start()
        self.ping_test_button.config(state=tk.NORMAL)
        self.cancel_button.config(state=tk.DISABLED)
        # self.execute_button.config(state=tk.DISABLED)


    def ping_tcp_port(self, device):
        try:
            hostname = device["hostname"]
            # port = device.get("port", 22)  # 默认端口为22
            port = int(float(device.get("port", "22")))
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
           # sock.settimeout(self.ping_timeout / 1000)  # 使用用户设置的超时时间
            sock.settimeout(self.ping_timeout)  # 使用用户设置的超时时间
            result = sock.connect_ex((hostname, int(port)))  # 确保端口号是整数
            # sock.close()
            if result == 0:
                # self.log(f"[成功]：Tcp Ping  测试结果 - {hostname}:{port} 端口连通")
                return True
            else:
                # self.log(f"[失败]：Tcp Ping  测试结果 - {hostname}:{port} 端口不通")
                return False
        except Exception as e:
            self.log(f"[错误]：Tcp Ping 测试结果-{hostname}:{port} 错误: {str(e)}")
            return False
        finally:
            sock.close()  # 确保关闭套接字

    def start_test_connections(self):
        self.cancel_event.clear()
        self.log_audit("[操作]：执行 Login 测试", "用户点击 Login 测试按钮")
        self.test_button.config(state=tk.DISABLED)
        self.cancel_button.config(state=tk.NORMAL)
        self.ping_test_button.config(state=tk.NORMAL)
        self.execute_button.config(state=tk.DISABLED)


        threading.Thread(target=self.test_connections, daemon=True).start()
        self.log("[提示]：用户点击Login测试")
        self.log("[警告]：已启动连接设备测试")
        time.sleep(0.1)

    def test_connections(self):
        selected_items = self.tree.selection()
        if not selected_items:
            self.log("[错误]：执行操作请先选择设备")
            self.test_button.config(state=tk.NORMAL)
            self.cancel_button.config(state=tk.DISABLED)
            return

        self.progress_bar["maximum"] = len(selected_items)
        self.progress_bar["value"] = 0
        self.status_label.config(text="状态：正在测试连接...")
        print(f"[调试]: def test_connections后打印原始密码{self.original_passwords} {self.super_passwords}")

        success_count = 0
        failure_count = 0
        connection_results = []  # 用于收集每个设备的测试结果

        def test_device(device, item_id):
            nonlocal success_count, failure_count
            if self.cancel_event.is_set():
                return
            try:
                # 获取端口号并尝试转换为整数
                # print(f"[调试]: def test_device后打印原始密码{self.original_passwords}")
                port = device.get("port", None)
                port = int(float(port))
                if port is not None:
                    try:
                        # 将端口号转换为浮点数，再转换为整数
                        port = int(float(port))
                    except ValueError:
                        self.log(f"[错误]：设备 {hostname} 的端口号无效: {port}")
                        messagebox.showerror("错误", f"设备 {hostname} 的端口号无效: {port}")
                        return
                connection_protocol = device["connection_protocol"].lower()
                device_type = device["device_type"].lower()
                # port = int(float(port))
                # port = device.get("port", None)
                # 使用原始密码
                # original_password = self.original_passwords.get(hostname, password)
                # 尝试从原始密码字典获取密码，如果为空则使用设备信息中的密码
                # ora_password = self.original_passwords.get(hostname, device["password"])
                password = self.original_passwords.get(hostname, "")
                # 获取Super密码
                super_password = self.super_passwords.get(hostname, "")  # 获取Super密码
                outputs = []

                # print(f"打印原始密码3-def test_device(device, item_id):-{hostname}-{password}")

                if connection_protocol == "telnet":
                    # 支持无用户名的Telnet连接
                    if not device["username"]:  # 如果用户名为空
                        connection = {
                            "device_type": "cisco_ios_telnet",
                            "host": device["hostname"],
                            "password": password,
                            "port": port or 23,
                            "timeout": self.ssh_telnet_timeout
                        }
                    else:
                        connection = {
                            "device_type": "cisco_ios_telnet",
                            "host": device["hostname"],
                            "username": device["username"],
                            "password": password,
                            "port": port or 23,
                            "timeout": self.ssh_telnet_timeout
                        }
                elif connection_protocol == "ssh":
                    if device_type == "h3c":
                        connection = {
                            "device_type": "hp_comware",
                            "host": device["hostname"],
                            "username": device["username"],
                            "password": password,
                            "port": port or 22,
                            "timeout": self.ssh_telnet_timeout
                        }
                        # print(f"打印原始密码-test-内部-{hostname}{password}")

                    elif device_type == "huawei":
                        connection = {
                            "device_type": "huawei",
                            "host": device["hostname"],
                            "username": device["username"],
                            "password": password,
                            "port": port or 22,
                            "timeout": self.ssh_telnet_timeout
                        }
                    elif device_type == "cisco":
                        connection = {
                            "device_type": "cisco_ios",
                            "host": device["hostname"],
                            "username": device["username"],
                            "password": password,
                            "port": port or 22,
                            "timeout": self.ssh_telnet_timeout
                        }
                    elif device_type == "linux":
                        connection = {
                            "device_type": "linux_ssh",
                            "host": device["hostname"],
                            "username": device["username"],
                            "password": password,
                            "port": port or 22,
                            "timeout": self.ssh_telnet_timeout,
                            "global_delay_factor": 2
                        }
                    else:
                        connection = {
                            "device_type": "generic",
                            "host": device["hostname"],
                            "username": device["username"],
                            "password": password,
                            "port": port or 22,
                            "timeout": self.ssh_telnet_timeout
                        }

                        # 连接到设备
                with ConnectHandler(**connection) as conn:
                            # 处理Super密码
                            if super_password and super_password != "":
                                print(f"处理Super密码逻辑，找到super密码 {device['hostname']}")
                                # prompt = conn.find_prompt()
                                if device_type in ["h3c", "huawei"]:
                                    print(f"[调试]: 匹配到设备类型为h3c或huawei设备-实际设备类型-{device_type}")
                                    output =  conn.send_command_timing("super")
                                    # time.sleep(0.2)
                                    print(f"[调试]: 打印super命令输出之后的结果-{output}")
                                    time.sleep(0.5)
                                    output = conn.send_command_timing(super_password)
                                    print(f"检测H3C-HUAWEI设备-到Super密码{super_password},发送Super密码")
                                    # conn.find_prompt()  # 确保Super密码已正确输入并找到新的提示符
                                    print(f"[调试]：发送Super密码后的结果：{output}")
                                    outputs.append(output)

                                    # 验证Super密码是否正确
                                    if self.is_super_password_privilege(output.lower()):
                                        # self.log(f"[成功]：设备 {hostname} Super密码验证成功")
                                        self.log(f"[成功]：Super密码验证成功{device['hostname']}")
                                        print("[成功]: Super密码成功")
                                        result_text = "1[成功]: 测试正常"
                                        success_count += 1
                                        self.tree.item(item_id, tags=("success",))
                                    else:
                                        print(f"[调试2]：发送Super密码后的结果：{output}")
                                        self.log(f"[错误]：Super密码验证失败{device['hostname']}")
                                        print("0[错误]: Super密码验证失败")
                                        result_text = "0[错误]: Super密码验证失败"
                                        failure_count += 1
                                        self.tree.item(item_id, tags=("failure",))
                                        return

                                elif device_type in ["cisco"]:
                                    output = conn.send_command_timing("enable")
                                    # conn.send_command_timing(super_password)
                                    time.sleep(0.5)
                                    output = conn.send_command_timing(super_password)
                                    # conn.find_prompt()  # 确保Super密码已正确输入并找到新的提示符
                                    print(f"检测CISCO设备-到Super密码{super_password},发送Super密码")
                                    print(f"[调试]：发送Super密码后的结果：{output}")
                                    outputs.append(output)
                                    # 验证Super密码是否正确
                                    if self.is_super_password_privilege(output):
                                        # self.log(f"[成功]：设备 {hostname} Super密码验证成功")
                                        self.log(f"[成功]：Super密码验证成功 {hostname}")
                                        print("[成功]: Super密码成功")
                                        result_text = "1[成功]: 测试正常"
                                        success_count += 1
                                        self.tree.item(item_id, tags=("success",))
                                    else:
                                        self.log(f"[错误]：Super密码验证失败{hostname}")
                                        print("0[错误]: Super密码验证失败")
                                        result_text = "0[错误]: Super密码验证失败"
                                        failure_count += 1
                                        self.tree.item(item_id, tags=("failure",))
                                        return

                                else:
                                    self.log(f"[提示]：不支持的设备类型，跳过super密码{device['hostname']}")
                                    print("[提示]: 不支持的设备类型，跳过super密码验证")
                                    result_text = "1[成功]: 测试正常"
                                    success_count += 1
                                    self.tree.item(item_id, tags=("success",))

                                conn.find_prompt(delay_factor=2)
                                print(f"[调试]：特权模式提示符为: {conn.find_prompt()}")

                            else:
                                print(f"处理Super密码逻辑，未-找到super密码")
                                # self.log(f"[提示]：设备未找到Super密码 {device['hostname']} ")
                                output = conn.send_command_timing("screen disable")
                                outputs.append(output)
                                output = conn.send_command_timing("display version")
                                outputs.append(output)
                                # conn.send_command_timing(super_password)
                                # output = conn.send_command_timing(super_password)

                                if any("Permission denied" in output for output in outputs):
                                    # self.root.after(0, self.update_tree_item, item_id, None, "[错误] 用户选择中止执行")
                                    self.log(f"[错误]：设备权限异常 {device['hostname']}")
                                    result_text = "0[错误]: 设备权限异常"
                                    failure_count += 1
                                    self.tree.item(item_id, tags=("failure",))
                                    connection_results.append(f"{device['hostname'].ljust(20)} | {result_text[:35]}")
                                    return

                                else:
                                    self.log(f"[提示]：连接到设备成功 {device['hostname']}")
                                    result_text = "1[成功]: 测试正常"
                                    print("[成功]: 连接成功，无需要Super密码")
                                    conn.find_prompt(delay_factor=2)
                                    success_count += 1
                                    self.tree.item(item_id, tags=("success",))
                                    connection_results.append(f"{device['hostname'].ljust(20)} | {result_text[:35]}")
                            # 测试成功
                            # self.root.after(0, self.update_tree_item, item_id, None, "[成功] 测试成功")
                            # 直接更新设备信息窗口的 test_result 列
                            self.root.after(0, self.update_tree_item, item_id, "1[成功]: 测试成功", None)
                            # success_count += 1
                            # 收集测试结果
                            connection_results.append(f"{device['hostname'].ljust(20)} | {result_text[:35]}")

            except (NetMikoTimeoutException, NetMikoAuthenticationException) as e:
                result_text = f"0[失败]: {str(e)}"  # 明确设置 result_text
                failure_count += 1
                self.tree.item(item_id, tags=("failure",))
                connection_results.append(f"{device['hostname'].ljust(20)} | {result_text[:35]}")
                self.log(f"[错误]：连接设备 {device['hostname']} 失败: {str(e)}")
            except Exception as e:
                result_text = f"0[错误]: {str(e)}"  # 明确设置 result_text
                failure_count += 1
                self.tree.item(item_id, tags=("failure",))
                connection_results.append(f"{device['hostname'].ljust(20)} | {result_text[:35]}")
                self.log(f"[错误]：连接设备 {device['hostname']} 时出错: {str(e)}")
            finally:
                if 'result_text' in locals():  # 检查 result_text 是否已定义
                    # self.root.after(0, self.update_tree_item, item_id, None, result_text)
                    # 直接更新设备信息窗口的 test_result 列
                    self.root.after(0, self.update_tree_item, item_id, result_text, None)
                else:
                    self.root.after(0, self.update_tree_item, item_id, None, "[错误]: 测试失败")
                self.root.after(0, lambda: self.update_progress(success_count, failure_count, len(selected_items)))


        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = []
            for item_id in selected_items:
                values = self.tree.item(item_id, "values")
                hostname = values[0]
                username = values[1]
                password = self.original_passwords.get(hostname, "")
                # ora_password = self.original_passwords.get(hostname, device["password"])
                # print(f"打印原始密码22{password}")
                connection_protocol = values[3]
                device_type = values[4]
                path = values[5]
                port = values[6] if len(values) > 6 else None
                device = {
                    "hostname": hostname,
                    "username": username,
                    "password": password,
                    "connection_protocol": connection_protocol,
                    "device_type": device_type,
                    "path": path,
                    "port": port
                }
                futures.append(executor.submit(test_device, device, item_id))

            for future in futures:
                if self.cancel_event.is_set():
                    break
                future.result()

        # 发送通知
        if connection_results:
            title = "Login 测试结果"
            content = "\n".join(["主机名      | 测试结果\n"] + connection_results)
            self.log(f"[提示]：Login 测试通知内容:\n {content}")
            self.send_server_chan_notification(title, content, test_type="Login测试")

        self.status_label.config(text="状态：Login 测试完成")
        self.test_button.config(state=tk.NORMAL)
        self.cancel_button.config(state=tk.NORMAL)
        self.ping_test_button.config(state=tk.NORMAL)
        self.execute_button.config(state=tk.NORMAL)

        self.log("[成功]：Login 测试完成")

    # 添加一个方法来显示命令执行模式选择对话框
    def show_command_mode_dialog(self):
        mode_window = tk.Toplevel(self.root)
        mode_window.title("命令执行模式设置")
        mode_window.geometry("300x200")
        self.center_window(mode_window)

        # 创建提示标签
        tk.Label(mode_window, text="选择命令执行模式:").pack(pady=10)

        # 创建单选按钮
        self.mode_var = tk.StringVar(value=self.execution_mode)
        tk.Radiobutton(mode_window, text="自适应模式", variable=self.mode_var,
                       value="send_command_expect").pack()
        tk.Radiobutton(mode_window, text="定时模式", variable=self.mode_var,
                       value="send_command_timing").pack()

        # 创建按钮框架
        button_frame = tk.Frame(mode_window)
        button_frame.pack(fill=tk.X, padx=10, pady=10)

        # 创建确认和取消按钮，并使用 grid 布局居中对齐
        confirm_button = tk.Button(button_frame, text="确定", command=lambda: self.save_mode_and_close(mode_window))
        cancel_button = tk.Button(button_frame, text="取消", command=mode_window.destroy)

        confirm_button.grid(row=0, column=0, padx=5)
        cancel_button.grid(row=0, column=1, padx=5)

        # 设置按钮框架的列权重，使按钮居中对齐
        button_frame.columnconfigure(0, weight=1)
        button_frame.columnconfigure(1, weight=1)

    # 添加一个方法来保存选择的模式并关闭对话框
    def save_mode_and_close(self, window):
        self.execution_mode = self.mode_var.get()
        if self.execution_mode == "send_command_expect":
            # self.log(f"[提示]：已设置命令执行模式为：{self.execution_mode}")
            self.log(f"[提示]：已设置为自适应模式")
        else:
            self.log(f"[提示]：已设置为定时模式")
        window.destroy()

    def run_commands_scheduled(self):
        # 定时任务执行脚本时不需要弹窗确认
        title = "定时任务触发"
        content = "定时任务执行脚本已触发，开始运行定时任务。"
        # self.log(f"[提示]：即将运行定时任务:\n {content}")
        self.send_server_chan_notification(title, content, test_type="定时任务")

        # messagebox.showinfo("任务执行", "执行命令脚本（定时任务触发）")
        self.log("[警告]：执行命令脚本（定时任务触发）")
        self.log_audit("[警告]：执行命令脚本", "执行命令脚本（定时任务触发）")
        self.run_commands()

    def start_command_execution(self):
        self.execute_button.config(state=tk.DISABLED)
        if not self.is_executing:
            # 弹出确认对话框
            self.log("[提示]：点击执行脚本")
            self.log_audit("[操作]：执行脚本", "用户点击执行脚本按钮")

            confirm = messagebox.askyesno("确认执行", "确定要执行脚本吗？", default=messagebox.NO)

            if confirm:
                self.log("[警告]：用户已点击确认脚本执行")
                self.log_audit("[操作]：执行脚本？", "用户已点击确认脚本执行")
            else:
                # self.log("[警告]：用户已点击取消脚本执行")
                self.log_audit("[操作]：执行脚本？", "用户已点击取消脚本执行")
                self.execute_button.config(state=tk.NORMAL)
                return

            self.cancel_event.clear()
            # self.execute_button.config(state=tk.DISABLED)
            self.cancel_button.config(state=tk.NORMAL)

            self.test_button.config(state=tk.DISABLED)

            self.ping_test_button.config(state=tk.NORMAL)

            # 创建一个线程来执行延迟后的命令
            def delayed_run_commands():
                # 延迟5秒
                # self.log("[提示]: 暂停5秒等待")
                self.log("[警告]：若选已择设备，5秒后开始脚本执行…")
                time.sleep(5)
                self.log("[提示]：脚本开始执行")
                self.is_executing = True
                self.execute_button.config(text="暂停执行")
                self.execute_button.config(state=tk.NORMAL)
                self.cancel_button.config(state=tk.NORMAL)
                # self.log("[提示]：脚本开始执行")
                # 启动命令执行
                self.run_commands()
                # self.toggle_execution()
            # self.log(f"打印变更状态{self.is_executing}")

            threading.Thread(target=delayed_run_commands, daemon=True).start()

        elif self.is_paused:
            # 继续执行脚本
            self.log("[提示]：用户点击继续执行脚本")
            self.log_audit("[操作]：暂停脚本", "用户点击继续执行脚本按钮")

            confirm = messagebox.askyesno("继续执行", "确定要继续执行脚本吗？", default=messagebox.NO)

            if confirm:
                self.is_paused = False
                self.execute_button.config(text="暂停执行")
                self.execute_button.config(state=tk.NORMAL)
                self.log("[提示]：已继续执行脚本")
                self.status_label.config(text="状态：正在执行命令...")
            else:
                self.is_paused = True
                self.log("[警告]：用户未确认取消脚本执行")
                self.execute_button.config(state=tk.NORMAL)
                self.log_audit("[操作]：取消暂停脚本？", "用户未确认取消脚本执行")
                return

        else:
            # 暂停执行脚本
            confirm = messagebox.askyesno("暂停执行", "确认暂停脚本吗？", default=messagebox.NO)

            if confirm:
                self.log("[警告]：用户已点击确认脚本暂停")
                self.log_audit("[操作]：暂停脚本？", "用户已点击确认脚本暂停")
                self.status_label.config(text="状态：暂停执行命令...")
            else:
                # self.log("[警告]：用户已点击取消脚本执行")
                self.log_audit("[操作]：暂停脚本？", "用户已点击取消脚本暂停")
                self.execute_button.config(text="暂停执行")
                self.execute_button.config(state=tk.NORMAL)
                return

            self.is_paused = True
            self.execute_button.config(text="继续执行")
            self.execute_button.config(state=tk.NORMAL)
            self.log("[提示]：已暂停执行下一条脚本")

    def manage_super_password_chars(self):
        # 创建一个新的窗口用于管理Super密码检测字符
        self.super_password_chars_window = tk.Toplevel(self.root)
        self.super_password_chars_window.title("管理Super密码检测字符")
        self.super_password_chars_window.geometry("400x300")
        self.center_window(self.super_password_chars_window)
        self.super_password_chars_window.grab_set()  # 使窗口获得焦点，防止用户与其他窗口交互

        # 创建一个标签和文本框用于显示和编辑Super密码检测字符
        tk.Label(self.super_password_chars_window, text="Super密码检测字符（每行一个）:").pack(pady=10)
        self.super_password_chars_text = tk.Text(self.super_password_chars_window, height=10, width=40)
        self.super_password_chars_text.pack(fill=tk.BOTH, expand=True)

        # 加载当前的Super密码检测字符
        config_file = os.path.join(self.config_dir, "super_password_chars.json")
        if os.path.exists(config_file):
            try:
                with open(config_file, "r") as f:
                    self.super_password_chars = json.load(f)
            except Exception as e:
                self.super_password_chars = ["success", "successed", "成功", "privilege", "network-admin","has not been set","level is 3"]
                self.log(f"[错误]：加载Super密码检测字符失败: {str(e)}")
        else:
            self.super_password_chars = ["success", "successed", "成功", "network-admin","has not been set","level is 3"]

        # 将当前字符显示在文本框中
        for char in self.super_password_chars:
            self.super_password_chars_text.insert(tk.END, char + "\n")

        # 创建保存按钮，并在点击后关闭窗口
        save_button = tk.Button(
            self.super_password_chars_window,
            text="保存",
            command=self.save_and_close_super_password_chars
        )
        save_button.pack(pady=10)

    def save_and_close_super_password_chars(self):
        # 获取用户输入的Super密码检测字符
        new_chars = [
            line.strip()
            for line in self.super_password_chars_text.get("1.0", tk.END).split("\n")
            if line.strip()
        ]
        # 保存到配置文件
        config_file = os.path.join(self.config_dir, "super_password_chars.json")
        try:
            with open(config_file, "w") as f:
                json.dump(new_chars, f)
            self.super_password_chars = new_chars
            self.log(f"[成功]：Super密码检测字符已保存{self.super_password_chars}")
            self.log_audit("[操作]：保存Super密码检测字符", f"用户已保存Super密码检测字符：{new_chars}")
            # 关闭窗口
            self.super_password_chars_window.destroy()
        except Exception as e:
            self.log(f"[错误]：保存Super密码检测字符失败: {str(e)}")
            messagebox.showerror("错误", f"保存Super密码检测字符失败: {str(e)}")


    def load_super_password_chars(self):
        config_file = os.path.join(self.config_dir, "super_password_chars.json")
        if os.path.exists(config_file):
            try:
                with open(config_file, "r") as f:
                    self.super_password_chars = json.load(f)
            except Exception as e:
                self.super_password_chars = ["success",  "successed", "成功", "network-admin","has not been set","level is 3"]
        else:
            self.super_password_chars = ["success",  "successed", "成功", "network-admin","has not been set","level is 3"]

    def is_super_password_privilege(self, output):
        # 检查是否进入特权模式
        print(f"[调试]：Super密码检测字符：{self.super_password_chars}")
        if any(char in output for char in self.super_password_chars):
            return True
        return False

    def manage_error_chars(self):
        error_chars_window = tk.Toplevel(self.root)
        error_chars_window.title("管理错误字符规则")
        error_chars_window.geometry("400x300")

        tk.Label(error_chars_window, text="错误字符规则（每行一个）:").pack(pady=10)
        self.error_chars_text = tk.Text(error_chars_window, height=10, width=40)
        self.error_chars_text.pack(pady=10)

        # 加载当前错误字符规则
        self.error_chars_text.insert(tk.END, "\n".join(self.error_chars))

        # 创建保存按钮，并在点击后关闭窗口
        save_button = tk.Button(error_chars_window, text="保存",
                                command=lambda: self.save_and_close(error_chars_window))
        save_button.pack(pady=10)

    def save_and_close(self, window):
        # 获取用户输入的错误字符规则
        new_error_chars = self.error_chars_text.get("1.0", tk.END).strip().split("\n")
        # 过滤掉空行
        new_error_chars = [char for char in new_error_chars if char.strip()]
        # 更新错误字符规则
        self.error_chars = new_error_chars
        # 自动保存错误字符规则到JSON文件
        self.save_error_chars_to_json(self.error_chars)  # 调用此方法保存设置
        # 关闭窗口
        window.destroy()

    def save_error_chars_to_json(self,error_chars):
        # 保存错误字符规则到JSON文件
        config_dir = os.path.join(os.getcwd(), "config")
        os.makedirs(config_dir, exist_ok=True)
        error_chars_file = os.path.join(config_dir, "error_chars.json")

        try:
            with open(error_chars_file, "w") as f:
                json.dump(self.error_chars, f, indent=4)
            # self.log("[成功]：错误字符规则已保存到: {}".format(error_chars_file))
            self.log(f"[成功]：匹配字符已设置为{error_chars}")
        except Exception as e:
            self.log("[错误]：保存错误字符规则失败: {}".format(str(e)))

    def on_button_click(self, window, result):
        self.error_dialog_result = result
        window.destroy()

    def set_retry_count(self):
        retry_count = simpledialog.askinteger("设置超时", "请输入超时时间:", initialvalue=self.retry_count)
        if retry_count is not None:
            self.retry_count = retry_count
            self.log(f"[提示]：脚本执行超时已设置为{retry_count}秒")

    def show_error_dialog(self, hostname, command, output):
        # 创建弹窗变量
        self.error_dialog_result = None
        self.log(f"[提示]：命令执行结果检测错误的匹配规则在设置中修改，默认为:Unrecognized，Ambiguous，")

        # 创建弹窗
        error_window = tk.Toplevel(self.root)
        error_window.title("命令执行错误")
        error_window.geometry("800x600")

        # 创建标题区域
        title_frame = tk.Frame(error_window)
        title_frame.pack(fill=tk.X, padx=10, pady=10)

        title_label = tk.Label(
            title_frame,
            text="请选择处理（错误匹配规划可在设置中修改）",
            font=("Arial", 12, "bold"),
            fg="red"
        )
        title_label.pack(side=tk.LEFT)

        # 创建提示信息区域
        prompt_frame = tk.Frame(error_window)
        prompt_frame.pack(fill=tk.X, padx=10, pady=5)

        prompt_label = tk.Label(
            prompt_frame,
            text=f"设备 {hostname} 执行命令 '{command}' 时出现错误。\n\n是否继续执行后续命令？",
            font=("Arial", 10)
        )
        prompt_label.pack(side=tk.LEFT)

        # 创建一个文本框用于显示错误信息，并添加滚动条
        output_frame = tk.Frame(error_window)
        output_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        error_text = ScrolledText(output_frame, wrap=tk.WORD)
        error_text.pack(fill=tk.BOTH, expand=True)
        error_text.insert(tk.END, f"{output}")
        error_text.configure(state='disabled')  # 设置为只读

        # 创建按钮框架并居中按钮
        button_frame = tk.Frame(error_window)
        button_frame.pack(fill=tk.X, padx=10, pady=10)

        # 添加“继续”和“取消”按钮，并居中对齐
        yes_button = tk.Button(button_frame, text="继续", command=lambda: self.on_button_click(error_window, True))
        no_button = tk.Button(button_frame, text="中断", command=lambda: self.on_button_click(error_window, False))
        yes_button.pack(side=tk.LEFT, padx=5, expand=True)  # expand=True 让按钮在框架中均匀分布
        no_button.pack(side=tk.RIGHT, padx=5, expand=True)

        # 确保按钮框架的列权重相同，以便按钮可以居中对齐
        button_frame.grid_columnconfigure(0, weight=1)
        button_frame.grid_columnconfigure(1, weight=1)


        # 等待用户交互
        error_window.lift()  # 确保弹窗显示在最前端
        error_window.grab_set()  # 阻止用户与主窗口交互
        error_window.wait_window()  # 等待弹窗关闭

        # 返回用户选择的结果
        return self.error_dialog_result

    def manage_prompt_patterns(self):
        # 创建一个新的窗口用于管理提示符
        self.prompt_window = tk.Toplevel(self.root)
        self.prompt_window.title("管理提示符")
        self.prompt_window.geometry("400x300")

        # 创建一个标签和文本框用于显示和编辑提示符
        tk.Label(self.prompt_window, text="当前提示符：").pack(pady=10)
        self.prompt_text = tk.Text(self.prompt_window, height=10, width=40)
        self.prompt_text.pack(fill=tk.BOTH, expand=True)

        # 加载当前的提示符
        self.load_prompt_patterns()

        # 创建按钮用于保存提示符
        save_button = tk.Button(self.prompt_window, text="保存提示符", command=self.save_and_close_prompt_window)
        save_button.pack(pady=10)

    def save_and_close_prompt_window(self):
        # 保存提示符并关闭窗口
        self.save_patterns_to_json("prompt_patterns")
        print(f"[调试]: 已保存错误提示符")
        if hasattr(self, 'prompt_window') and self.prompt_window:
            self.prompt_window.destroy()

    def load_prompt_patterns(self):
        # 从配置文件加载提示符
        config_file = os.path.join(self.config_dir, "prompt_patterns.json")
        default_patterns = [r'>', r'#', r']', r'}', r'$']  # 默认提示符
        if os.path.exists(config_file):
            try:
                with open(config_file, "r") as f:
                    self.prompt_patterns = json.load(f)
            except Exception as e:
                self.log(f"[错误]：加载提示符时出错: {str(e)}")
                self.prompt_patterns = default_patterns
        else:
            self.prompt_patterns = default_patterns

        # 将提示符显示在文本框中
        self.prompt_text.delete(1.0, tk.END)
        for pattern in self.prompt_patterns:
            self.prompt_text.insert(tk.END, pattern + "\n")

    def save_patterns_to_json(self, pattern_type):
        # 保存提示符或错误字符到JSON文件
        if pattern_type == "prompt_patterns":
            # 从文本框获取提示符
            patterns = [line.strip() for line in self.prompt_text.get(1.0, tk.END).split("\n") if line.strip()]
            file_name = "prompt_patterns.json"
        elif pattern_type == "error_chars":
            patterns = self.error_chars
            file_name = "error_chars.json"
        else:
            self.log("[错误]：未知的模式类型")
            return

        config_file = os.path.join(self.config_dir, file_name)
        try:
            with open(config_file, "w") as f:
                json.dump(patterns, f)
            self.log(f"[成功]：输入提示符已设置为 {patterns} ")
            self.prompt_patterns = patterns  # 更新当前提示符列表
        except Exception as e:
            self.log(f"[错误]：输入提示符保存 {patterns} 时出错: {str(e)}")

    # 在初始化方法中添加加载提示符的代码
    def initialize_settings(self):
        # 其他初始化代码
        self.prompt_patterns = []
        self.load_prompt_patterns()
        # 加载Super密码检测字符

    def run_commands(self):
        selected_items = self.tree.selection()
        if not selected_items:
            self.log("[错误]：执行操作请先选择设备")
            messagebox.showwarning("提示", "请先选择设备")
            self.execute_button.config(state=tk.NORMAL)
            self.cancel_button.config(state=tk.DISABLED)
            self.test_button.config(state=tk.NORMAL)

            self.is_executing = False
            self.is_paused = False
            self.execute_button.config(text="执行脚本")

            return

        devices = {}
        for item_id in selected_items:
            values = self.tree.item(item_id, "values")
            hostname = values[0]
            username = values[1]
            password = self.original_passwords.get(hostname, "")
            connection_protocol = values[3]
            device_type = values[4]
            path = values[5]
            port = values[6] if len(values) > 6 else None
            super_password = self.super_passwords.get(hostname, "")  # 获取Super密码
            devices[item_id] = {
                "hostname": hostname,
                "username": username,
                "password": password,
                "connection_protocol": connection_protocol,
                "device_type": device_type,
                "path": path,
                "port": port,
                "super_password": super_password
            }

        try:
            self.cmd_read_interval = int(self.cmd_interval_entry.get() or 3)
            self.log(f"[提示]：已设置命令执行间隔为{self.cmd_read_interval}秒")
        except ValueError:
            messagebox.showerror("错误", "命令执行间隔时间必须为整数")
            self.log("[错误]：命令执行间隔时间必须为整数，退出执行命令。")
            self.execute_button.config(state=tk.NORMAL)
            self.cancel_button.config(state=tk.DISABLED)
            self.status_label.config(text="状态：命令执行完成")
            self.is_executing = False
            self.is_paused = False
            self.execute_button.config(text="执行脚本")

            return

        self.progress_bar["maximum"] = len(selected_items)
        self.progress_bar["value"] = 0
        self.status_label.config(text="状态：正在执行命令...")

        success_count = 0
        failure_count = 0

        def run_commands_for_device(device, item_id):
            nonlocal success_count, failure_count
            if self.cancel_event.is_set():
                return
            connection_protocol = device["connection_protocol"].lower()
            device_type = device["device_type"].lower()
            port = int(float(device.get("port", "22")))
            hostname = device["hostname"]  # 获取设备的主机名
            super_password = device["super_password"]  # 获取Super密码
            password = device["password"]  # 获取密码
            # print(port)  # 输出: 22
            # port = device.get("port", None)
            # 添加一个字典来跟踪“不再提醒”的设备
            if not hasattr(self, "muted_devices"):
                self.muted_devices = set()

            try:
                command_file_path = self.command_file_path or device["path"]
                if not command_file_path or not os.path.exists(command_file_path):
                    self.root.after(0, self.update_tree_item, item_id, None, "[错误]: 命令文件不存在")
                    self.log(f"[错误]：连接设备 {hostname} 失败: 命令文件不存在")
                    failure_count += 1
                    self.root.after(0, lambda: self.update_progress(success_count, failure_count, len(devices)))
                    return

                with open(command_file_path, "r") as file:
                    commands = file.read().strip().split("\n")

                # 跳过空行和注释行
                commands = [cmd.strip() for cmd in commands if cmd.strip() and not cmd.startswith("#")]

                if connection_protocol == "telnet":
                    # 支持无用户名的Telnet连接
                    if not device["username"]:  # 如果用户名为空
                        connection = {
                            "device_type": "cisco_ios_telnet",
                            "host": device["hostname"],
                            "password": password,
                            "port": port or 23,
                            "timeout": self.ssh_telnet_timeout
                        }
                    else:
                        connection = {
                            "device_type": "cisco_ios_telnet",
                            "host": device["hostname"],
                            "username": device["username"],
                            "password": password,
                            "port": port or 23,
                            "timeout": self.ssh_telnet_timeout
                        }
                elif connection_protocol == "ssh":
                    if device_type == "h3c":
                        connection = {
                            "device_type": "hp_comware",
                            "host": device["hostname"],
                            "username": device["username"],
                            "password": device["password"],
                            "port": port or 22,
                            "timeout": self.ssh_telnet_timeout
                        }
                    elif device_type == "huawei":
                        connection = {
                            "device_type": "huawei",
                            "host": device["hostname"],
                            "username": device["username"],
                            "password": device["password"],
                            "port": port or 22,
                            "timeout": self.ssh_telnet_timeout
                        }
                    elif device_type == "cisco":
                        connection = {
                            "device_type": "cisco_ios",
                            "host": device["hostname"],
                            "username": device["username"],
                            "password": device["password"],
                            "port": port or 22,
                            "timeout": self.ssh_telnet_timeout
                        }
                    elif device_type == "linux":
                        connection = {
                            "device_type": "linux_ssh",
                            "host": device["hostname"],
                            "username": device["username"],
                            "password": device["password"],
                            "port": port or 22,
                            "timeout": self.ssh_telnet_timeout,
                            "global_delay_factor": 2
                        }
                    else:
                        connection = {
                            "device_type": "generic",
                            "host": device["hostname"],
                            "username": device["username"],
                            "password": device["password"],
                            "port": port or 22,
                            "timeout": self.ssh_telnet_timeout
                        }

                with (ConnectHandler(**connection) as conn):
                    #处理Super密码逻辑
                    if super_password and super_password != "":
                        print(f"处理Super密码逻辑，找到super密码 {device['hostname']}")
                        # prompt = conn.find_prompt()
                        if device_type in ["h3c", "huawei"]:
                            print(f"[调试]: 设备类型为h3c设备")
                            output = conn.send_command_timing("super")
                            # time.sleep(0.5)
                            print(f"[调试]: 打印super命令输出之后的结果-{output}")
                            # time.sleep(1)
                            output = conn.send_command_timing(super_password)
                            print(f"检测H3C-HUAWEI设备-到Super密码{super_password},发送Super密码")
                            # conn.find_prompt()  # 确保Super密码已正确输入并找到新的提示符
                            print(f"[调试]：发送Super密码后的结果：{output}")
                            # 验证Super密码是否正确
                            if self.is_super_password_privilege(output.lower()):
                                self.log(f"[成功]：Super密码验证成功 {hostname}")
                                print("[成功]: Super密码成功")
                                result_text = "1[成功]: 测试正常"
                                output = conn.send_command_timing("terminal length 0")
                                output = conn.send_command_timing("screen-length disable")
                                print(f"[调试]：发送禁止分页命令：terminal length 0和screen disable")
                                # success_count += 1
                                # conn.find_prompt()  # 确保Super密码已正确输入并找到新的提示符
                            else:
                                # self.log(f"[错误]：Super密码输入失败 {device['hostname']} ")
                                self.log(f"[错误]：Super密码错误,退出执行 {hostname}")
                                print(f"[错误]：Super密码错误,退出执行 {hostname}")
                                result_text = "0[错误]: Super密码错误,退出执行"
                                self.root.after(0, self.update_tree_item, item_id, None, result_text)
                                failure_count += 1
                                return

                        elif device_type in ["cisco"]:
                            output = conn.send_command_timing("enable")
                            # conn.send_command_timing(super_password)
                            output = conn.send_command_timing(super_password)
                            # conn.find_prompt()  # 确保Super密码已正确输入并找到新的提示符
                            print(f"检测CISCO设备-到Super密码{super_password},发送Super密码")
                            print(f"[调试]：发送Super密码后的结果：{output}")
                            # 验证Super密码是否正确
                            if self.is_super_password_privilege(output.lower()):
                                self.log(f"[成功]：Super密码验证成功 {hostname}")
                                print("[成功]: Super密码成功")
                                result_text = "1[成功]: 测试正常"
                                # success_count += 1
                                # conn.find_prompt()  # 确保Super密码已正确输入并找到新的提示符
                            else:
                                self.log(f"[错误]：Super密码失败,退出执行 {hostname}")
                                print(f"[错误]：Super密码失败,退出执行 {hostname}")
                                result_text = "[错误]: Super密码失败,退出执行"
                                self.root.after(0, self.update_tree_item, item_id, None, result_text)
                                failure_count += 1
                                return


                        else:
                            self.log(f"[提示]：不支持的设备类型，跳过super密码{device['hostname']}")
                            print("[提示]: 不支持的设备类型，跳过super密码验证")
                            result_text = "1[成功]: 测试正常"
                            success_count += 1
                            self.tree.item(item_id, tags=("success",))

                        #验证结果
                        conn.find_prompt(delay_factor=2)
                        print(f"[调试]：特权模式提示符为: {conn.find_prompt()}")

                    else:
                        print(f"处理Super密码逻辑，未找到super密码")
                        outputs = []
                        # self.log(f"[提示]：设备未找到Super密码 {device['hostname']} ")
                        output = conn.send_command_timing("terminal length 0")
                        output = conn.send_command_timing("screen disable")                      
                        outputs.append(output)
                        output = conn.send_command_timing("display version")
                        outputs.append(output)
                        # outputs.append("")

                        # conn.send_command_timing(super_password)
                        # output = conn.send_command_timing(super_password)

                        if any("Permission denied" in output for output in outputs):
                            # self.root.after(0, self.update_tree_item, item_id, None, "[错误] 用户选择中止执行")
                            self.log(f"[错误]：设备权限异常 {device['hostname']}")
                            result_text = "0[错误]: Super密码错误"
                            self.root.after(0, self.update_tree_item, item_id, None, result_text)
                            failure_count += 1
                            return

                        else:
                            self.log(f"[提示]：连接到设备成功 {hostname}")
                            result_text = "[成功]: 连接成功"
                            print("[成功]: 连接成功，无需要Super密码")
                            conn.find_prompt(delay_factor=2)
                            success_count += 1
                            self.tree.item(item_id, tags=("success",))

                    # 短暂等待，确保设备响应完成
                    time.sleep(0.2)  # 显式增加延迟
                    conn.find_prompt()
                    outputs = []
                    execution_times = []
                    start_time = time.time()
                    start_cmd_time = time.time()
                    for command in commands:
                        if self.cancel_event.is_set():
                            # self.log(f"执行脚本前检查变量{self.is_paused}")
                            break

                        if self.is_paused:
                            while self.is_paused and not self.cancel_event.is_set():
                                # self.log(f"循环暂停，执行脚本前检查变量{self.is_paused}")
                                time.sleep(1)
                            if self.cancel_event.is_set():
                                break

                        # 根据模式执行命令
                        if self.execution_mode == "send_command_expect":
                            output = conn.send_command_expect(command, delay_factor=20,max_loops=200)
                            # self.log(f"[调试063][提示]：执行自适应模式send_command_expect")
                            # 判断命令是否执行完成

                            # while not self.is_command_output_complete(command, conn, output):
                            #     if self.cancel_event.is_set():
                            #         break
                            #     time.sleep(0.5)
                            #     # output += conn.read_channel()


                        else:  # send_command_timing 模式
                            output = conn.send_command_timing(command, delay_factor=20,max_loops=200)
                            # self.log(f"[调试063][提示]：执行定时模式send_command_timing")
                            # 判断命令是否执行完成

                            # while not self.is_command_output_complete(command, conn, output):
                            #     if self.cancel_event.is_set():
                            #         break
                            #     time.sleep(0.5)
                            #     # output += conn.read_channel()

                        outputs.append(output)
                        execution_times.append(time.time() - start_cmd_time)

                        if self.error_chars and len(self.error_chars) > 2:
                            err_length = len(self.error_chars)
                            print(f"[调试]: 错误提示符号非空,进行检查{err_length}")
                            if any(char in output for char in self.error_chars):
                                # 检查是否在此设备的“不再提醒”列表中
                                confirm = False
                                # error_dialog = "null"
                                if hostname not in self.muted_devices:
                                    # 弹出确认对话框
                                    confirm = self.show_error_dialog(hostname, command, output)
                                    print("[调试]: 执行错误判断，结束，调试")
                                    if confirm is None:  # 用户点击“取消”按钮
                                        # 添加到“不再提醒”列表
                                        self.muted_devices.add(hostname)
                                        # self.root.after(0, self.update_tree_item, item_id, None, "[错误] 用户选择不再提醒")
                                        self.log(f"[错误]：{hostname}:{command}: 命令脚本错误,用户选择不再提醒执行!")
                                        self.last_error_command[device["hostname"]] = command

                                        self.last_error_output[device["hostname"]] = output  # 记录错误输出
                                        error_msg = f"{command}: 命令错误,继续执行脚本! \n"
                                        self.root.after(0, self.update_tree_item, item_id, None, error_msg)
                                        self.log(f"[错误]: {device['hostname']}: {error_msg} \n")
                                        outputs.append("")
                                        execution_times.append(0)
                                        outputs[-1] = output
                                    elif not confirm:  # 用户点击“否”
                                        # self.root.after(0, self.update_tree_item, item_id, None, "[错误] 用户选择中止执行")
                                        self.log(f"[错误]：{hostname}:{command}: 命令脚本错误,用户选择中止执行!")
                                        # self.log(f"[提示]：用户选择不再提醒此设备 {hostname}")
                                        self.last_error_command[device["hostname"]] = command
                                        # outputs[-1] = output
                                        self.last_error_output[device["hostname"]] = output  # 记录错误输出
                                        error_msg = f"{command}: 命令错误,退出执行脚本！ \n"
                                        self.root.after(0, self.update_tree_item, item_id, None, error_msg)
                                        self.log(f"[错误]: {device['hostname']}: {error_msg} \n")
                                        # outputs.append("")
                                        execution_times.append(0)
                                        outputs[-1] = output
                                        break
                                    else:  # 用户点击“是”
                                        # self.root.after(0, self.update_tree_item, item_id, None, "[警告] 用户选择继续执行")
                                        # self.log(f"[警告]：用户选择继续执行设备 {hostname} 的命令")
                                        # self.log(f"[错误]：命令错误，用户选择继续执行脚本 {hostname}")
                                        self.log(f"[错误]：{hostname}:{command}: 命令脚本错误,用户选择继续执行!")
                                        self.last_error_command[device["hostname"]] = command
                                        # outputs[-1] = output
                                        self.last_error_output[device["hostname"]] = output  # 记录错误输出
                                        error_msg = f"{command}: 命令错误,继续执行脚本! \n"
                                        self.root.after(0, self.update_tree_item, item_id, None, error_msg)
                                        self.log(f"[错误]: {device['hostname']}: {error_msg} \n")
                                        # outputs.append("")
                                        execution_times.append(0)
                                        outputs[-1] = output
                                else:
                                    self.root.after(0, self.update_tree_item, item_id, None, "[警告]: 跳过错误提醒")
                                    # self.log(f"[警告]：跳过设备 {hostname} 的错误提醒")
                                    self.log(f"[错误]：{hostname}:{command}: 跳过错误提醒,继续执行!")
                                    self.last_error_command[device["hostname"]] = command
                                    # outputs[-1] = output
                                    self.last_error_output[device["hostname"]] = output  # 记录错误输出
                                    error_msg = f"{command}: 命令错误,跳过错误提醒,继续执行! \n"
                                    self.root.after(0, self.update_tree_item, item_id, None, error_msg)
                                    self.log(f"[错误]: {device['hostname']}: {error_msg} \n")
                                    # outputs.append("")
                                    execution_times.append(0)
                                    outputs[-1] = output

                            else :
                                self.log(f"[成功]: {device['hostname']}: {command} \n")
                                self.root.after(0, self.update_tree_item, item_id, None, "[执行]: 执行脚本中，请等待")
                                # outputs.append("")
                                execution_times.append(0)
                                outputs[-1] = output
                                pass

                            # 执行脚本间隔
                            # time.sleep(self.cmd_read_interval)

                        else :
                            self.log(f"[成功]: {device['hostname']}: {command} \n")
                            self.root.after(0, self.update_tree_item, item_id, None, "[执行]: 执行脚本中，请等待")
                            # outputs.append("")
                            execution_times.append(0)
                            outputs[-1] = output
                            pass
                        # 执行脚本间隔
                        time.sleep(self.cmd_read_interval)

                    # 判断每条命令的执行结果
                    results = []
                    for output in outputs:
                        if not output.strip():
                            results.append("错误")
                        elif any("失败" in output or "错误" in output for output in outputs):
                            results.append("失败")
                        else:
                            results.append("成功")

                    # 保存命令执行历史
                    self.save_command_history(device["hostname"], commands, results, execution_times)


                    # 确保 commands 和 outputs 长度一致
                    if len(commands) != len(outputs):
                        # 如果命令执行出错，outputs 可能比 commands 短，填充空字符串
                        while len(outputs) < len(commands):
                            outputs.append("Null")
                        # 如果命令执行出错，commands 可能比 outputs 短，截断 commands
                        if len(outputs) > len(commands):
                            commands = commands[:len(outputs)]


                    if not self.cancel_event.is_set() and not self.is_paused:
                        self.save_command_result(device["hostname"], commands, outputs)
                        self.root.after(0, self.update_tree_item, item_id, None, "[成功]: 脚本执行成功")
                        success_count += 1
                        self.tree.item(item_id, tags=("success",))

                    elif self.cancel_event.is_set() or self.is_paused:
                        self.save_command_result(device["hostname"], commands, outputs)
                        self.root.after(0, self.update_tree_item, item_id, None, "[成功]: 脚本执行成功")
                        failure_count += 1
                        self.tree.item(item_id, tags=("failure",))
                    else:
                        failure_count += 1
                        self.tree.item(item_id, tags=("failure",))
            except Exception as e:
                error_msg = f"[失败] {str(e)}"
                self.root.after(0, self.update_tree_item, item_id, None, error_msg)
                failure_count += 1
                self.tree.item(item_id, tags=("failure",))
            finally:
                self.root.after(0, lambda: self.update_progress(success_count, failure_count, len(devices)))
                failure_count += 1
                self.tree.item(item_id, tags=("failure",))


        threads = []
        for item_id, device in devices.items():
            thread = threading.Thread(target=run_commands_for_device, args=(device, item_id))
            threads.append(thread)
            thread.start()

        for thread in threads:
            while self.is_paused and not self.cancel_event.is_set():
                print(f"[调试]: 循环暂停，执行脚本前检查变量{self.is_paused}")
                time.sleep(1)
            if self.cancel_event.is_set():
                break

            thread.join()

        self.status_label.config(text="状态：脚本执行完成")
        self.execute_button.config(state=tk.NORMAL)

        self.cancel_button.config(state=tk.NORMAL)
        self.test_button.config(state=tk.NORMAL)
        self.ping_test_button.config(state=tk.NORMAL)
        self.is_executing = False
        self.is_paused = False
        self.execute_button.config(text="执行脚本")
        if hasattr(self, "muted_devices"):
            self.muted_devices.clear()

        self.send_server_chan_notification("Command 脚本执行完成", "所有设备脚本已执行完成，请到软件查看结果。",
                                           test_type="Command测试")
        self.log("[成功]：所有设备脚本执行完成。")
        # 在命令执行完成后调用通知
        self.show_notification("脚本执行完成", 2)
        self.log_audit("[操作]：脚本执行", "所有设备脚本已执行完成，请到软件查看结果")


    def is_command_output_complete(self, command, conn, output):
        """
            检测命令输出是否完整，通过两种方法同时检测：
            1. 正则表达式匹配提示符
            2. Netmiko 的 find_prompt 方法

            当其中一种方法检测到命令完成时立即返回成功，
            如果两种方法均在超时时间内未检测到命令完成则返回失败
            """
        # 初始化标志
        prompt_detected = False
        find_prompt_success = False
        prompt_patterns = self.prompt_patterns or [r'>', r'#', r']', r'}', r'\$']
        # if self.cancel_event.is_set():
        #     return

        # 获取用户设置的超时时间（秒）
        timeout = int(self.retry_count or 120)  # 默认为120秒
        checks_count = 1

        # 结合超时机制和现有判断逻辑
        start_time = time.time()
        # self.log(f"[调试063][提示]：时间{start_time}")
        while checks_count <= timeout:
            # 检查取消事件
            if self.cancel_event.is_set():
                self.log("[警告]：任务已被取消")
                return False

            # 读取通道输出
            output += conn.read_channel()

            # 初始化每轮检测的标志
            prompt_detected = False
            find_prompt_success = False

            # 检测提示符
            # prompt_patterns = [r'>', r'#', r']', r'}', r'\$']

            # 方法一：正则表达式匹配提示符
            for pattern in prompt_patterns:
                if self.cancel_event.is_set():
                    return
                try:
                    if re.search(pattern, output):
                        # prompt_detected = False
                        # self.log(f"[调试063][1111提示]：提示符匹配：prompt_patterns = True,匹配到回显字符{pattern}, 字符集为{prompt_patterns}")
                        # self.log(f"[调试063][555执行]：耗时{time.time() - start_time},第{checks_count}次")
                        prompt_detected = True
                        break  # 退出循环，已找到提示符

                except Exception as e:
                    # self.log(f"[调试063][222错误]：提示符匹配检测命令完成时出错: {str(e)}")
                    # self.log(f"[调试063][555执行]：耗时{time.time() - start_time},第{checks_count}次")
                    prompt_detected = False

                # 使用 Netmiko 的 find_prompt 方法
            if not prompt_detected:
                try:
                    conn.find_prompt()
                    # find_prompt_success = False
                    # self.log(f"[调试063][222提示]：函数检测：find_prompt_success = True,返回函数检测状态{conn.find_prompt()}")
                    # self.log(f"[调试063][555执行]：耗时{time.time() - start_time},第{checks_count}次")
                    find_prompt_success = True
                    # continue
                    # return True
                except Exception as e:
                    # self.log(f"[调试063][222错误]：函数检测命令完成时出错: {str(e)}")
                    # self.log(f"[调试063][555执行]：耗时{time.time() - start_time},第{checks_count}次")
                    find_prompt_success = False

                # 如果提示符检测和 find_prompt 都成功，则返回 True
            if prompt_detected or find_prompt_success:
                # self.log(f"[调试063][333提示]：ind_prompt_success命令输出完成，返回return true")
                # self.log(f"[调试063][555执行]：耗时{time.time() - start_time},第{checks_count}次")
                return True

            # # 检查是否已经超时
            # # elapsed_time = time.time() - start_time
            # if checks_count*2 >= timeout:
            #     self.log(f"[警告]：命令 '{command}' 执行超时，已达到最大等待时间 {timeout} 秒")
            #     prompt_detected = False
            #     find_prompt_success = False
            #     return False  # 超时

            # 短暂等待，避免高 CPU 占用
            time.sleep(0.3)
            checks_count += 1

        # 在 while 循环结束后检查是否最后失败,如果超时时间内未同时满足两个条件，记录超时日志并返回 False
        if checks_count*2 >= timeout:
            self.log(f"[警告]：{command} 命令执行超时{self.retry_count}秒，执行下一条命令")
            return False

        # self.log(f"[警告]：命令 '{command}' 超时，未检测到提示符变化，已达到最大等待时间 {timeout} 秒")
        # return False  # 超时

    #错误命令弹窗
    def show_error_with_scrollbar(hostname, command, output):
        # 创建一个自定义弹窗
        error_window = tk.Toplevel(app.root)
        error_window.title("命令执行错误")
        error_window.geometry("800x6.00")
        # app.center_window(error_window)  # 如果需要居中显示，请取消注释

        # 创建一个ScrolledText组件用于显示错误信息
        text_frame = tk.Frame(error_window)
        text_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        error_text = ScrolledText(text_frame, wrap=tk.WORD)
        error_text.pack(fill=tk.BOTH, expand=True)
        error_text.insert(tk.END,
                          f"设备 {hostname} 执行命令 '{command}' 时出现错误：\n{output}\n\n是否继续执行后续命令？")
        error_text.configure(state='disabled')  # 设置为只读

        # 添加按钮框架
        button_frame = tk.Frame(error_window)
        button_frame.pack(fill=tk.X, padx=10, pady=10)

        # 模拟 messagebox.askyesno 的功能
        confirm = tk.BooleanVar(value=False)

        def on_yes():
            confirm.set(True)
            error_window.destroy()

        def on_no():
            confirm.set(False)
            error_window.destroy()

        yes_button = tk.Button(button_frame, text="继续", command=on_yes)
        yes_button.pack(side=tk.LEFT, padx=5)

        no_button = tk.Button(button_frame, text="取消", command=on_no)
        no_button.pack(side=tk.RIGHT, padx=5)

        # 等待窗口关闭
        error_window.wait_window()
        return confirm.get()



    def save_command_result(self, hostname, commands, outputs):
        # 清理 ANSI 转义序列
        ansi_escape = re.compile(r'(\x9B|\x1B\[)[0-?]*[ -/]*[@-~]')
        cleaned_outputs = [ansi_escape.sub('', output) for output in outputs]

        # 确保 commands 和 cleaned_outputs 长度一致
        if len(commands) != len(cleaned_outputs):
            max_length = max(len(commands), len(cleaned_outputs))
            commands = commands + [''] * (max_length - len(commands))
            cleaned_outputs = cleaned_outputs + [''] * (max_length - len(cleaned_outputs))

        # 保存命令执行历史
        # self.save_command_history(hostname, commands, cleaned_outputs, [])

        # 确保设备目录存在
        directory = os.path.join(self.selected_save_directory or self.default_device_log_dir, hostname)
        os.makedirs(directory, exist_ok=True)

        # 验证目录是否存在和可写
        if not os.path.exists(directory):
            self.log(f"[错误]：目录 {directory} 不存在，无法保存文件")
            return
        if not os.access(directory, os.W_OK):
            self.log(f"[错误]：目录 {directory} 不可写，无法保存文件")
            return

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # 保存为 Excel 文件
        df = pd.DataFrame({'命令': commands, '输出': cleaned_outputs})
        excel_filename = f"{hostname}_{timestamp}.xlsx"
        excel_path = os.path.join(directory, excel_filename)
        df.to_excel(excel_path, index=False)
        self.log(f"[提示]：脚本执行结果已保存为 Excel 文件：{excel_path}")

        # 验证 Excel 文件内容
        if os.path.exists(excel_path):
            # self.log(f"[调试]：Excel 文件大小：{os.path.getsize(excel_path)} 字节")
            # 读取并打印前几行数据
            try:
                test_df = pd.read_excel(excel_path)
                # self.log(f"[调试]：Excel 文件内容验证成功，前 5 行数据：\n{test_df.head()}")
            except Exception as e:
                self.log(f"[错误]：验证 Excel 文件内容失败：{str(e)}")

        # 保存为 TXT 文件
        txt_filename = f"{hostname}_{timestamp}.txt"
        txt_path = os.path.join(directory, txt_filename)
        with open(txt_path, "w", encoding='utf-8') as txt_file:
            for command, output in zip(commands, cleaned_outputs):
                txt_file.write(f"命令: {command}\n")
                txt_file.write(f"输出:\n{output.strip()}\n\n")
        self.log(f"[提示]：脚本执行结果已保存为TXT文本文件：{txt_path}")

        # 验证 TXT 文件内容
        if os.path.exists(txt_path):
            # self.log(f"[调试]：TXT 文件大小：{os.path.getsize(txt_path)} 字节")
            # 读取并打印前几行数据
            try:
                with open(txt_path, "r", encoding='utf-8') as txt_file:
                    content = txt_file.read()
                    # self.log(f"[调试]：TXT 文件内容验证成功，前 200 字符：\n{content[:200]}")
            except Exception as e:
                self.log(f"[错误]：验证 TXT 文件内容失败：{str(e)}")

        # 保存为 HTML 文件
        html_filename = f"{hostname}_{timestamp}.html"
        html_path = os.path.join(directory, html_filename)
        html_content = """
        <html>
        <head>
            <title>设备命令执行结果 - {hostname}</title>
            <style>
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
            </style>
        </head>
        <body>
            <h1>设备命令执行结果 - {hostname}</h1>
            <table>
                <tr>
                    <th>命令</th>
                    <th>输出</th>
                </tr>
                {table_rows}
            </table>
        </body>
        </html>
        """.format(
            hostname=hostname,
            table_rows="\n".join([
                f"<tr><td>{row['命令']}</td><td><pre>{row['输出']}</pre></td></tr>"
                for _, row in df.iterrows()
            ])
        )
        with open(html_path, "w", encoding='utf-8') as html_file:
            html_file.write(html_content)
        self.log(f"[提示]：脚本执行结果已保存为 HTML 文件：{html_path}")

        # 验证 HTML 文件内容
        if os.path.exists(html_path):
            # self.log(f"[调试]：HTML 文件大小：{os.path.getsize(html_path)} 字节")
            # 读取并打印前几行数据
            try:
                with open(html_path, "r", encoding='utf-8') as html_file:
                    content = html_file.read()
                    # self.log(f"[调试]：HTML 文件内容验证成功，前 200 字符：\n{content[:200]}")
            except Exception as e:
                self.log(f"[错误]：验证 HTML 文件内容失败：{str(e)}")

    def update_tree_item(self, item_id, test_result=None, execute_result=None):
        current_values = self.tree.item(item_id, "values")
        new_values = list(current_values)
        if test_result is not None:
            new_values[7] = test_result
        if execute_result is not None:
            new_values[8] = execute_result
        self.tree.item(item_id, values=new_values)

    def get_device_info(self):
        devices = {}
        for item_id in self.tree.get_children():
            values = self.tree.item(item_id, "values")
            hostname = values[0]
            username = values[1]
            password = self.original_passwords.get(hostname, "")
            connection_protocol = values[3]
            device_type = values[4]
            path = values[5]
            port = values[6] if len(values) > 6 else None
            device = {
                "hostname": hostname,
                "username": username,
                "password": password,
                "connection_protocol": connection_protocol,
                "device_type": device_type,
                "path": path,
                "port": port
            }
            devices[item_id] = device
        return devices

    def update_progress(self, success_count, failure_count, total_count):
        current_value = self.progress_bar["value"]
        max_value = self.progress_bar["maximum"]

        # 计算成功和失败的比例
        if total_count > 0:
            success_rate = success_count / total_count
            failure_rate = failure_count / total_count
        else:
            success_rate = 0
            failure_rate = 0

        # 根据任务执行结果设置进度条颜色
        if failure_rate == 1:
            self.progress_bar["style"] = "red.Horizontal.TProgressbar"
        elif failure_rate > 0:
            self.progress_bar["style"] = "yellow.Horizontal.TProgressbar"
        else:
            self.progress_bar["style"] = "green.Horizontal.TProgressbar"

        self.progress_bar["value"] = current_value + 1
        self.root.update_idletasks()

        # 重置静止检测计时器
        self.reset_static_timer()

    def reset_static_timer(self):
        # 取消之前的计时器
        if self.static_progress_timer is not None:
            self.static_progress_timer.cancel()

        # 重置静止检测计时器
        self.progress_last_update = time.time()
        self.static_progress_timer = threading.Timer(60.0, self.check_static_progress)
        self.static_progress_timer.daemon = True
        self.static_progress_timer.start()

    def check_static_progress(self):
        elapsed_time = time.time() - self.progress_last_update
        if elapsed_time >= 1800:  # 静止1800秒后启动跑灯效果
            self.start_breathing_effect()
            # 重新启动计时器
            self.reset_static_timer()



    def start_breathing_effect(self):
        if self.breathing_effect_active:
            return

        self.breathing_effect_active = True
        self.original_progress_color = self.progress_bar.cget("style")  # 记录原始颜色状态
        self.breathing_effect_thread = threading.Thread(target=self.breathing_effect)
        self.breathing_effect_thread.daemon = True
        self.breathing_effect_thread.start()

    def breathing_effect(self):
        # 模拟从0%到100%的跑灯效果
        for value in range(0, 101, 5):
            if not self.breathing_effect_active:
                break
            self.root.after(0, lambda v=value: self.update_progress_bar_value(v))
            time.sleep(0.1)

        # 模拟从100%到0%的跑灯效果
        for value in range(100, -1, -5):
            if not self.breathing_effect_active:
                break
            self.root.after(0, lambda v=value: self.update_progress_bar_value(v))
            time.sleep(0.1)

        # 跑灯效果完成后恢复原始颜色
        self.root.after(0, self.stop_breathing_effect)

    def update_progress_bar_value(self, value):
        if self.breathing_effect_active:
            self.progress_bar["value"] = value
            # 根据进度值切换颜色
            if value < 33:
                self.progress_bar["style"] = "green.Horizontal.TProgressbar"
            elif value < 66:
                self.progress_bar["style"] = "yellow.Horizontal.TProgressbar"
            else:
                self.progress_bar["style"] = "red.Horizontal.TProgressbar"

    def stop_breathing_effect(self):
        # self.breathing_effect_active = False
        # self.progress_bar["style"] = "green.Horizontal.TProgressbar"
        # self.progress_bar["value"] = 0

        self.breathing_effect_active = False
        if hasattr(self, 'original_progress_color'):
            self.progress_bar.configure(style=self.original_progress_color)  # 恢复原始颜色
        self.progress_bar["value"] = 0
        # self.log("保活进度栏闪烁停止")

    def task_completed(self):
        # 任务完成后停止跑灯效果
        self.stop_breathing_effect()
        #self.log("状态栏效果停止")
        if self.static_progress_timer is not None:
            self.static_progress_timer.cancel()


    def cancel_operation(self):
        # 弹出确认对话框
        confirm = messagebox.askyesno("确认取消", "确定要取消操作吗？", default=messagebox.NO)
        if not confirm:
            self.log("[警告]：取消操作已被取消")
            self.log_audit("[操作]：取消操作？", "用户点击取消操作但未点击确认")
            return

        self.cancel_event.set()
        self.cancel_button.config(state=tk.DISABLED)
        self.ping_test_button.config(state=tk.NORMAL)
        # self.test_button.config(state=tk.NORMAL)
        # self.execute_button.config(state=tk.NORMAL)
        self.status_label.config(text="状态：操作已取消")
        self.log("[警告]：用户点击取消操作，正在停止")
        self.log_audit("[操作]：操作取消", "用户点击并确认取消操作")


    def export_log(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".log", filetypes=[("Log files", "*.log")])
        self.log_audit("[操作]：导出日志", "用户点击导出操作日志按钮")
        if file_path:
            try:
                with open(self.log_file_path, 'r', encoding='utf-8') as source, \
                        open(file_path, 'w', encoding='utf-8') as target:
                    target.write(source.read())
                messagebox.showinfo("成功", "操作日志导出成功")
                self.log(f"[成功]：操作日志已导出至: {file_path}")
                self.log_audit("[操作]：导出日志", "用户已导出日志成功")
            except Exception as e:
                messagebox.showerror("错误", f"导出操作日志失败: {str(e)}")
                self.log(f"[错误]：导出操作日志失败: {str(e)}")

    def export_command_results_to_excel(self, zipf):
        # 创建一个字典来存储每个设备的命令执行结果
        all_results = {}
        self.log_audit("[操作]：导出报告", "用户点击导出报告按钮")

        # 遍历设备目录，获取每个设备的最新命令执行结果
        for item_id in self.tree.get_children():
            values = self.tree.item(item_id, "values")
            hostname = values[0]
            device_directory = os.path.join(self.default_device_log_dir, hostname)  # 使用 device-log 目录
            if os.path.exists(device_directory):
                latest_file = None
                latest_time = None
                for filename in os.listdir(device_directory):
                    if filename.endswith(".xlsx"):
                        file_path = os.path.join(device_directory, filename)
                        file_time = os.path.getmtime(file_path)
                        if latest_file is None or file_time > latest_time:
                            latest_file = file_path
                            latest_time = file_time
                if latest_file:
                    df = pd.read_excel(latest_file)
                    commands = df['命令'].tolist()
                    results = df['输出'].tolist()
                    all_results[hostname] = {cmd: res for cmd, res in zip(commands, results)}

        # 初始化 summary_df 为一个空的 DataFrame
        summary_df = pd.DataFrame()

        if all_results:
            # 创建一个新的Excel文件来汇总所有设备的命令执行结果
            summary_df = pd.DataFrame(index=all_results.keys())  # 行索引为设备IP
            # 遍历所有命令，确保每个命令都有对应的列
            all_commands = set()
            for results in all_results.values():
                all_commands.update(results.keys())
            for command in all_commands:
                summary_df[command] = summary_df.index.map(lambda ip: all_results[ip].get(command, ""))

            # 将汇总结果保存到Excel文件
            summary_excel_path = os.path.join(self.default_device_log_dir, "command_results_summary.xlsx")
            os.makedirs(self.default_device_log_dir, exist_ok=True)  # 确保目录存在
            summary_df.to_excel(summary_excel_path)
            # 将汇总Excel文件添加到ZIP
            zipf.write(summary_excel_path, os.path.basename(summary_excel_path))
            # 删除临时创建的汇总Excel文件
            os.remove(summary_excel_path)
            self.log("[成功]：命令执行结果汇总到一张Excel表并添加到ZIP文件")
            self.log_audit("[操作]：导出报告", "用户导出命令执行结果汇总成功")


    def generate_summary_html_report(self):
        # 创建一个字典来存储每个设备的命令执行结果
        all_results = {}
        # 遍历设备目录，获取每个设备的最新命令执行结果
        for item_id in self.tree.get_children():
            values = self.tree.item(item_id, "values")
            hostname = values[0]
            # 根据用户是否选择了保存目录，决定日志保存路径
            if self.selected_save_directory:
                device_directory = os.path.join(self.selected_save_directory, hostname)
            else:
                device_directory = os.path.join(self.default_device_log_dir, hostname)
            if os.path.exists(device_directory):
                # 获取最新的Excel文件
                excel_files = [f for f in os.listdir(device_directory) if f.endswith(".xlsx")]
                if excel_files:
                    latest_excel = max(excel_files, key=lambda x: os.path.getmtime(os.path.join(device_directory, x)))
                    df = pd.read_excel(os.path.join(device_directory, latest_excel))
                    commands = df['命令'].tolist()
                    results = df['输出'].tolist()
                    all_results[hostname] = {cmd: res for cmd, res in zip(commands, results)}

        # 如果没有找到任何结果，返回空字符串
        if not all_results:
            return "<h1>无执行结果</h1>"

        # 创建汇总的HTML内容
        html_content = """
        <html>
        <head>
            <title>命令执行结果汇总</title>
            <style>
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
            </style>
        </head>
        <body>
            <h1>命令执行结果汇总</h1>
            <table>
                <tr>
                    <th>设备</th>
                    <th>命令</th>
                    <th>输出</th>
                </tr>
                {}
            </table>
        </body>
        </html>
        """.format(
            "\n".join([
                f"<tr><td>{hostname}</td><td>{cmd}</td><td><pre>{res}</pre></td></tr>"
                for hostname, cmds in all_results.items()
                for cmd, res in cmds.items()
            ])
        )

        return html_content

    def export_command_result_zip(self):
        zip_file_path = filedialog.asksaveasfilename(defaultextension=".zip", filetypes=[("Zip files", "*.zip")])
        if zip_file_path:
            try:
                with zipfile.ZipFile(zip_file_path, 'w') as zipf:
                    # 遍历设备目录，将每个设备的命令执行结果文件添加到ZIP
                    for item_id in self.tree.get_children():
                        values = self.tree.item(item_id, "values")
                        hostname = values[0]
                        # 根据用户是否选择了保存目录，决定日志保存路径
                        if self.selected_save_directory:
                            device_directory = os.path.join(self.selected_save_directory, hostname)
                        else:
                            device_directory = os.path.join(self.default_device_log_dir, hostname)
                        if os.path.exists(device_directory):
                            # 添加TXT和Excel文件
                            for filename in os.listdir(device_directory):
                                file_path = os.path.join(device_directory, filename)
                                if os.path.isfile(file_path) and (
                                        filename.endswith(".txt") or filename.endswith(".xlsx")):
                                    # 确保相对路径正确
                                    arcname = os.path.relpath(file_path,
                                                              self.selected_save_directory or self.default_device_log_dir)
                                    zipf.write(file_path, arcname=arcname)

                            # 添加HTML文件
                            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                            html_content = self.generate_device_html_report(hostname)
                            html_filename = f"{hostname}_{timestamp}_report.html"

                            # 保存为 HTML 文件
                            html_file_path = os.path.join(device_directory, html_filename)
                            with open(html_file_path, "w", encoding="utf-8") as f:
                                f.write(html_content)
                            # 确保相对路径正确
                            arcname = os.path.relpath(html_file_path,
                                                      self.selected_save_directory or self.default_device_log_dir)
                            zipf.write(html_file_path, arcname=arcname)
                            os.remove(html_file_path)  # 删除临时HTML文件

                    # 添加汇总Excel文件
                    self.export_command_results_to_excel(zipf)

                    # 添加汇总HTML文件
                    summary_html_content = self.generate_summary_html_report()
                    summary_html_filename = "command_results_summary.html"
                    summary_html_path = os.path.join(self.default_device_log_dir, summary_html_filename)
                    with open(summary_html_path, "w", encoding="utf-8") as f:
                        f.write(summary_html_content)
                    zipf.write(summary_html_path, arcname=summary_html_filename)
                    os.remove(summary_html_path)  # 删除临时HTML文件

                messagebox.showinfo("成功", "脚本执行结果汇总ZIP 文件已导出")
                self.log(f"[成功]：脚本执行结果汇总ZIP 文件已导出至: {zip_file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"导出脚本执行结果汇总ZIP 文件失败: {str(e)}")
                self.log(f"[错误]：导出脚本执行结果汇总ZIP 文件失败: {str(e)}")


    def generate_device_html_report(self, hostname):
        # 根据用户是否选择了保存目录，决定日志保存路径
        if self.selected_save_directory:
            device_directory = os.path.join(self.selected_save_directory, hostname)
        else:
            device_directory = os.path.join(self.default_device_log_dir, hostname)

        excel_files = [f for f in os.listdir(device_directory) if f.endswith(".xlsx")]
        if not excel_files:
            return "<h1>无执行结果</h1>"

        # 读取最新的Excel文件
        latest_excel = max(excel_files, key=lambda x: os.path.getmtime(os.path.join(device_directory, x)))
        df = pd.read_excel(os.path.join(device_directory, latest_excel))

        # 生成HTML内容
        html_content = """
        <html>
        <head>
            <title>设备执行结果 - {hostname}</title>
            <style>
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
            </style>
        </head>
        <body>
            <h1>设备执行结果 - {hostname}</h1>
            <table>
                <tr>
                    <th>命令</th>
                    <th>输出</th>
                </tr>
                {table_rows}
            </table>
        </body>
        </html>
        """.format(
            hostname=hostname,
            table_rows="\n".join([
                f"<tr><td>{row['命令']}</td><td><pre>{row['输出']}</pre></td></tr>"
                for _, row in df.iterrows()
            ])
        )

        return html_content

    def download_template(self):
        template_data = {
            "hostname": ["192.168.1.1", "192.168.2.2", "192.168.2.3"],
            "username": ["admin", "admin", "root"],
            "password": ["password1", "password2", "password2"],
            "protocol": ["telnet", "ssh", "ssh"],
            "device_type": ["h3c", "huawei", "linux"],
            "path": ["c:/commands1.txt", "c:/commands2.txt", "c:/commands2.txt"],
            "port": ["23", "22", "22"],
            "test_result": ["", "", ""],
            "execute_result": ["", "", ""],
            "super_password": ["", "", ""]
        }

        df = pd.DataFrame(template_data)
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            try:
                df.to_excel(file_path, index=False)
                messagebox.showinfo("成功", "设备模板下载成功")
                self.log(f"[成功]：设备模板已下载至: {file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"下载设备模板失败: {str(e)}")
                self.log(f"[错误]：下载设备模板失败: {str(e)}")

    def create_context_menu(self):
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="编辑", command=self.edit_task_context)
        self.context_menu.add_command(label="删除", command=self.remove_task_context)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="暂停", command=self.pause_task_context)
        self.context_menu.add_command(label="恢复", command=self.resume_task_context)

        self.task_listbox.bind("<Button-3>", self.show_context_menu)

    def show_context_menu(self, event):
        self.task_listbox.selection_set(self.task_listbox.nearest(event.y))
        self.context_menu.post(event.x_root, event.y_root)

    def edit_task_context(self):
        selected_task = self.task_listbox.curselection()
        if not selected_task:
            return
        self.edit_task(None)

    def edit_task_context(self):
        selected_tasks = self.task_listbox.curselection()
        if len(selected_tasks) > 1:
            messagebox.showinfo("提示", "不支持多任务的编辑操作")
            return
        self.edit_task(None)

    def remove_task_context(self):
        self.remove_task()

    def pause_task_context(self):
        self.pause_task()

    def resume_task_context(self):
        self.resume_task()

    def add_task(self):
        task_time = self.task_entry.get().strip()
        selected_operation = self.task_menu.get()

        valid_operations = ["Login 测试", "执行脚本", "Ping 测试"]

        if selected_operation not in valid_operations:
            # messagebox.showerror("错误", "请在下拉框中选择操作类型")
            self.log("[错误]：添加定时任务失败：无效的操作类型，请在下拉框中选择")
            return

        if not task_time or not selected_operation:
            # messagebox.showerror("错误", "请输入任务时间和选择操作类型")
            self.log("[错误]：添加定时任务失败：任务时间或操作类型为空")
            return

        try:
            hour, minute = map(int, task_time.split(":"))
            if hour < 0 or hour > 23 or minute < 0 or minute > 59:
                raise ValueError
        except ValueError:
            # messagebox.showerror("错误", "无效的时间格式，请输入 HH:MM 格式")
            self.log("[错误]：添加定时任务失败：无效的时间格式")
            return

        job_name = f"{selected_operation} @ {task_time}"
        if job_name in self.jobs:
            # messagebox.showerror("错误", f"任务 '{job_name}' 已存在")
            self.log(f"[错误]：添加定时任务失败：任务 '{job_name}' 已存在")
            return

        def task_wrapper():
            try:
                if selected_operation == "Login 测试":
                    self.start_test_connections()
                elif selected_operation == "执行脚本":
                    self.run_commands_scheduled()
                elif selected_operation == "Ping 测试":
                    self.start_ping_test()
                self.update_task_status(job_name)
            except Exception as e:
                self.update_task_status(job_name, success=False)
                # messagebox.showerror("任务执行失败", f"任务 '{job_name}' 执行失败: {str(e)}")

        trigger = CronTrigger(hour=hour, minute=minute)
        job = self.scheduler.add_job(
            task_wrapper,
            trigger=trigger,
            id=job_name,
            name=job_name
        )
        self.jobs[job.id] = job
        self.task_listbox.insert(tk.END, f"{job.name} [状态: 待执行]")
        self.task_statuses[job.id] = "待执行"
        self.log(f"[成功]：添加定时任务：{job_name}")
        self.log_audit("[操作]：添加定时任务", f"添加定时任务成功：{job_name}")

        config_file = "config.json"
        task_data = {
            "name": job_name,
            "time": task_time,
            "operation": selected_operation,
            "status": "待执行"
        }
        if os.path.exists(config_file):
            with open(config_file, "r") as f:
                config = json.load(f)
        else:
            config = {}
        if "scheduled_tasks" not in config:
            config["scheduled_tasks"] = []
        config["scheduled_tasks"].append(task_data)
        with open(config_file, "w") as f:
            json.dump(config, f, indent=4)

        self.save_tasks_to_json()

    def select_all_tasks(self, event):
        self.task_listbox.select_set(0, tk.END)
        return "break"

    def ctrl_click_select(self, event):
        # 获取鼠标点击的位置
        index = self.task_listbox.nearest(event.y)
        # 切换选择状态
        if index in self.task_listbox.curselection():
            self.task_listbox.selection_clear(index)
        else:
            self.task_listbox.selection_set(index)

    def edit_task(self, event):
        selected_task = self.task_listbox.curselection()
        if not selected_task:
            return

        task_name = self.task_listbox.get(selected_task)
        job_name = task_name.split(" [")[0]
        operation, time_str = job_name.split(" @ ")

        new_time = simpledialog.askstring("编辑任务时间", "请输入新的任务时间 (HH:MM):", parent=self.root,
                                          initialvalue=time_str)
        if not new_time:
            return

        # new_operation = simpledialog.askstring("编辑任务操作", "请输入新的任务操作类型:", parent=self.root,
        #                                         initialvalue=operation)
        # if not new_operation:
        #    return

        try:
            hour, minute = map(int, new_time.split(":"))
            if hour < 0 or hour > 23 or minute < 0 or minute > 59:
                raise ValueError
        except ValueError:
            messagebox.showerror("错误", "无效的时间格式，请输入 HH:MM 格式")
            return

        new_job_name = f"{operation} @ {new_time}"
        if new_job_name in self.jobs and new_job_name != job_name:
            messagebox.showerror("错误", f"任务 '{new_job_name}' 已存在")
            return

        # 更新任务
        job = self.jobs[job_name]
        job.remove()
        del self.jobs[job_name]

        def task_wrapper():
            try:
                if new_operation == "Login 测试":
                    self.start_test_connections()
                elif new_operation == "执行脚本":
                    self.run_commands_scheduled()
                elif new_operation == "Ping 测试":
                    self.start_ping_test()
                self.update_task_status(new_job_name)
            except Exception as e:
                self.update_task_status(new_job_name, success=False)
                messagebox.showerror("任务执行失败", f"任务 '{new_job_name}' 执行失败: {str(e)}")

        trigger = CronTrigger(hour=hour, minute=minute)
        new_job = self.scheduler.add_job(
            task_wrapper,
            trigger=trigger,
            id=new_job_name,
            name=new_job_name
        )
        self.jobs[new_job.id] = new_job
        self.task_listbox.delete(selected_task)
        self.task_listbox.insert(selected_task, f"{new_job.name} [状态: 待执行]")
        self.task_statuses[new_job.id] = "待执行"
        self.log(f"[成功]：已修改定时任务：{job.name}->{new_job.name}")
        self.log_audit("[操作]：修改定时任务", f"已修改定时任务：{job.name}->{new_job.name}")
        self.save_tasks_to_json()

    def remove_task(self):
        selected_tasks = self.task_listbox.curselection()
        if not selected_tasks:
            messagebox.showerror("错误", "未选择任务")
            return

        for index in reversed(selected_tasks):
            task_name = self.task_listbox.get(index)
            job_name = task_name.split(" [")[0]
            if job_name in self.jobs:
                job = self.jobs[job_name]
                job.remove()
                del self.jobs[job_name]
                del self.task_statuses[job.id]
            self.log("[成功]："f"删除定时任务：{job_name}")
            self.log_audit("[操作]：删除定时任务", f"删除定时任务：{job_name}")
            self.task_listbox.delete(index)
        # self.log("[警告]："f"删除定时任务：{job_name}")
        # self.log_audit("[操作]：删除定时任务", f"删除定时任务：{selected_tasks}")
        self.log("[警告]："f"删除定时任务")
        self.save_tasks_to_json()



    def pause_task(self):
        selected_tasks = self.task_listbox.curselection()
        if not selected_tasks:
            messagebox.showerror("错误", "未选择任务")
            return

        for index in selected_tasks:
            task_name = self.task_listbox.get(index)
            job_name = task_name.split(" [")[0]
            if job_name in self.jobs:
                job = self.jobs[job_name]
                job.pause()
                self.task_listbox.delete(index)
                self.task_listbox.insert(index, f"{job.name} [状态: 已暂停]")
                self.task_statuses[job.id] = "已暂停"
                self.log(f"[提示]：定时任务已暂停: {job.name}")
                self.log_audit("[操作]：用户暂停任务", f"定时任务已暂停: {job.name}")
        self.save_tasks_to_json()

    # def resume_task(self):
    #     selected_task = self.task_listbox.curselection()
    #     if not selected_task:
    #         messagebox.showerror("错误", "未选择任务")
    #         return
    #
    #     task_name = self.task_listbox.get(selected_task)
    #     for job_id, job in self.jobs.items():
    #         if job.name == task_name.split(" [")[0]:
    #             job.resume()
    #             self.task_listbox.delete(selected_task)
    #             self.task_listbox.insert(selected_task, f"{job.name} [状态: 待执行]")
    #             self.task_statuses[job_id] = "待执行"
    #             self.log(f"[提示]：定时任务已恢复: {job.name}")
    #             self.log_audit("[操作]：恢复任务", f"定时任务已恢复: {job.name}")
    #             self.save_tasks_to_json()
    #             return
    #     messagebox.showerror("错误", "任务未找到")

    def resume_task(self):
        selected_tasks = self.task_listbox.curselection()
        if not selected_tasks:
            messagebox.showerror("错误", "未选择任务")
            return

        for index in selected_tasks:
            task_name = self.task_listbox.get(index)
            job_name = task_name.split(" [")[0]
            if job_name in self.jobs:
                job = self.jobs[job_name]
                job.resume()
                self.task_listbox.delete(index)
                self.task_listbox.insert(index, f"{job.name} [状态: 待执行]")
                self.task_statuses[job.id] = "待执行"
                self.log(f"[提示]：定时任务已恢复: {job.name}")
                self.log_audit("[操作]：恢复任务", f"定时任务已恢复: {job.name}")
                self.save_tasks_to_json()

    def update_task_status(self, job_name, success=True):
        for i in range(self.task_listbox.size()):
            item = self.task_listbox.get(i)
            if job_name in item:
                new_status = "[状态: 已执行]" if success else "[状态: 执行失败]"
                self.task_listbox.delete(i)
                self.task_listbox.insert(i, f"{job_name} {new_status}")
                break

    def save_tasks_to_json(self):
        tasks = []
        for job_id, job in self.jobs.items():
            tasks.append({
                "name": job.name,
                "time": job.name.split(" @ ")[1],
                "operation": job.name.split(" @ ")[0],
                "status": self.task_statuses.get(job_id, "待执行")
            })
        tasks_json_path = os.path.join(self.config_dir, "tasks.json")
        with open(tasks_json_path, "w") as f:
            json.dump(tasks, f, indent=4)
            # self.log(f"[成功]：自动保存定时任务")
            self.log_audit("[系统]：定时任务保存", f"自动保存定时任务")
            # else:
            #     self.log_audit("[系统]：定时任务保存", f"自动保存定时任务：{job.name}")

    def load_saved_tasks(self):
        tasks_json_path = os.path.join(self.config_dir, "tasks.json")
        if os.path.exists(tasks_json_path):
            with open(tasks_json_path, "r") as f:
                tasks = json.load(f)
            for task in tasks:
                job_name = f"{task['operation']} @ {task['time']}"
                try:
                    hour, minute = map(int, task['time'].split(":"))
                    trigger = CronTrigger(hour=hour, minute=minute)

                    def task_wrapper():
                        try:
                            if task['operation'] == "Login 测试":
                                self.start_test_connections()
                            elif task['operation'] == "执行脚本":
                                self.run_commands_scheduled()
                            elif task['operation'] == "Ping 测试":
                                self.start_ping_test()
                            self.update_task_status(job_name)
                        except Exception as e:
                            self.update_task_status(job_name, success=False)
                            messagebox.showerror("任务执行失败", f"任务 '{job_name}' 执行失败: {str(e)}")

                    job = self.scheduler.add_job(
                        task_wrapper,
                        trigger=trigger,
                        id=job_name,
                        name=job_name
                    )
                    self.jobs[job.id] = job
                    self.task_listbox.insert(tk.END, f"{job.name} [状态: {task['status']}]")
                    self.task_statuses[job.id] = task['status']
                except Exception as e:
                    messagebox.showerror("加载任务失败", f"加载任务 '{job_name}' 失败: {str(e)}")

    # 修改 save_keys_to_json 方法，将 Key.json 保存到 config 目录
    def save_keys_to_json(self):
        config = {
            "server_chan_key": self.server_chan_key_entry.get().strip(),
            "bark_device_key": self.notification_key_entry.get().strip(),
            "scheduled_tasks": []
        }
        for job_id in self.jobs:
            job = self.jobs[job_id]
            status = self.task_statuses.get(job.id, "待执行")
            task_data = {
                "name": job.name,
                "time": job.name.split(" @ ")[1],
                "operation": job.name.split(" @ ")[0],
                "status": status
            }
            config["scheduled_tasks"].append(task_data)
        keys_json_path = os.path.join(self.config_dir, "Key.json")

        with open(keys_json_path, "w") as f:
            json.dump(config, f)

    # 添加以下方法来加载保存的密钥
    # 修改 load_saved_keys 方法，从 config 目录加载 Key.json
    def load_saved_keys(self):
        keys_json_path = os.path.join(self.config_dir, "Key.json")
        # self.log_audit("[提示]：加载默认API密钥配置")
        # self.log(f"[提示]：加载默认API密钥配置")

        # 如果文件不存在，创建一个默认文件
        if not os.path.exists(keys_json_path):
            default_keys = {
                "server_chan_key": "请输入API密钥",
                "bark_device_key": "请输入API密钥"
            }
            with open(keys_json_path, "w") as f:
                json.dump(default_keys, f)

        try:
            with open(keys_json_path, "r") as f:
                config = json.load(f)

            # 设置 Server酱密钥
            server_chan_key = config.get("server_chan_key", "请输入API密钥")
            self.log_audit("[系统]：已加载默认API密钥信息")
            self.log(f"[提示]：已加载默认API密钥信息")
            self.server_chan_key_entry.delete(0, tk.END)
            self.server_chan_key_entry.insert(0, server_chan_key)

            # 设置 Bark设备密钥
            bark_device_key = config.get("bark_device_key", "请输入API密钥")

            self.notification_key_entry.delete(0, tk.END)
            self.notification_key_entry.insert(0, bark_device_key)

        except Exception as e:
            self.log(f"[错误]：加载保存的密钥时出错: {str(e)}")

    def update_notification_config(self, *args):
        # 根据通知类型，更新通知密钥的提示信息
        notification_type = self.notification_type.get()
        placeholder = ""
        if notification_type == "Server酱":
            placeholder = "输入Server酱的API密钥"
        elif notification_type == "Bark-Server":
            placeholder = "输入BARK设备密钥"
        elif notification_type == "Contoso":
            placeholder = "输入Contoso通知密钥"

        self.notification_key_entry.delete(0, tk.END)  # 清空输入框
        self.notification_key_entry.insert(0, placeholder)  # 设置占位符
        self.notification_key_entry.config(fg="gray")  # 设置文本颜色为灰色
        self.save_keys_to_json()  # 存储通知密钥

    def send_server_chan_notification(self, title, content, test_type):
        if not self.enable_notification.get():
            return

        notification_type = self.notification_type.get()  # 获取通知类型
        key = self.notification_key_entry.get().strip()  # 获取通知密钥

        if not key:
            self.log("[错误]：通知密钥未填写")
            return

        try:
            response = None  # 初始化 response
            if notification_type == "Server酱":
                url = f"https://sctapi.ftqq.com/{key}.send"
                data = {"title": title, "desp": content, "tags": '网络工具'}
                response = requests.post(url, data=data)
            elif notification_type == "Bark-Server":
                url = "https://us3.rebo.asia/barknetwork/push"
                data = {
                    "body": content,
                    "device_key": key,
                    "title": title,
                    "badge": 1,
                    "sound": "minuet.caf",
                    "icon": "https://day.app/assets/images/avatar.jpg",
                    "group": "test"
                }
                response = requests.post(url, json=data)
            elif notification_type == "Contoso":
                # 假设 Contoso 的通知逻辑
                pass

            if response is not None:
                if response.status_code == 200:
                    self.log(f"[成功]：{notification_type} ：通知发送成功，请到手机查看。")
                else:
                    self.log(f"[失败]：{notification_type} ： 通知发送失败，状态码：{response.status_code}")
            else:
                self.log("[错误]：通知发送失败，无法获取响应(Server酱)")
        except Exception as e:
            self.log(f"[错误]：发送通知失败: {str(e)}")

    def send_bark_notification(self, title, content, test_type):
        if not self.enable_notification.get():
            return

        notification_type = self.notification_type.get()
        key = self.notification_key_entry.get().strip()

        if not key:
            self.log("[错误]：密钥未填写")
            return

        response = None  # 初始化 response
        try:
            if notification_type == "Bark-Server":
                url = "https://us3.rebo.asia/barknetwork/push"
                headers = {
                    'Content-Type': 'application/json; charset=utf-8'
                }
                data = {
                    "body": content,
                    "device_key": key,
                    "title": title,
                    "badge": 1,
                    # "sound": "minuet.caf",
                    "icon": "https://day.app/assets/images/avatar.jpg",
                    "group": "网络工具服务",
                    "action": "none"
                }
                response = requests.post(url, headers=headers, json=data)
        except Exception as e:
            self.log(f"[错误]：发送通知失败：{str(e)}")

        if response is not None:
            if response.status_code == 200:
                self.log(f"[成功]：{notification_type} 通知发送成功，请到手机查看。")
            else:
                self.log(f"[错误]：{notification_type} 通知发送失败，状态码：{response.status_code}")
        else:
            self.log("[错误]：通知发送失败，无法获取响应(Bark-Server)")

    def auto_save_device_info(self):
        selected_group = self.current_group_var.get()
        devices_by_group = {}

        # 尝试加载现有设备信息
        device_json_path = os.path.join(self.config_dir, "device.json")
        
        print(f"当前局点-{selected_group}")
        # 更新设备信息面板的标题
        if hasattr(self, 'device_frame') and self.device_frame:
            self.device_frame.config(text=f"设备&局点信息 - {selected_group}")
            print(f"当前局点-{selected_group}")
        
        if os.path.exists(device_json_path):
            try:
                with open(device_json_path, "r", encoding="utf-8") as f:
                    loaded_data = json.load(f)
                    if isinstance(loaded_data, dict):
                        devices_by_group = loaded_data
                    else:
                        self.log("[警告]：设备文件格式错误，已创建新的设备文件")
                        devices_by_group = {}
            except json.JSONDecodeError:
                self.log("[警告]：设备文件损坏，已创建新的设备文件")
                devices_by_group = {}
            except Exception as e:
                self.log(f"[错误]：加载设备文件时出错: {str(e)}")
                devices_by_group = {}

        # 确保当前分组存在
        if selected_group not in devices_by_group:
            devices_by_group[selected_group] = []

        # 清空当前分组的设备信息
        devices_by_group[selected_group] = []

        # 遍历 Treeview 获取设备信息并保存到当前分组
        for item_id in self.tree.get_children():
            values = list(self.tree.item(item_id, "values"))
            hostname = values[0]
            username = values[1]
            password = self.original_passwords.get(hostname, "")
            super_password = self.super_passwords.get(hostname, "")
            connection_protocol = values[3]
            device_type = values[4]
            path = values[5]
            port = values[6]
            test_result = values[7] if len(values) > 7 else ""
            execute_result = values[8] if len(values) > 8 else ""

            # 构建设备信息列表
            device = [
                hostname,
                username,
                password,
                connection_protocol,
                device_type,
                path,
                port,
                test_result,
                execute_result,
                super_password,
            ]

            # 添加到当前分组的设备列表
            devices_by_group[selected_group].append(device)

        # 保存到 device.json 文件
        try:
            with open(device_json_path, "w", encoding="utf-8") as f:
                json.dump(devices_by_group, f, ensure_ascii=False, indent=4)
            self.log(f"[成功]：设备信息已自动保存到局点 {selected_group}")
        except Exception as e:
            self.log(f"[错误]：保存设备信息失败: {str(e)}")


    # 从 config 目录加载设备信息
    def load_device_info_from_json(self):
        # 从 config 目录加载设备信息
        try:
            selected_group = self.current_group_var.get()
            device_json_path = os.path.join(self.config_dir, "device.json")
            devices_by_group = {}

            if os.path.exists(device_json_path):
                try:
                    with open(device_json_path, "r", encoding="utf-8") as f:
                        loaded_data = json.load(f)
                        if isinstance(loaded_data, dict):
                            devices_by_group = loaded_data
                        else:
                            self.log("[警告]：设备文件格式错误，已使用默认局点")
                except json.JSONDecodeError:
                    self.log("[警告]：设备文件损坏，已使用默认局点")
                    self.load_default_device_info()
                except Exception as e:
                    self.log(f"[错误]：加载设备文件时出错: {str(e)}")
                    self.load_default_device_info()
            else:
                try:
                    self.log("[提示]：设备文件不存在，已创建默认局点，尝试加载设备模板")
                    # self.load_default_device_info()
                    devices_by_group["默认局点"] = []
                except Exception as e:
                    self.log(f"[错误]：创建局点失败，加载设备信息失败，请手工创建局点: {str(e)}")
                    messagebox.showerror("错误", f"加载设备信息失败: {str(e)}")

            # 确保至少有一个默认局点
            if "默认局点" not in devices_by_group:
                devices_by_group["默认局点"] = []

            self.tree.delete(*self.tree.get_children())
            self.original_passwords = {}
            self.super_passwords = {}

            if selected_group in devices_by_group:
                for device in devices_by_group[selected_group]:
                    if len(device) < 10:
                        continue  # 跳过无效的设备记录
                    hostname = device[0]
                    username = device[1]
                    password = device[2]
                    connection_protocol = device[3]
                    device_type = device[4]
                    path = device[5]
                    port = device[6]
                    test_result = device[7] if len(device) > 7 else ""
                    execute_result = device[8] if len(device) > 8 else ""
                    super_password = device[9] if len(device) > 9 else ""
                    # # 存储原始密码
                    # self.original_passwords[hostname] = password
                    # # print(f"[调试]: 加载device.json后打印原始密码{self.original_passwords}")
                    # self.super_passwords[hostname] = super_password  # 存储 super 密码

                    # 校验IP地址格式
                    if not self.validate_ip(hostname):
                        self.log(f"[错误]: 设备 {hostname} IP无效")
                        continue

                    # 校验端口格式
                    try:
                        port = int(float(port))
                        if port < 1 or port > 65535:
                            raise ValueError
                    except ValueError:
                        self.log(f"[错误]: 设备 {hostname} 端口无效")
                        continue

                    # 插入到 Treeview
                    masked_password = '*' * len(password) if password else ""
                    masked_super_password = '*' * len(super_password) if super_password else ""
                    self.tree.insert("", "end", values=(
                        hostname,
                        username,
                        masked_password,
                        connection_protocol,
                        device_type,
                        path,
                        port,
                        test_result,
                        execute_result,
                        masked_super_password
                    ))

                    # 更新原始密码和Super密码字典
                    if password:
                        self.original_passwords[hostname] = password
                    if super_password:
                        self.super_passwords[hostname] = super_password

                dev_count = len(self.tree.get_children())
                self.log(f"[提示]：设备信息加载成功, 共计{dev_count}条信息")
            else:
                self.log(f"[提示]: 当前局点 {selected_group} 无设备，已加载缺省模板信息")
                self.load_default_device_info()

        except Exception as e:
            self.log(f"[错误]：加载设备信息失败: {str(e)}，已加载缺省模板信息")
            messagebox.showerror("错误", f"加载设备信息失败: {str(e)}")
            self.load_default_device_info()

    def create_group(self):
        group_name = simpledialog.askstring("创建新局点", "输入新局点名称:")
        if group_name:
            device_json_path = os.path.join(self.config_dir, "device.json")
            devices_by_group = {}

            if os.path.exists(device_json_path):
                try:
                    with open(device_json_path, "r", encoding="utf-8") as f:
                        loaded_data = json.load(f)
                        if isinstance(loaded_data, dict):
                            devices_by_group = loaded_data
                        else:
                            self.log("[警告]：设备文件格式错误，已创建新的设备文件")
                            devices_by_group = {}
                except json.JSONDecodeError:
                    self.log("[警告]：设备文件损坏，已创建新的设备文件")
                    devices_by_group = {}
                except Exception as e:
                    self.log(f"[错误]：加载设备文件时出错: {str(e)}")
                    devices_by_group = {}

            if group_name in devices_by_group:
                messagebox.showerror("错误", "局点名称已存在")
                return

            # 初始化新局点为空列表
            devices_by_group[group_name] = []

            # 保存到文件
            try:
                with open(device_json_path, "w", encoding="utf-8") as f:
                    json.dump(devices_by_group, f, ensure_ascii=False, indent=4)
                self.update_group_menu()
                self.log(f"[成功]：创建新局点 {group_name}")
                self.current_group_var.set(group_name)
                self.load_device_info_from_json()
            except Exception as e:
                self.log(f"[错误]：创建新局点失败: {str(e)}")

    def delete_group(self):
        selected_group = self.current_group_var.get()
        if selected_group == "默认局点":
            messagebox.showerror("错误", "默认局点不能删除")
            return
        confirm = messagebox.askyesno("确认删除", f"确定要删除局点 {selected_group} 吗?")
        if confirm:
            device_json_path = os.path.join(self.config_dir, "device.json")
            devices_by_group = {}

            if os.path.exists(device_json_path):
                try:
                    with open(device_json_path, "r", encoding="utf-8") as f:
                        loaded_data = json.load(f)
                        if isinstance(loaded_data, dict):
                            devices_by_group = loaded_data
                        else:
                            self.log("[警告]：设备文件格式错误，已创建新的设备文件")
                            devices_by_group = {}
                except json.JSONDecodeError:
                    self.log("[警告]：设备文件损坏，已创建新的设备文件")
                    devices_by_group = {}
                except Exception as e:
                    self.log(f"[错误]：加载设备文件时出错: {str(e)}")
                    devices_by_group = {}

            if selected_group in devices_by_group:
                del devices_by_group[selected_group]

                # 保存到文件
                try:
                    with open(device_json_path, "w", encoding="utf-8") as f:
                        json.dump(devices_by_group, f, ensure_ascii=False, indent=4)
                    self.update_group_menu()
                    self.log(f"[成功]：删除局点 {selected_group}")
                    self.current_group_var.set("默认局点")
                    self.load_device_info_from_json()
                except Exception as e:
                    self.log(f"[错误]：删除局点失败: {str(e)}")
            else:
                self.log(f"[警告]：局点 {selected_group} 不存在")



    def on_group_change(self, event=None):
        # self.auto_save_device_info()

        selected_group = self.current_group_var.get()
        self.log(f"[提示]：切换到局点：{selected_group}")
        print(f"[调试]:当前局点-{selected_group}")
        # 自动保存设备信息

        self.load_device_info_from_json()  # 重新加载对应分组的设备信息
        time.sleep(0.1)
        
        # self.update_device_frame_title(selected_group)

        # 更新设备信息面板的标题
        if hasattr(self, 'device_frame') and self.device_frame:
            print(f"[调试]: 父窗口状态：{self.root.state()}")
            self.device_frame.config(text=f"设备&局点信息 - {selected_group}")


    def switch_group(self, event=None):
        selected_group = self.current_group_var.get()
        self.load_device_info_from_json()

    def update_group_menu(self):
        device_json_path = os.path.join(self.config_dir, "device.json")
        devices_by_group = {}

        if os.path.exists(device_json_path):
            try:
                with open(device_json_path, "r", encoding="utf-8") as f:
                    loaded_data = json.load(f)
                    if isinstance(loaded_data, dict):
                        devices_by_group = loaded_data
                    else:
                        self.log("[警告]：设备文件格式错误，已使用默认局点")
            except json.JSONDecodeError:
                self.log("[警告]：设备文件损坏，已使用默认局点")
            except Exception as e:
                self.log(f"[错误]：加载设备文件时出错: {str(e)}")
        else:
            try:
                # self.log("[提示]：设备文件不存在，已创建默认局点")
                self.load_default_device_info()
                devices_by_group["默认局点"] = []
            except Exception as e:
                self.log(f"[错误]：创建缺省局点失败，加载设备信息失败，请手工创建: {str(e)}")
                # messagebox.showerror("错误", f"加载设备信息失败: {str(e)}")

        # 确保至少有一个默认局点
        if "默认局点" not in devices_by_group:
            devices_by_group["默认局点"] = []

        group_names = list(devices_by_group.keys())
        self.group_menu['values'] = group_names
        self.group_menu.set("默认局点" if group_names else "")

    def cleanup(self):
        # 销毁所有 Tkinter 的 Variable 对象
        for var in [self.enable_notification, self.bark_device_key, self.notification_type]:
            if hasattr(var, 'destroy'):
                var.destroy()

        # # 关闭日志文件
        # if self.log_file:
        #     self.log_file.close()

        # 销毁所有 Tkinter 的 Image 对象
        if hasattr(self, 'gradient_image') and self.gradient_image:
            self.gradient_image.close()

        if hasattr(self, "ping_stop_event"):
            self.ping_stop_event.set()
            # 等待线程完成
            time.sleep(2)

        # 关闭日志文件
        if self.log_file:
            time.sleep(1)
            self.log_file.close()
            time.sleep(1)

        # if self.del_program_log:
        #     # self.log("[提示]：操作日志正在被使用，程序关闭后日志将删除")
        #     time.sleep(1)
        #     self.del_program_log = False
        #     self.log_file.close()
        #     if os.path.exists(self.log_file_path):
        #         os.remove(self.log_file_path)
        #     # self.log("[提示]：操作日志文件已关闭,并删除")

        # # 清理操作日志
        # if self.del_program_log:
        #     try:
        #         time.sleep(1)
        #         # 尝试关闭日志文件
        #         if hasattr(self, 'log_file') and not self.log_file.closed:
        #             # self.log("[提示]：操作日志正在被使用，程序关闭后日志将删除")
        #             self.del_program_log = False
        #             self.log_file.close()
        #             # self.log("[提示]：操作日志文件已关闭")
        #         if os.path.exists(self.log_file_path):
        #             os.remove(self.log_file_path)
        #         #     self.log("[成功]：操作日志已清理")
        #         else:
        #             print("[提示]：操作日志文件不存在")
        #     except Exception as e:
        #         print (f"[错误]：清理操作日志失败: {str(e)}")


        # 清理设备执行结果
        if self.del_device_log:
                    try:
                        device_log_dir = os.path.join(self.default_device_log_dir)
                        self.del_device_log = False
                        time.sleep(1)
                        print(f"[提示]: 设备执行结果目录路径: {device_log_dir}")  # 打印路径
                        if os.path.exists(device_log_dir):
                            for filename in os.listdir(device_log_dir):
                                file_path = os.path.join(device_log_dir, filename)
                                # print(f"删除文件: {file_path}")  # 打印每个文件路径
                                try:
                                    os.remove(file_path)
                                except PermissionError:
                                    print (f"[警告]：文件 {file_path} 正在被使用，无法删除")
                                except Exception as e:
                                    print(f"[错误]：删除文件 {file_path} 时出错: {str(e)}")
                            # self.log("[成功]：设备执行结果目录已清空")
                        else:
                            print("[提示]：设备执行结果目录不存在")
                    except Exception as e:
                            print (f"[错误]：清理设备执行结果失败: {str(e)}")


        # 销毁Ping窗口
        # if hasattr(self, "ping_window") and self.ping_window.winfo_exists():
        # self.ping_window.destroy()

        # # 销毁主窗口
        # if self.root:
        #     self.root.destroy()
        # self.save_ping_results()

        # 销毁主窗口
        self.root.destroy()

    def on_closing(self):
        self.cleanup()
        # 确保在主线程中销毁 Tkinter 对象
        self.root.after(0, self.cleanup)



    def update_ui_from_thread(self, data):
        self.root.after(0, lambda: self.update_ui(data))
        # 确保 UI 更新在主线程中完成

    def update_ui(self, data):
        # 在这里更新 UI
        pass

    def load_default_device_info(self):
        # 获取当前选择的局点
        print("[调试]: 加载设备模板测试1")
        current_group = self.current_group_var.get()
        print("[调试]: 加载设备模板测试2")

        # 如果没有选择局点，则使用默认局点
        if not current_group:
            current_group = "默认局点"

        # 创建默认设备信息
        default_data = {
            "hostname": "39.156.66.10",
            "username": "admin",
            "password": "password",
            "protocol": "ssh",
            "device_type": "h3c",
            "path": "c:/commands.txt",
            "port": 2222,
            "test_result": "none",
            "execute_result": "none",
            "super_password": "Super密码"
        }

        # 将设备信息插入到设备信息表格中
        self.tree.insert("", "end", values=list(default_data.values()))

        # 更新设备信息面板的标题，显示当前局点
        self.device_frame.config(text=f"设备&局点信息 - {current_group}")

        # 记录日志
        self.log(f"[提示]：填充模板示例设备信息到局点-{current_group} ")

    def update_marquee_title(self):
        # 更新跑马灯效果
        try:
            # 获取当前局点名称
            current_group = self.current_group_var.get()
            if not current_group:
                current_group = "默认局点"  # 如果没有选择局点，则使用默认局点

            # 设置跑马灯标题
            self.marquee_title = f"设备&局点信息 - {current_group}"

            # 获取标题的长度
            title_length = len(self.marquee_title)

            # 计算当前偏移量
            if self.marquee_offset >= title_length:
                self.marquee_offset = 0  # 重置偏移量

            # 更新标题文字
            self.device_frame.config(text=self.marquee_title[self.marquee_offset:])

            # 更新偏移量
            self.marquee_offset += 1

        except Exception as e:
            self.log(f"[错误]：跑马灯效果更新失败: {str(e)}")

        # 重新启动定时器
        self.root.after(200, self.update_marquee_title)

    def start_help_button_flash(self):
        # 启动“使用说明”按钮的闪烁效果
        self.help_button_flash()

    def help_button_flash(self):
        # “使用说明”按钮的闪烁效果
        if hasattr(self, 'help_button') and self.help_button:
            current_bg = self.help_button.cget("bg")
            new_bg = "lightgray" if current_bg == "white" else "white"
            self.help_button.config(bg=new_bg)
            self.help_button_flash_timer = self.root.after(500, self.help_button_flash)


    def show_help(self):
        # 获取当前脚本的目录
        current_dir = os.path.dirname(os.path.abspath(inspect.getfile(inspect.currentframe())))
        # 构建HTML帮助文件路径
        help_file_path = os.path.join(current_dir, "help.html")

        # HTML内容（这里可以替换为上面生成的完整HTML代码）


        # 将HTML内容写入文件
        with open(help_file_path, "w", encoding="utf-8") as f:
            f.write(self.html_content)

        # 打开HTML文件
        webbrowser.open(help_file_path)

    def show_about(self):
        about_window = tk.Toplevel(self.root)
        about_window.title("关于")
        about_window.geometry("300x150")
        about_label = tk.Label(about_window,
                               text=f"网络大剑仙 {self.ver}\n"
                                    "作者：WB\n"
                                    "日期：2025-05-01\n"
                                    "版权所有，禁止商用。",
                               font=("Arial", 12))
        about_label.pack(pady=20)

    def load_command_templates(self):
        config_dir = os.path.join(os.getcwd(), "config")
        os.makedirs(config_dir, exist_ok=True)
        template_path = os.path.join(config_dir, "command_templates.json")
        if os.path.exists(template_path):
            try:
                with open(template_path, "r") as f:
                    self.command_templates = json.load(f)
                self.template_listbox.delete(0, tk.END)
                for template_name in self.command_templates:
                    self.template_listbox.insert(tk.END, template_name)
                    # self.log(f"[提示]：加载命令模板成功")
                self.command_templates.keys()
            except Exception as e:
                self.log(f"[错误]：加载命令模板时出错: {str(e)}")
        else:
            self.command_templates = {}

    def load_selected_template(self):
        selected = self.template_listbox.curselection()
        if not selected:
            messagebox.showerror("错误", "请选择模板")
            self.log(f"[错误]：未选择要加载的模板")
            return

        template_name = self.template_listbox.get(selected[0])
        self.template_name_entry.delete(0, tk.END)
        self.template_name_entry.insert(0, template_name)
        self.template_commands_text.delete("1.0", tk.END)
        self.template_commands_text.insert(tk.END, "\n".join(self.command_templates[template_name]))

    def apply_command_template(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showinfo("提示", "请先选择设备")
            return

        # 弹出窗口，让用户选择命令模板
        template_window = tk.Toplevel(self.root)
        template_window.title("应用命令模板")
        template_window.geometry("300x200")
        self.center_window(template_window)

        # 创建一个下拉框，列出所有可用的命令模板
        template_label = tk.Label(template_window, text="选择命令模板:")
        template_label.pack(pady=10)

        template_var = tk.StringVar()
        template_menu = ttk.Combobox(template_window, textvariable=template_var,
                                     values=list(self.command_templates.keys()))
        template_menu.insert(0, "右侧按钮下拉选择")
        template_menu.pack(pady=10)

        # 确认按钮
        def apply_selected_template():
            selected_template = template_var.get()
            if not selected_template:
                messagebox.showerror("错误", "请选择一个命令模板")
                return

            template_commands = self.command_templates[selected_template]

            # 保存命令脚本到文件
            script_dir = os.path.join(self.default_device_log_dir, "scripts")
            os.makedirs(script_dir, exist_ok=True)
            script_path = os.path.join(script_dir, selected_template + ".txt")

            try:
                with open(script_path, "w") as f:
                    for command in template_commands:
                        f.write(command + "\n")
                self.log(f"[成功]：命令脚本已保存到 {script_path}")
            except Exception as e:
                self.log(f"[错误]：保存命令脚本失败: {str(e)}")
                messagebox.showerror("错误", f"保存命令脚本失败: {str(e)}")
                return

            for item_id in selected_items:
                values = self.tree.item(item_id, "values")
                hostname = values[0]
                username = values[1]
                password = self.original_passwords.get(hostname, "")
                connection_protocol = values[3]
                device_type = values[4]
                path = values[5]
                port = values[6] if len(values) > 6 else None

                # 更新设备的命令脚本路径
                self.tree.set(item_id, "path", script_path)

                # 保存更新后的设备信息
                self.auto_save_device_info()

            messagebox.showinfo("成功", f"模板 '{selected_template}' 已成功应用到选中的设备")
            template_window.destroy()

        apply_button = tk.Button(template_window, text="应用", command=apply_selected_template)
        apply_button.pack(pady=10)

    def create_new_template(self):
        template_name = self.template_name_entry.get().strip()
        if not template_name:
            messagebox.showerror("错误", "请输入模板名称")
            self.log(f"[错误]：保存命令模板错误，请输入模板名称")
            return

        if template_name in self.command_templates:
            messagebox.showerror("错误", "模板名称已存在")
            self.log(f"[错误]：保存命令模板错误，模板名称已存在")
            return

        else:
            self.log(f"[成功]：保存命令模板成功,名称:{template_name}")
            # return

        self.command_templates[template_name] = self.template_commands_text.get("1.0", tk.END).strip().split("\n")
        self.save_command_templates()
        self.template_listbox.insert(tk.END, template_name)
        self.log_audit("[操作]：创建命令模板", f"模板创建成功，名称: {template_name}")

    def save_template(self):
        selected = self.template_listbox.curselection()

        # 1. 取用户输入的模板名称（去除首尾空格）
        user_name = self.template_name_entry.get().strip()

        if user_name:  # 用户填写了名称
            template_name = user_name
        elif selected:  # 用户没填，但列表里已选中
            template_name = self.template_listbox.get(selected[0])
        else:  # 既没填也没选，自动命名
            template_name = f"模板_{time.strftime('%Y%m%d_%H%M%S')}"

        # 2. 保存
        self.command_templates[template_name] = self.template_commands_text.get("1.0", tk.END).strip().split("\n")
        self.save_command_templates()

        # 3. 如果列表里没有该名称，则追加
        if template_name not in self.template_listbox.get(0, tk.END):
            self.template_listbox.insert(tk.END, template_name)

        self.log(f"[提示]：已保存命令模板: {template_name}")
        self.log_audit("[操作]：保存命令模板", f"模板名称: {template_name}")

    # def save_template(self):
    #     selected = self.template_listbox.curselection()
    #     # 如果用户未选择模板，自动以当前时间戳创建新模板
    #     if not selected:
    #         template_name = f"模板_{time.strftime('%Y%m%d_%H%M%S')}"
    #         self.command_templates[template_name] = []
    #         self.template_listbox.insert(tk.END, template_name)  # 可选：立即在列表中显示
    #     else:
    #         template_name = self.template_listbox.get(selected[0])
    #
    #     # 保存命令
    #     self.command_templates[template_name] = self.template_commands_text.get("1.0", tk.END).strip().split("\n")
    #     self.save_command_templates()
    #     self.log(f"[提示]：已保存命令模板: {template_name}")
    #     self.log_audit("[操作]：保存命令模板", f"模板名称: {template_name}")

    def delete_template(self):
        selected = self.template_listbox.curselection()
        if not selected:
            messagebox.showerror("错误", "请选择要删除的模板")
            self.log(f"[错误]：删除命令模板错误，未选择要删除的模板")
            return

        template_name = self.template_listbox.get(selected[0])
        del self.command_templates[template_name]
        self.save_command_templates()
        self.template_listbox.delete(selected[0])
        self.log(f"[警告]：用户已删除命令模板： {template_name}")
        self.log_audit("[操作]：删除命令模板", f"模板名称: {template_name}")

    def export_command_template(self):
        self.log_audit("[操作]：命令模板", "用户点击导出命令模板按钮")
        config_dir = os.path.join(os.getcwd(), "config")
        os.makedirs(config_dir, exist_ok=True)
        default_path = os.path.join(config_dir, "command_templates.json")
        file_path = filedialog.asksaveasfilename(
            initialdir=config_dir,
            initialfile="command_templates.json",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json")]
        )
        if file_path:
            try:
                with open(file_path, 'w') as f:
                    json.dump(self.command_templates, f, indent=4)
                messagebox.showinfo("成功", "命令模板已导出")
                self.log(f"[成功]：命令模板已导出至: {file_path}")
                self.log_audit("[操作]：导出命令模板", f"文件路径: {file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"导出命令模板失败: {str(e)}")
                self.log(f"[错误]：导出命令模板失败: {str(e)}")

    def import_command_template(self):
        config_dir = os.path.join(os.getcwd(), "config")
        os.makedirs(config_dir, exist_ok=True)
        file_path = filedialog.askopenfilename(
            initialdir=config_dir,
            filetypes=[("JSON files", "*.json")]
        )
        if file_path:
            try:
                with open(file_path, 'r') as f:
                    imported_templates = json.load(f)
                self.command_templates.update(imported_templates)
                self.save_command_templates()
                self.template_listbox.delete(0, tk.END)
                for template_name in self.command_templates:
                    self.template_listbox.insert(tk.END, template_name)
                messagebox.showinfo("成功", "命令模板已导入")
                self.log(f"[成功]：命令模板已从: {file_path} 导入")
                self.log_audit("[操作]：导入命令模板", f"文件路径: {file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"导入命令模板失败: {str(e)}")
                self.log(f"[错误]：导入命令模板失败: {str(e)}")

    # 在 open_command_template_manager 方法中添加按钮
    def open_command_template_manager(self):
        template_window = tk.Toplevel(self.root)
        template_window.title("命令模板管理")
        template_window.geometry("550x450")
        self.center_window(template_window)

        # 创建模板列表
        tk.Label(template_window, text="命令模板列表:").grid(row=0, column=0, padx=5, pady=5)
        self.template_listbox = tk.Listbox(template_window, width=20, height=15)
        self.template_listbox.grid(row=1, column=0, padx=5, pady=5)

        # 单击模板加载
        self.template_listbox.bind("<ButtonRelease-1>", lambda event: self.load_selected_template())

        # 加载已保存的模板
        self.load_command_templates()

        # 创建模板编辑区域
        edit_frame = tk.Frame(template_window)
        edit_frame.grid(row=1, column=1, padx=5, pady=5)

        tk.Label(edit_frame, text="模板名称:").grid(row=0, column=0, padx=5, pady=5)
        self.template_name_entry = tk.Entry(edit_frame, width=15)
        self.template_name_entry.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(edit_frame, text="命令内容:").grid(row=1, column=0, padx=5, pady=5)
        self.template_commands_text = ScrolledText(edit_frame, width=35, height=15)
        self.template_commands_text.grid(row=2, column=0, columnspan=2, padx=5, pady=5)

        # 创建操作按钮
        button_frame = tk.Frame(template_window)
        button_frame.grid(row=2, column=0, columnspan=2, padx=5, pady=5)

        tk.Button(button_frame, text="新建模板", command=self.create_new_template).grid(row=0, column=0, padx=5, pady=5)
        tk.Button(button_frame, text="保存模板", command=self.save_template).grid(row=0, column=1, padx=5, pady=5)
        tk.Button(button_frame, text="删除模板", command=self.delete_template).grid(row=0, column=2, padx=5, pady=5)
        tk.Button(button_frame, text="导出模板", command=self.export_command_template).grid(row=0, column=4, padx=5,
                                                                                            pady=5)
        tk.Button(button_frame, text="导入模板", command=self.import_command_template).grid(row=0, column=5, padx=5,
                                                                                            pady=5)
        tk.Button(button_frame, text="关闭", command=template_window.destroy).grid(row=0, column=6, padx=5, pady=5)

    def save_command_templates(self):
        config_dir = os.path.join(os.getcwd(), "config")
        os.makedirs(config_dir, exist_ok=True)
        template_path = os.path.join(config_dir, "command_templates.json")
        # with open(template_path, "w") as f:
        #     json.dump(self.command_templates, f, indent=4)

        try:
            with open(template_path, "w") as f:
                json.dump(self.command_templates, f, indent=4)
            self.log(f"[成功]：命令模板保存成功")
        except Exception as e:
            self.log(f"[错误]：保存命令模板失败: {str(e)}")

    def apply_template_to_selected_devices(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showinfo("提示", "请先选择设备")
            return

        template_name = simpledialog.askstring("应用模板", "请输入模板名称:")
        if not template_name:
            return

        if template_name not in self.command_templates:
            messagebox.showerror("错误", f"模板 '{template_name}' 不存在")
            return

        template_commands = self.command_templates[template_name]

        for item_id in selected_items:
            values = self.tree.item(item_id, "values")
            hostname = values[0]
            username = values[1]
            password = self.original_passwords.get(hostname, "")
            connection_protocol = values[3]
            device_type = values[4]
            path = values[5]
            port = values[6] if len(values) > 6 else None

            self.tree.set(item_id, "path", template_name + ".txt")

            self.auto_save_device_info()

        messagebox.showinfo("成功", f"模板 '{template_name}' 已成功应用到选中的设备")

    def initialize_command_history(self):
        self.command_history = []
        command_history_path = os.path.join(self.config_dir, "command_history.json")
        if os.path.exists(command_history_path):
            with open(command_history_path, "r") as f:
                self.command_history = json.load(f)

    def save_command_history(self, device, commands, results, execution_times):
        entry = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "device": device,
            "commands": commands,
            "results": results,
            "execution_times": execution_times
        }
        self.command_history.append(entry)
        command_history_path = os.path.join(self.config_dir, "command_history.json")
        with open(command_history_path, "w") as f:
            json.dump(self.command_history, f, indent=4)

    def view_command_history(self):
        history_window = tk.Toplevel(self.root)
        history_window.title("命令执行历史记录")
        history_window.geometry("1000x600")
        self.center_window(history_window)

        # 创建一个包含表格和滚动条的框架
        log_frame = tk.Frame(history_window)
        log_frame.pack(fill=tk.BOTH, expand=True)

        # 创建表格显示历史记录
        columns = ("timestamp", "device", "command", "result", "execution_time")
        self.history_tree = ttk.Treeview(log_frame, columns=columns, show='headings')

        # 创建垂直滚动条并绑定到表格
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=scrollbar.set)

        # 使用 grid 布局表格和滚动条
        self.history_tree.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        # 设置框架的列和行权重，使表格可伸缩
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)

        # 加载历史记录
        self.load_command_history_to_tree()

        # 添加按钮框架
        button_frame = tk.Frame(history_window)
        button_frame.pack(fill=tk.X, padx=5, pady=5)

        # 添加刷新和关闭按钮，并居中对齐
        refresh_button = tk.Button(button_frame, text="刷新", command=self.refresh_command_history)
        close_button = tk.Button(button_frame, text="关闭", command=history_window.destroy)

        # 使用 grid 布局将按钮放在中间
        refresh_button.grid(row=0, column=0, padx=10, pady=5)
        close_button.grid(row=0, column=1, padx=10, pady=5)

        # 设置按钮框架的列权重，使按钮居中
        button_frame.columnconfigure(0, weight=1)
        button_frame.columnconfigure(1, weight=1)

    def load_command_history_to_tree(self):
        # 清空表格
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)

        # 加载历史记录
        for entry in self.command_history:
            commands = entry.get("commands", [])
            results = entry.get("results", [])
            execution_times = entry.get("execution_times", [])
            timestamp = entry.get("timestamp", "未知时间")
            device = entry.get("device", "未知设备")

            # 确保 commands、results 和 execution_times 长度一致
            max_length = max(len(commands), len(results), len(execution_times))
            for i in range(max_length):
                command = commands[i] if i < len(commands) else ""
                result = results[i] if i < len(results) else "无结果"
                execution_time = execution_times[i] if i < len(execution_times) else "未知时长"
                self.history_tree.insert("", tk.END, values=(
                    timestamp,
                    device,
                    command,
                    result,
                    execution_time
                ))

    def refresh_command_history(self):
        self.load_command_history_to_tree()


    def initialize_audit_log(self):
        self.audit_log = []
        audit_log_path = os.path.join(self.config_dir, "audit_log.json")
        if os.path.exists(audit_log_path):
            try:
                with open(audit_log_path, "r") as f:
                    content = f.read()
                    try:
                        self.audit_log = json.loads(content)
                        # 验证加载的数据是否为列表
                        if not isinstance(self.audit_log, list):
                            raise ValueError("日志文件内容格式不正确，应为JSON数组")
                        # 验证每个日志条目是否为字典
                        for entry in self.audit_log:
                            if not isinstance(entry, dict):
                                self.audit_log.remove(entry)
                                self.log("[警告]：已跳过一条格式不正确的审计日志记录")
                    except json.JSONDecodeError as e:
                        # 处理JSON解码错误
                        self.log(f"[警告]：审计日志文件包含格式错误，已尝试修复并加载有效部分: {str(e)}")
                        # 尝试逐行读取并解析JSON对象
                        self.audit_log = []
                        lines = content.strip().split('\n')
                        for line in lines:
                            try:
                                log_entry = json.loads(line.strip())
                                if isinstance(log_entry, dict):
                                    self.audit_log.append(log_entry)
                                else:
                                    self.log("[警告]：已跳过一条格式不正确的审计日志记录")
                            except json.JSONDecodeError:
                                self.log("[警告]：已跳过一条无法解析的审计日志记录")
            except Exception as e:
                self.log(f"[错误]：加载审计日志时出错: {str(e)}")
                # 清空审计日志并创建新的空文件
                self.audit_log = []
                try:
                    with open(audit_log_path, "w") as f:
                        json.dump([], f)
                    self.log("[提示]：已清空审计日志并创建新的空文件")
                except Exception as e:
                    self.log(f"[错误]：清空审计日志并创建新文件时出错: {str(e)}")
        else:
            # 如果文件不存在，创建一个空文件
            with open(audit_log_path, "w") as f:
                json.dump([], f)
            self.audit_log = []

    def check_and_create_default_templates(self):
        """检查并创建默认的命令模板文件"""
        template_path = os.path.join(self.config_dir, "command_templates.json")
        if not os.path.exists(template_path):
            # 创建默认模板
            default_templates = {
                "默认模板": ["screen disable\n"
                             "display version"]
            }
            try:
                with open(template_path, "w") as f:
                    json.dump(default_templates, f, indent=4)
                self.log(f"[提示]：已创建默认命令模板文件：{template_path}")
                self.log_audit("[提示]：已创建默认命令模板文件 {template_path}")
            except Exception as e:
                self.log(f"[错误]：创建默认命令模板文件失败: {str(e)}")

    def create_command_template_management(self):
        # 创建命令模板管理菜单
        template_menu = tk.Menu(self.menu_bar, tearoff=0)
        template_menu.add_command(label="打开命令模板管理", command=self.open_command_template_manager)
        self.menu_bar.add_cascade(label="命令模板", menu=template_menu)

    # def show_feedback(self):
    #     feedback_window = tk.Toplevel(self.root)
    #     feedback_window.title("用户反馈")
    #     feedback_window.geometry("400x300")
    #     self.center_window(feedback_window)
    #
    #     # 创建文本框
    #     tk.Label(feedback_window, text="请输入您的反馈：").pack(pady=10)
    #     self.feedback_text = tk.Text(feedback_window, height=10, width=40)
    #     self.feedback_text.pack(pady=10)
    #
    #     # 创建按钮
    #     try:
    #         tk.Button(feedback_window, text="提交反馈", command=self.submit_feedback).pack(pady=10)
    #
    #     except Exception as e:
    #         messagebox.showerror("错误", f"提交失败：{e}")

    def show_feedback(self):
        def _feedback():
            feedback_content = self.feedback_text.get("1.0", tk.END).strip()
            if not feedback_content:
                messagebox.showwarning("提示", "请输入反馈内容")
                return

            local_content = self.get_local_info()
            feedback_content_all = "用户反馈:" + feedback_content + "。\n运行环境:" + local_content
            try:
                url = f"https://sctapi.ftqq.com/sctp5121tx5qrgap3xf2mlrzemn16po.send"
                data = {
                    "title": "用户反馈",
                    "desp": feedback_content_all
                }
                response = requests.post(url, data=data)
                if response.status_code == 200:
                    messagebox.showinfo("成功", "反馈已提交，感谢您的支持")
                    self.log(f"[成功]：反馈已提交，感谢支持。")
                    self.log_audit("[操作]：用户反馈", "反馈已提交，感谢支持")
                    win.destroy()
                else:
                    messagebox.showerror("失败", f"提交反馈失败，请检查网络")
                    self.log(f"[错误]：提交反馈失败，请检查网络")
                    self.log_audit("[操作]：用户反馈", f"提交反馈失败，请检查网络")
                    win.destroy()
            except Exception as e:
                messagebox.showerror("错误", f"提交反馈失败，请检查网络")
                self.log(f"[错误]：提交反馈失败:，请检查网络")
                self.log_audit("[操作]：用户反馈", f"提交反馈失败，请检查网络")
                win.destroy()

        win = tk.Toplevel(self.root)
        win.title("用户反馈")
        win.geometry("400x300")
        self.center_window(win)

        tk.Label(win, text="请填写反馈内容：").pack(pady=5)
        self.feedback_text = tk.Text(win, height=12)
        self.feedback_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        btn_frame = tk.Frame(win)
        btn_frame.pack(pady=5)
        try:
            tk.Button(btn_frame, text="提交", command=_feedback).pack(side=tk.LEFT, padx=10)
            # win.destroy()
        except Exception as e:
            messagebox.showerror("错误", f"提交失败：{e}")
        tk.Button(btn_frame, text="取消", command=win.destroy).pack(side=tk.RIGHT, padx=10)

    # def submit_feedback(self):
    #     feedback_content = self.feedback_text.get("1.0", tk.END).strip()
    #     if not feedback_content:
    #         messagebox.showwarning("提示", "请输入反馈内容")
    #         return
    #
    #     local_content = self.get_local_info()
    #     feedback_content_all = feedback_content + local_content
    #     try:
    #         url = f"https://sctapi.ftqq.com/sctp5121tx5qrgap3xf2mlrzemn16po.send"
    #         data = {
    #             "title": "用户反馈",
    #             "desp": feedback_content_all
    #         }
    #         response = requests.post(url, data=data)
    #         if response.status_code == 200:
    #             messagebox.showinfo("成功", "反馈已提交，感谢您的支持")
    #             self.log(f"[成功]：反馈已提交，感谢支持。")
    #             self.log_audit("[操作]：用户反馈", "反馈已提交，感谢支持")
    #         else:
    #             messagebox.showerror("失败", f"提交反馈失败，请检查网络")
    #             self.log(f"[错误]：提交反馈失败，请检查网络")
    #             self.log_audit("[操作]：用户反馈", f"提交反馈失败，请检查网络")
    #     except Exception as e:
    #         messagebox.showerror("错误", f"提交反馈失败，请检查网络")
    #         self.log(f"[错误]：提交反馈失败:，请检查网络")
    #         self.log_audit("[操作]：用户反馈", f"提交反馈失败，请检查网络")

    def sort_ping_tree_column(self, col):
        # 确保 column_sort_direction 中包含当前列
        if col not in self.column_sort_direction:
            self.column_sort_direction[col] = False  # 默认为升序

        # 获取当前列的数据
        data = [(self.ping_tree.set(item, col), item) for item in self.ping_tree.get_children("")]

        # 检查是否是IP地址列
        if col == "设备":
            # 按照IP地址格式进行排序
            data.sort(key=lambda x: self.parse_ip_for_sorting(x[0]), reverse=self.column_sort_direction[col])
        else:
            # 根据数据类型选择排序方式
            try:
                # 尝试将列数据转换为浮点数进行排序（适用于数值列）
                data.sort(key=lambda x: float(x[0]) if x[0].replace('.', '', 1).isdigit() else x[0],
                          reverse=self.column_sort_direction[col])
            except:
                # 按字符串进行排序（适用于文本列）
                data.sort(key=lambda x: x[0], reverse=self.column_sort_direction[col])

        # 重新排列行
        for index, (_, item) in enumerate(data):
            self.ping_tree.move(item, '', index)

        # 切换排序方向
        self.column_sort_direction[col] = not self.column_sort_direction[col]

        # 设置标题样式以指示排序方向
        dir_str = " ↓" if self.column_sort_direction[col] else " ↑"
        self.ping_tree.heading(col, text=f"{col}{dir_str}")

    def start_specified_ping(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showinfo("提示", "请先选择设备")
            return

        # 弹出输入框，获取用户输入的Ping包个数
        count = simpledialog.askinteger("Ping指定次数", "请输入Ping包的个数:", initialvalue=10)
        if count is None or count <= 0:
            return

        self.log(f"[提示]：用户启动Ping指定次数: {count}")
        self.log_audit("[操作]：Ping指定次数", f"用户启动Ping指定次数: {count}")

        # 启动Ping线程
        threading.Thread(target=self.specified_ping, args=(count,), daemon=True).start()

    def start_continuous_ping(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showinfo("提示", "请先选择设备")
            return

        # 弹出确认对话框
        confirm = messagebox.askyesno("确认", "确定要持续Ping选中的设备吗？", default=messagebox.NO)
        if not confirm:
            return
        confirm = messagebox.askyesno("警告", "此功能仅用于测试，不稳定，请谨慎使用", default=messagebox.NO)
        if not confirm:
            return

        self.log("[提示]：用户启动长Ping")
        self.log_audit("[操作]：持续Ping", "用户启动长Ping")

        # 启动Ping线程
        threading.Thread(target=self.continuous_ping, daemon=True).start()

    def continuous_ping(self):
        selected_items = self.tree.selection()
        if not selected_items:
            return
        # 重置Ping状态
        self.ping_stop_event = threading.Event()  # 重置事件
        self.ping_paused = False

        # 如果Ping窗口已经存在，复用它
        if hasattr(self, "ping_window") and self.ping_window.winfo_exists():
            # 清空现有数据
            for item in self.ping_tree.get_children():
                self.ping_tree.delete(item)
        else:
            # 创建新的Ping窗口
            self.ping_window = tk.Toplevel(self.root)
            self.ping_window.title("Ping结果")
            self.ping_window.geometry("800x600")
            self.center_window(self.ping_window)
            # self.ping_window.protocol("WM_DELETE_WINDOW", lambda: None)  # 禁用窗口关闭功能

            # 创建表格和图表
            self.create_ping_tree_and_chart()

        # 重置Ping状态
        self.ping_results = {}
        for item_id in selected_items:
            values = self.tree.item(item_id, "values")
            hostname = values[0]
            self.ping_results[hostname] = {
                "total": 0,
                "lost": 0,
                "latencies": []
            }
            self.ping_tree.insert("", "end", values=(
                hostname,
                "0",
                "0",
                "0",
                "0%",
                "0 ms"
            ))

        # 启动Ping线程
        threads = []
        for item_id in selected_items:
            values = self.tree.item(item_id, "values")
            hostname = values[0]
            thread = threading.Thread(target=self.ping_device, args=(hostname, True))
            threads.append(thread)
            thread.start()

        # 启动实时更新UI的线程
        threading.Thread(target=self.update_ping_results, daemon=True).start()

        # 更改按钮文本和功能
        if hasattr(self, "pause_button"):
            self.pause_button.config(
                text="暂停Ping",
                command=self.toggle_pause_ping,
                bg="#ffcc00",
                fg="black"
            )

    def update_ping_results(self, count=None):
        while True:
            # 检查是否需要停止
            if self.ping_stop_event.is_set():
                break

            # 检查窗口和 ping_tree 是否仍然存在
            if not hasattr(self, "ping_window") or not self.ping_window.winfo_exists():
                break
            if not hasattr(self, "ping_tree"):
                break

            try:
                # 更新表格
                for item in self.ping_tree.get_children():
                    hostname = self.ping_tree.item(item, "values")[0]
                    if hostname in self.ping_results:
                        result = self.ping_results[hostname]
                        total = result["total"]
                        lost = result["lost"]
                        latencies = result["latencies"]
                        max_latency = result.get("max_latency", 0)  # 获取最高延时
                        min_latency = result.get("min_latency", 0)  # 获取最低延时

                        # # 计算丢包率和平均延时
                        # if total > 0:
                        #     lost_rate = (lost / total) * 100
                        #     avg_latency = sum(latencies) / len(latencies) if latencies else 0
                        # else:
                        #     lost_rate = 0
                        #     avg_latency = 0

                        # 计算丢包率和平均延时
                        if total > 0:
                            lost_rate = (lost / total) * 100
                            if lost_rate == 100:
                                avg_latency = "未响应"
                                max_latency = "未响应"
                                min_latency = "未响应"
                            else:
                                lost_rate = (lost / total) * 100
                                avg_latency = sum(latencies) / len(latencies) if latencies else "0"

                        else:
                            lost_rate = 0
                            avg_latency = 0
                        #
                        # 更新表格
                        self.ping_tree.item(item, values=(
                            hostname,  # 总包数
                            str(total),  # 已发送包的个数
                            str(total),  # 已发送包的个数
                            str(total - lost),  # 已接收包的个数
                            f"{lost_rate:.1f}%",
                            f"{avg_latency:.2f} ms" if isinstance(avg_latency, (int, float)) else avg_latency,
                            f"{max_latency:.2f} ms" if isinstance(max_latency, (int, float)) else max_latency,
                            f"{min_latency:.2f} ms" if isinstance(min_latency, (int, float)) else min_latency
                        ))

                 
                # 更新图表
                if hasattr(self, "figure") and hasattr(self, "subplot"):
                    self.subplot.clear()
                    has_data = False
                    for hostname in self.ping_results:
                        result = self.ping_results[hostname]
                        latencies = result["latencies"]
                        if latencies:
                            x = list(range(1, len(latencies) + 1))
                            line, = self.subplot.plot(x, latencies)
                            line.set_label(hostname)  # 设置线的标签为设备IP
                            has_data = True
                    self.subplot.set_xlabel("Ping次数")
                    self.subplot.set_ylabel("延时 (ms)")
                    self.subplot.set_title("Ping延时图表")
                    self.figure.tight_layout()  # 自动调整布局，消除灰色横条
                    self.canvas.draw()

                # 如果是指定次数Ping，达到次数后停止
                if count is not None:
                    count -= 1
                    if count <= 0:
                        break

            except Exception as e:
                self.log(f"[错误]：更新Ping结果时出错: {str(e)}")
                break

            # 间隔时间
            time.sleep(1)
            # 检查是否需要关闭Ping窗口
        if hasattr(self, "ping_window_should_close") and self.ping_window_should_close:
            if hasattr(self, "ping_window") and self.ping_window.winfo_exists():
                self.root.after(0, lambda: self.ping_window.destroy())
                del self.ping_window_should_close

    def specified_ping(self, count):
        selected_items = self.tree.selection()
        if not selected_items:
            return

        # 如果Ping窗口已经存在，复用它
        if hasattr(self, "ping_window") and self.ping_window.winfo_exists():
            # 清空现有数据
            for item in self.ping_tree.get_children():
                self.ping_tree.delete(item)
        else:
            # 创建新的Ping窗口
            self.ping_window = tk.Toplevel(self.root)
            self.ping_window.title("Ping结果")
            self.ping_window.geometry("800x600")
            self.center_window(self.ping_window)
            # self.ping_window.protocol("WM_DELETE_WINDOW", lambda: None)  # 禁用窗口关闭功能

            # 创建表格和图表
            self.create_ping_tree_and_chart()

        # 重置Ping状态
        self.ping_results = {}
        self.ping_stop_event = threading.Event()
        self.ping_paused = False

        for item_id in selected_items:
            values = self.tree.item(item_id, "values")
            hostname = values[0]
            self.ping_results[hostname] = {
                "total": 0,
                "lost": 0,
                "latencies": []
            }
            self.ping_tree.insert("", "end", values=(
                hostname,
                "0",
                "0",
                "0",
                "0%",
                "0 ms"
            ))

        # 启动Ping线程
        threads = []
        for item_id in selected_items:
            values = self.tree.item(item_id, "values")
            hostname = values[0]
            thread = threading.Thread(target=self.ping_device, args=(hostname, False, count))
            threads.append(thread)
            thread.start()

        # 启动实时更新UI的线程
        threading.Thread(target=self.update_ping_results, args=(count,), daemon=True).start()

    def create_ping_tree_and_chart(self):
        # 创建一个PanedWindow用于布局
        self.paned_window = ttk.PanedWindow(self.ping_window, orient=tk.VERTICAL)
        self.paned_window.pack(fill=tk.BOTH, expand=True)

        # 创建表格显示Ping结果
        frame_tree = ttk.Frame(self.paned_window)
        self.paned_window.add(frame_tree, weight=1)

        # 创建一个包含表格和滚动条的框架
        tree_frame = ttk.Frame(frame_tree)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        columns = ("设备", "总包数", "已发送", "已接收", "丢包率", "平均延时", "最高延时", "最低延时")  # 增加两列
        self.ping_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=10)
        for col in columns:
            self.ping_tree.heading(col, text=col, command=lambda c=col: self.sort_ping_tree_column(c))
            self.ping_tree.column(col, width=100)
        self.ping_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 添加垂直滚动条
        tree_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.ping_tree.yview)
        tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.ping_tree.configure(yscrollcommand=tree_scrollbar.set)

        # 创建图表区域
        frame_chart = ttk.Frame(self.paned_window)
        self.paned_window.add(frame_chart, weight=1)

        self.figure = Figure(figsize=(8, 4), dpi=100)
        self.subplot = self.figure.add_subplot(111)
        self.canvas = FigureCanvasTkAgg(self.figure, master=frame_chart)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        # 创建按钮区域
        button_frame = ttk.Frame(self.ping_window)
        button_frame.pack(fill=tk.X, pady=5)

        # 创建暂停/继续按钮
        self.pause_button = tk.Button(button_frame, text="暂停Ping", command=self.toggle_pause_ping, bg="#ffcc00",
                                      fg="black", padx=10, pady=5)
        self.pause_button.pack(side=tk.LEFT, padx=5)

        # 创建停止按钮
        stop_button = tk.Button(button_frame, text="停止&保存结果", command=self.stop_ping, bg="#ff6b6b", fg="white",
                                padx=10, pady=5)
        stop_button.pack(side=tk.LEFT, padx=5)

    def toggle_pause_ping(self):
        if hasattr(self, "pause_button"):
            if self.ping_paused:
                self.pause_button.config(text="暂停Ping", command=self.toggle_pause_ping, bg="#ffcc00", fg="black")
                self.ping_paused = False
                self.log("[提示]：用户继续Ping")
                self.log_audit("[操作]：继续Ping", "用户继续Ping")
            else:
                self.pause_button.config(text="继续Ping", command=self.toggle_pause_ping, bg="#ffcc00", fg="black")
                self.ping_paused = True
                self.log("[提示]：用户暂停Ping")
                self.log_audit("[操作]：暂停Ping", "用户暂停Ping")

    def ping_device(self, hostname, continuous=False, count=None):
        total = 0
        lost = 0
        latencies = []
        max_latency = 0  # 最高延时
        min_latency = 0  # 最低延时

        while not self.ping_stop_event.is_set():
            # 检查是否暂停Ping
            if hasattr(self, "ping_paused") and self.ping_paused:
                time.sleep(1)
                continue

            try:
                #response_time = ping(hostname, timeout=self.ping_timeout / 1000)
                response_time = ping(hostname, timeout=self.ping_timeout)
                total += 1
                if response_time is None:
                    lost += 1
                    latencies.append(0)
                else:
                    latency = response_time * 1000  # 转换为毫秒
                    latencies.append(latency)

                    # 更新最高和最低延时
                    if latency > max_latency:
                        max_latency = latency
                    if min_latency == 0 or latency < min_latency:
                        min_latency = latency

                # 更新结果
                if hasattr(self, "ping_results"):
                    self.ping_results[hostname] = {
                        "total": total,
                        "lost": lost,
                        "latencies": latencies,
                        "max_latency": max_latency,  # 添加最高延时
                        "min_latency": min_latency  # 添加最低延时
                    }

                # 如果是指定次数Ping，达到次数后停止
                if not continuous and count is not None:
                    count -= 1
                    if count <= 0:
                        break

            except Exception as e:
                self.log(f"[错误]：Ping {hostname} 失败: {str(e)}")
                break

            # 间隔时间
            time.sleep(1)

    def stop_ping(self):
        if hasattr(self, "ping_stop_event"):
            self.ping_stop_event.set()
            self.log("[提示]：用户停止Ping")
            self.log_audit("[操作]：停止Ping", "用户停止Ping")
            self.save_ping_results()  # 调用保存Ping结果的方法

        # 不关闭窗口，而是更改按钮文本和功能
        if hasattr(self, "pause_button"):
            self.pause_button.config(text="开始Ping", command=self.continuous_ping, bg="#4CAF50", fg="white")

        # 设置一个标志位，表示需要关闭Ping窗口
        if hasattr(self, "ping_window"):
            self.ping_window_should_close = True

    def restart_ping(self):
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showinfo("提示", "请先选择设备")
            return

        # 重置Ping状态
        self.ping_stop_event = threading.Event()
        self.ping_paused = False

        # 更改按钮文本和功能
        if hasattr(self, "pause_button"):
            self.pause_button.config(text="暂停Ping", command=self.toggle_pause_ping, bg="#ffcc00", fg="black")

        # 启动Ping线程
        threads = []
        for item_id in selected_items:
            values = self.tree.item(item_id, "values")
            hostname = values[0]
            thread = threading.Thread(target=self.ping_device, args=(hostname, True))
            threads.append(thread)
            thread.start()

        # 启动实时更新UI的线程
        threading.Thread(target=self.update_ping_results, daemon=True).start()

    def save_ping_results(self):
        if not hasattr(self, "ping_results") or not self.ping_results:
            self.log("[提示]：没有Ping结果需要保存")
            return

        try:
            # 确保目录存在
            device_log_dir = os.path.join(os.getcwd(), "device-log")
            os.makedirs(device_log_dir, exist_ok=True)

            # 获取当前时间戳
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

            # 创建Excel文件
            excel_filename = f"Ping_Results_{timestamp}.xlsx"
            excel_path = os.path.join(device_log_dir, excel_filename)

            # 创建PDF文件
            pdf_filename = f"Ping_Results_{timestamp}.pdf"
            pdf_path = os.path.join(device_log_dir, pdf_filename)

            # 准备数据
            data = []
            for hostname, result in self.ping_results.items():
                total = result["total"]
                lost = result["lost"]
                latencies = result["latencies"]
                max_latency = result.get("max_latency", 0)  # 获取最高延时
                min_latency = result.get("min_latency", 0)  # 获取最低延时
                lost_rate = (lost / total) * 100 if total > 0 else 0
                avg_latency = sum(latencies) / len(latencies) if latencies else 0

                data.append({
                    "设备": hostname,
                    "总包数": total,
                    "已发送": total - lost,
                    "已接收": total - lost,
                    "丢包率": f"{lost_rate:.1f}%",
                    "平均延时": f"{avg_latency:.2f} ms",
                    "最高延时": f"{max_latency:.2f} ms",  # 添加最高延时
                    "最低延时": f"{min_latency:.2f} ms"  # 添加最低延时
                })

            # 保存为Excel
            df = pd.DataFrame(data)
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Ping Results', index=False)

                # 创建图表
                plt.figure(figsize=(10, 6))
                for hostname in self.ping_results:
                    result = self.ping_results[hostname]
                    latencies = result["latencies"]
                    if latencies:
                        x = list(range(1, len(latencies) + 1))
                        plt.plot(x, latencies, label=hostname)

                plt.xlabel("Ping次数")
                plt.ylabel("延时 (ms)")
                plt.title("Ping延时图表")
                plt.legend()

                # 将图表保存到Excel
                buf = io.BytesIO()
                plt.savefig(buf, format='png')
                buf.seek(0)
                worksheet = writer.sheets['Ping Results']
                img = Image(buf)
                img.width = 600
                img.height = 300
                worksheet.add_image(img, 'A10')  # 将图表插入到A10单元格
                plt.close()
                # img = Image.open(buf)
                # img.save('temp.png')
                # worksheet.insert_image('A10', 'temp.png')
                # os.remove('temp.png')
                # plt.close()

            self.log(f"[成功]：Ping结果和图表已保存为Excel文件: {excel_path}")

            # 保存为PDF
            with PdfPages(pdf_path) as pdf:
                # 创建图表页面
                plt.figure(figsize=(10, 6))
                for hostname in self.ping_results:
                    result = self.ping_results[hostname]
                    latencies = result["latencies"]
                    if latencies:
                        x = list(range(1, len(latencies) + 1))
                        plt.plot(x, latencies, label=hostname)

                plt.xlabel("Ping次数")
                plt.ylabel("延时 (ms)")
                plt.title("Ping延时图表")
                plt.legend()
                pdf.savefig()  # 保存图表到PDF
                plt.close()

                # 创建数据表格页面
                plt.figure(figsize=(10, 8))
                plt.axis('off')
                plt.table(cellText=df.values, colLabels=df.columns, cellLoc='center', loc='center')
                pdf.savefig()  # 保存表格到PDF
                plt.close()

            self.log(f"[成功]：Ping结果和图表已保存为PDF文件: {pdf_path}")

        except Exception as e:
            self.log(f"[错误]：保存Ping结果失败: {str(e)}")

    def open_compare_window(self):
        if self.compare_window is None or not self.compare_window.winfo_exists():
            self.compare_window = tk.Toplevel(self.root)
            self.compare_window.title("文本对比")
            self.compare_window.geometry("1000x800")
            self.center_window(self.compare_window)
            self.compare_window.protocol("WM_DELETE_WINDOW", self.close_compare_window)

            # 创建对比框架
            compare_frame = tk.Frame(self.compare_window)
            compare_frame.pack(fill=tk.BOTH, expand=True)

            # 创建文件选择区域
            file_frame = tk.Frame(compare_frame)
            file_frame.pack(fill=tk.X, pady=10)

            tk.Label(file_frame, text="文件1:").pack(side=tk.LEFT, padx=5)
            self.file1_label = tk.Label(file_frame, text="未选择")
            self.file1_label.pack(side=tk.LEFT, padx=5)
            tk.Button(file_frame, text="选择文件1", command=self.select_file1_for_compare).pack(side=tk.LEFT, padx=5)

            tk.Label(file_frame, text="文件2:").pack(side=tk.LEFT, padx=5)
            self.file2_label = tk.Label(file_frame, text="未选择")
            self.file2_label.pack(side=tk.LEFT, padx=5)
            tk.Button(file_frame, text="选择文件2", command=self.select_file2_for_compare).pack(side=tk.LEFT, padx=5)

            # 创建对比按钮
            compare_button_frame = tk.Frame(compare_frame)
            compare_button_frame.pack(fill=tk.X, pady=10)
            tk.Button(compare_button_frame, text="开始对比", command=self.start_comparison).pack(side=tk.LEFT, padx=5)

            # 创建对比结果表格
            results_frame = tk.Frame(compare_frame)
            results_frame.pack(fill=tk.BOTH, expand=True)

            columns = ("命令", "文件1输出", "文件2输出", "相似度", "差异")
            self.compare_tree = ttk.Treeview(results_frame, columns=columns, show='headings')
            for col in columns:
                self.compare_tree.heading(col, text=col)
                self.compare_tree.column(col, width=200)
            self.compare_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

            # 添加滚动条
            scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=self.compare_tree.yview)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            self.compare_tree.configure(yscrollcommand=scrollbar.set)

            # 添加进度条和状态标签
            progress_frame = tk.Frame(compare_frame)
            progress_frame.pack(fill=tk.X, pady=5)
            self.compare_progress_bar = ttk.Progressbar(progress_frame, mode="determinate")
            self.compare_progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True)
            self.compare_status_label = tk.Label(progress_frame, text="对比状态：准备开始")
            self.compare_status_label.pack(side=tk.RIGHT, padx=5)

            # 添加操作按钮
            operation_frame = tk.Frame(compare_frame)
            operation_frame.pack(fill=tk.X, pady=5)
            tk.Button(operation_frame, text="显示相同", command=self.show_same_extend).pack(side=tk.LEFT, padx=5)
            tk.Button(operation_frame, text="显示不同", command=self.show_diff_extend).pack(side=tk.LEFT, padx=5)
            tk.Button(operation_frame, text="重置筛选", command=self.reset_filter_extend).pack(side=tk.LEFT, padx=5)
            tk.Button(operation_frame, text="导出结果", command=self.export_comparison_results_extend).pack(
                side=tk.RIGHT, padx=5)
            tk.Button(operation_frame, text="关闭", command=self.close_compare_window).pack(side=tk.RIGHT, padx=5)
        else:
            self.compare_window.lift()

    def close_compare_window(self):
        if self.compare_window:
            self.compare_window.destroy()
            self.compare_window = None

    def select_file1_for_compare(self):
        file_path = filedialog.askopenfilename(filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")])
        if file_path:
            self.compare_file1 = file_path
            self.file1_label.config(text=os.path.basename(file_path))
            self.log(f"[提示]：选择对比文件1：{file_path}")

    def select_file2_for_compare(self):
        file_path = filedialog.askopenfilename(filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")])
        if file_path:
            self.compare_file2 = file_path
            self.file2_label.config(text=os.path.basename(file_path))
            self.log(f"[提示]：选择对比文件2：{file_path}")

    def start_comparison(self):
        if not hasattr(self, 'compare_file1') or not hasattr(self, 'compare_file2'):
            messagebox.showerror("错误", "请选择两个文件进行对比")
            return

        # 清空表格
        for item in self.compare_tree.get_children():
            self.compare_tree.delete(item)

        # 重置进度条和状态标签
        self.compare_progress_bar["value"] = 0
        self.compare_status_label.config(text="对比状态：正在对比...")

        # 启动对比线程
        self.compare_thread_pool.submit(self.compare_files)

    def compare_files(self):
        try:
            # 读取文件内容
            with open(self.compare_file1, 'r', encoding='utf-8') as f:
                content1 = f.read()
            with open(self.compare_file2, 'r', encoding='utf-8') as f:
                content2 = f.read()

            # 按命令分块
            commands1 = self.split_content_by_command_extend(content1)
            commands2 = self.split_content_by_command_extend(content2)

            # 计算总命令数
            total_commands = len(commands1) + len(commands2)
            self.compare_progress_bar["maximum"] = total_commands

            # 对比命令
            results = []
            for cmd1 in commands1:
                best_match = None
                best_similarity = 0
                for cmd2 in commands2:
                    # 计算命令相似度
                    similarity = difflib.SequenceMatcher(None, cmd1["command"], cmd2["command"]).ratio()
                    if similarity > best_similarity:
                        best_similarity = similarity
                        best_match = cmd2

                # 如果找到匹配的命令
                if best_match and best_similarity > 0.8:
                    # 对比输出
                    output_similarity = self.compare_command_outputs(cmd1["output"], best_match["output"])
                    diff_text = self.get_differences(cmd1["output"], best_match["output"])
                    results.append({
                        "command": cmd1["command"],
                        "file1_output": cmd1["output"],
                        "file2_output": best_match["output"],
                        "similarity": f"{output_similarity:.2f}",
                        "diff": diff_text
                    })
                else:
                    # 未找到匹配的命令
                    results.append({
                        "command": cmd1["command"],
                        "file1_output": cmd1["output"],
                        "file2_output": "未找到匹配命令",
                        "similarity": "0.00",
                        "diff": ""
                    })

            # 更新对比结果
            self.root.after(0, self.update_comparison_results, results)
            self.compare_status_label.config(text="对比状态：完成")
            self.log(f"[成功]：文本对比完成，共对比 {len(results)} 条命令，请导出查看对比结果。")

        except Exception as e:
            self.log(f"[错误]：文本对比失败：{str(e)}")
            self.compare_status_label.config(text=f"对比状态：失败 - {str(e)}")

    def compare_command_outputs(self, output1, output2):
        # 逐行对比命令输出
        lines1 = output1.splitlines()
        lines2 = output2.splitlines()
        total_lines = max(len(lines1), len(lines2))
        if total_lines == 0:
            return 1.0

        similar_lines = 0
        for line1, line2 in zip(lines1, lines2):
            similarity = difflib.SequenceMatcher(None, line1, line2).ratio()
            if similarity > 0.95:  # 相似度阈值设为95%
                similar_lines += 1

        # 处理行数不一致的情况
        if len(lines1) > len(lines2):
            similar_lines += len(lines1) - len(lines2)
        elif len(lines2) > len(lines1):
            similar_lines += len(lines2) - len(lines1)

        return similar_lines / total_lines

    def get_differences(self, output1, output2):
        # 获取差异
        diff = difflib.ndiff(output1.splitlines(), output2.splitlines())
        return '\n'.join(diff)

    def update_comparison_results(self, results):
        # 清空表格
        for item in self.compare_tree.get_children():
            self.compare_tree.delete(item)

        # 插入结果
        for result in results:
            item = self.compare_tree.insert("", tk.END, values=(
                result["command"],
                result["file1_output"],
                result["file2_output"],
                result["similarity"],
                result["diff"]
            ))

            # 设置差异文本为黄色背景
            if result["diff"]:
                self.compare_tree.item(item, tags=("diff",))
                self.compare_tree.tag_configure("diff", background="yellow")

    def show_same_extend(self):
        # 显示相同项
        for item in self.compare_tree.get_children():
            similarity = float(self.compare_tree.set(item, "相似度"))
            if similarity < 1.0:
                self.compare_tree.item(item, open=False)
            else:
                self.compare_tree.item(item, open=True)

    def show_diff_extend(self):
        # 显示不同项
        for item in self.compare_tree.get_children():
            similarity = float(self.compare_tree.set(item, "相似度"))
            if similarity >= 1.0:
                self.compare_tree.item(item, open=False)
            else:
                self.compare_tree.item(item, open=True)

    def reset_filter_extend(self):
        # 重置筛选
        for item in self.compare_tree.get_children():
            self.compare_tree.item(item, open=True)

    # def export_comparison_results_extend(self):
        # # 导出对比结果
        # file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        # if file_path:
            # try:
                # # 创建Excel文件
                # wb = Workbook()
                # ws = wb.active
                # ws.append(["命令", "文件1输出", "文件2输出", "相似度", "差异"])

                # # 写入结果
                # for item in self.compare_tree.get_children():
                    # values = self.compare_tree.item(item, "values")
                    # ws.append(values)

                # # 保存文件
                # wb.save(file_path)
                # self.log(f"[成功]：对比结果已导出至：{file_path}")
                # messagebox.showinfo("成功", f"对比结果已导出至：{file_path}")
            # except Exception as e:
                # self.log(f"[错误]：导出对比结果失败：{str(e)}")
                # messagebox.showerror("错误", f"导出对比结果失败：{str(e)}")

    def export_comparison_results_extend(self):
        # 导出对比结果
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("HTML files", "*.html")]
        )
        if file_path:
            try:
                # 根据文件扩展名判断导出格式
                if file_path.endswith('.xlsx'):
                    # 创建Excel文件
                    wb = Workbook()
                    ws = wb.active
                    ws.append(["命令", "文件1输出", "文件2输出", "相似度", "差异"])

                    # 写入结果
                    for item in self.compare_tree.get_children():
                        values = self.compare_tree.item(item, "values")
                        ws.append(values)

                    # 保存文件
                    wb.save(file_path)
                    self.log(f"[成功]：对比结果已导出至：{file_path}")
                    messagebox.showinfo("成功", f"对比结果已导出至：{file_path}")
                elif file_path.endswith('.html'):
                    # 创建HTML内容
                    html_content = """
                    <html>
                    <head>
                        <title>命令执行结果对比</title>
                        <style>
                            table {{ border-collapse: collapse; width: 100%; }}
                            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                            th {{ background-color: #f2f2f2; }}
                            tr:nth-child(even) {{ background-color: #f9f9f9; }}
                            .diff {{ background-color: #ffcccc; }}
                            .same {{ background-color: #e6ffe6; }}
                        </style>
                    </head>
                    <body>
                        <h1>命令执行结果对比</h1>
                        <table>
                            <tr>
                                <th>命令</th>
                                <th>文件1输出</th>
                                <th>文件2输出</th>
                                <th>相似度</th>
                                <th>差异</th>
                            </tr>
                            {}
                        </table>
                    </body>
                    </html>
                    """.format("\n".join([
                        f"<tr class='{'diff' if 'diff' in item else 'same'}'>"
                        f"<td>{values[0]}</td>"
                        f"<td>{values[1]}</td>"
                        f"<td>{values[2]}</td>"
                        f"<td>{values[3]}</td>"
                        f"<td>{values[4]}</td>"
                        "</tr>"
                        for item in self.compare_tree.get_children()
                        for values in [self.compare_tree.item(item, "values")]
                    ]))

                    # 保存HTML文件
                    with open(file_path, "w", encoding="utf-8") as f:
                        f.write(html_content)
                    self.log(f"[成功]：对比结果已导出至：{file_path}")
                    messagebox.showinfo("成功", f"对比结果已导出至：{file_path}")
                else:
                    messagebox.showerror("错误", "不支持的文件格式")
            except Exception as e:
                self.log(f"[错误]：导出对比结果失败：{str(e)}")
                messagebox.showerror("错误", f"导出对比结果失败：{str(e)}")

    def split_content_by_command_extend(self, content):
        # 按命令分块
        commands = []
        lines = content.split('\n')
        current_command = None
        current_output = []

        for line in lines:
            if line.startswith("命令: "):
                if current_command:
                    commands.append({
                        "command": current_command,
                        "output": '\n'.join(current_output)
                    })
                current_command = line.replace("命令: ", "").strip()
                current_output = []
            elif current_command:
                current_output.append(line.strip())

        # 添加最后一个命令
        if current_command:
            commands.append({
                "command": current_command,
                "output": '\n'.join(current_output)
            })

        return commands

    #一键备份与还原功能
    def create_backup(self):
        try:
            # 确保备份目录存在
            backup_dir = os.path.join(os.getcwd(), "backup")
            os.makedirs(backup_dir, exist_ok=True)

            # 获取当前时间戳
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_filename = f"backup_{timestamp}.zip"
            backup_path = os.path.join(backup_dir, backup_filename)

            # 创建ZIP文件
            with zipfile.ZipFile(backup_path, 'w') as zipf:
                # 备份 config 目录
                config_dir = os.path.join(os.getcwd(), "config")
                if os.path.exists(config_dir):
                    for root, dirs, files in os.walk(config_dir):
                        for file in files:
                            file_path = os.path.join(root, file)
                            # 使用相对路径避免重复
                            arcname = os.path.relpath(file_path, config_dir)
                            zipf.write(file_path, os.path.join("config", arcname))

                # 备份 device-log 目录
                log_dir = os.path.join(os.getcwd(), "device-log")
                if os.path.exists(log_dir):
                    for root, dirs, files in os.walk(log_dir):
                        for file in files:
                            file_path = os.path.join(root, file)
                            # 使用相对路径避免重复
                            arcname = os.path.relpath(file_path, log_dir)
                            zipf.write(file_path, os.path.join("device-log", arcname))

            self.log(f"[成功]：备份成功，备份文件已保存到：{backup_path}")
            messagebox.showinfo("备份成功", f"备份文件已保存到：{backup_path}")

        except Exception as e:
            self.log(f"[错误]：备份失败：{str(e)}")
            messagebox.showerror("备份失败", f"备份失败：{str(e)}")

    def restore_backup(self):
        # 弹出文件选择对话框，允许用户选择任意目录下的ZIP文件
        backup_file_path = filedialog.askopenfilename(
            filetypes=[("ZIP files", "*.zip")],
            title="选择备份文件"
        )

        if not backup_file_path:
            self.log("[提示]：用户取消了还原操作")
            return

        try:
            # 检查选中的文件是否是一个有效的ZIP文件
            if not zipfile.is_zipfile(backup_file_path):
                self.log("[错误]：所选文件不是一个有效的ZIP文件")
                messagebox.showerror("还原失败", "所选文件不是一个有效的ZIP文件")
                return

            # 获取备份文件名
            backup_filename = os.path.basename(backup_file_path)

            # 弹出确认对话框
            confirm = messagebox.askyesno("确认还原", f"确定要还原备份文件 {backup_filename} 吗？此操作将覆盖现有文件。")
            if not confirm:
                return

            # 解压备份文件
            with zipfile.ZipFile(backup_file_path, 'r') as zipf:
                zipf.extractall(os.getcwd())

            self.log(f"[成功]：还原成功，已还原备份文件：{backup_filename}")

            # 弹出消息框询问是否自动加载配置文件
            load_config = messagebox.askyesno("加载配置文件", "还原成功！是否自动加载还原后的配置文件？")
            if load_config:
                # 加载设备信息
                self.load_device_info_from_json()
                # 加载审计日志
                self.initialize_audit_log()
                # 加载密钥文件
                self.load_saved_keys()
                # 加载定时任务
                self.load_saved_tasks()
                # 加载提示符模式
                # self.load_prompt_patterns()
                # 加载错误字符检测规则
                # self.load_error_chars()
                # 加载命令历史
                self.initialize_command_history()
                # 加载命令模板
                # self.load_command_templates()
                self.log("[提示]：已自动加载所有还原后的配置文件")
            else:
                self.log("[提示]：用户选择不自动加载配置文件")

            messagebox.showinfo("还原成功", f"已还原备份文件：{backup_filename}\n请重启程序以确保所有配置生效。")

        except Exception as e:
            self.log(f"[错误]：还原失败：{str(e)}")
            messagebox.showerror("还原失败", f"还原失败：{str(e)}")

    #增加右键预览脚本txt文件功能
    def add_preview_script_to_context_menu(self):
        # 添加分隔线和“预览脚本”选项到右键菜单
        self.right_click_menu.add_separator()
        self.right_click_menu.add_command(label="在线预览脚本", command=self.preview_and_edit_script)
        self.right_click_menu.add_command(label="应用命令模板", command=self.apply_command_template)

    def preview_and_edit_script(self):
        # 获取选中的设备
        selected_items = self.tree.selection()
        if not selected_items:
            self.log("[提示]：请先选择设备")
            return
        elif len(selected_items) >1:
            messagebox.showinfo("错误", "每次只能选择一个脚本")
            return

        # 遍历选中的设备并预览对应的脚本文件
        for item_id in selected_items:
            values = self.tree.item(item_id, "values")
            hostname = values[0]
            script_path = values[5]  # 假设脚本路径存储在第6列

            if not script_path or not os.path.exists(script_path):
                self.log(f"[警告]：设备 {hostname} 没有关联的脚本文件或文件不存在")
                continue

            # 打开脚本文件并显示内容
            try:
                with open(script_path, 'r', encoding='utf-8') as file:
                    content = file.read()

                # 创建预览窗口
                preview_window = tk.Toplevel(self.root)
                preview_window.title(f"脚本预览 - {hostname}")
                preview_window.geometry("600x400")
                self.center_window(preview_window)

                # 创建文本编辑框
                text_widget = ScrolledText(preview_window)
                text_widget.pack(fill=tk.BOTH, expand=True)
                text_widget.insert(tk.END, content)

                # 保存原始内容以便检测修改
                original_content = content

                # 定义关闭窗口时的保存逻辑
                def save_and_close():
                    modified_content = text_widget.get("1.0", tk.END)
                    if modified_content != original_content:
                        try:
                            with open(script_path, 'w', encoding='utf-8') as file:
                                file.write(modified_content)
                            self.log(f"[成功]：脚本文件已保存：{script_path}")
                        except Exception as e:
                            self.log(f"[错误]：保存脚本文件失败：{str(e)}")
                    preview_window.destroy()

                # 绑定窗口关闭事件
                preview_window.protocol("WM_DELETE_WINDOW", save_and_close)

            except Exception as e:
                self.log(f"[错误]：预览脚本文件时出错: {str(e)}")

    def safe_get(fn, default=None):
        """捕获异常并返回默认值"""
        try:
            return fn()
        except Exception:
            return default

    def get_local_info(self):
        import socket, platform, psutil, getpass
        from datetime import datetime

        info = {
            "软件版本": f"网络大剑仙{self.ver}",
            "启动时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "本机IP": "获取失败",
            "公网IP": "获取失败",
            "系统": "获取失败",
            "CPU": "获取失败",
            "内存": "获取失败",
            "磁盘": "获取失败",
            "运行用户": "获取失败",
            "MAC地址": "获取失败",
            "主板序列号": "获取失败",
            "IP所在地": "获取失败"
        }

        def safe_get(func, default="获取失败"):
            try:
                return func()
            except:
                return default

        def _get_mac_addresses(self):
            """返回首块非回环网卡的真实 MAC，异常返回 '获取失败'"""
            try:
                for nic, addrs in psutil.net_if_addrs().items():
                    for snic in addrs:
                        if snic.family == psutil.AF_LINK and snic.address:
                            mac = snic.address.replace("-", ":")
                            if mac and not mac.startswith("00:00:00"):
                                return mac
            except:
                pass
            # 兜底 uuid
            try:
                return ":".join(["%02x" % ((uuid.getnode() >> (8 * i)) & 0xFF) for i in range(6)][::-1])
            except:
                return "获取失败"


        # 局域网 IP
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            info["本机IP"] = s.getsockname()[0]
            # info["公网IP"] = requests.get("https://checkip.amazonaws.com", timeout=3).text.strip()
            # info["本机IP"] = local_ip + public_ip
        except:
            info["本机IP"] = "获取失败"
            # info["公网IP"] = "获取失败"
        finally:
            s.close()

            # 公网 IP
        try:
                s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
                s.connect(("8.8.8.8", 80))
                # info["本机IP"] = s.getsockname()[0]
                info["公网IP"] = requests.get("https://checkip.amazonaws.com", timeout=3).text.strip()
                # info["本机IP"] = local_ip + public_ip
                self.public_ip = info["公网IP"]
        except:
                # info["本机IP"] = "获取失败"
                info["公网IP"] = "获取失败"
        finally:
                s.close()

        try:
            # response = requests.get("https://api.ipify.org?format=json", timeout=5)
            # public_ip = response.json().get("ip", "获取失败")
            geo = requests.get(f"https://ipinfo.io/{self.public_ip}/json", timeout=5).json()
            city = geo.get("city", "未知")
            region = geo.get("region", "未知")
            country = geo.get("country", "未知")
            loc = geo.get("loc", "未知")
            pos = geo.get("postal","未知")
            info["IP所在地"] = f"{country}-{region}-{city}-经伟度{loc}-邮政编码{pos}"
        except:
            pass

        try:
            info["系统"] = platform.platform()
        except:
            pass

        try:
            cpu_model = platform.processor() or "未知型号"
            info["CPU"] = f"{psutil.cpu_count(False)} 核 / {psutil.cpu_count()} 线程（{cpu_model}）"
        except:
            pass

        try:
            mem = psutil.virtual_memory()
            info["内存"] = f"{mem.total // 1024 ** 3} GB / {mem.percent:.1f}% 已用"
        except:
            pass

        try:
            disks = [f"{d.mountpoint} {psutil.disk_usage(d.mountpoint).total // 1024 ** 3}GB" for d in
                     psutil.disk_partitions()]
            info["磁盘"] = " | ".join(disks) if disks else "无磁盘信息"
        except:
            pass

        try:
            info["运行用户"] = getpass.getuser()
        except:
            pass

            # MAC 地址（首个非回环）
        try:
            info["MAC地址"] = self._get_mac_addresses()
        except:
            pass

            # 主板序列号
        try:
            info["主板序列号"] = safe_get(lambda: self._get_board_serial())
        except:
            pass

        return "\n".join(f"{k}：{v}" for k, v in info.items())

    #主板序列号
    def _get_board_serial(self):
        try:
            import subprocess
            if os.name == "nt":  # Windows
                import wmi
                return wmi.WMI().Win32_BaseBoard()[0].SerialNumber.strip()
            else:  # Linux / macOS
                return subprocess.check_output(
                    ["dmidecode", "-s", "baseboard-serial-number"],
                    stderr=subprocess.DEVNULL).decode().strip()
        except:
            return "获取失败"

    def _get_mac_addresses(self):
        """返回首块非回环网卡的真实 MAC，异常返回 '获取失败'"""
        try:
            for nic, addrs in psutil.net_if_addrs().items():
                for snic in addrs:
                    if snic.family == psutil.AF_LINK and snic.address:
                        mac = snic.address.replace("-", ":")
                        if mac and not mac.startswith("00:00:00"):
                            return mac
        except:
            pass
        # 兜底 uuid
        try:
            return ":".join(["%02x" % ((uuid.getnode() >> (8 * i)) & 0xFF) for i in range(6)][::-1])
        except:
            return "获取失败"
    def send_system_info(self):
        """把系统信息发 Server 酱"""
        title = "网络大剑仙 · 系统启动通知"
        content = self.get_local_info()
        # self.log(f"[提示]：推送系统信息\n{content}")
        print(f"[提示]：推送程序启动信息\n{content}")
        self.send_server_sauce_notification(title, content, test_type="网络工具")

    def download_user_manual(self):
        """将用户手册HTML保存为本地文件并打开"""
        # file_path = filedialog.asksaveasfilename(
        #     title="保存用户手册",
        #     initialfile="网络大剑仙软件使用手册.html",
        #     filetypes=[("HTML 文件", "*.html")]
        # )

        file_path = filedialog.asksaveasfilename(
            title="保存用户手册",
            defaultextension=".html",
            initialfile="网络大剑仙软件使用手册.html",
            filetypes=[("HTML 文件", "*.html")]
        )
        if file_path:
            try:
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write(self.html_content)
                messagebox.showinfo("成功", f"用户手册已保存：{file_path}")
                webbrowser.open(file_path)
                self.log(f"[成功] 用户手册已导出：{file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"保存失败：{e}")
                self.log(f"[错误] 用户手册导出失败：{e}")

class ToolTip:
    def __init__(self, widget):
        self.widget = widget
        self.tip_window = None
        self.id = None
        self.wait_time = 500  # 悬停500毫秒后显示提示

    def show_tip(self, text):
        self.text = text
        if self.tip_window or not self.text:
            return
        x, y, _, _ = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25
        self.tip_window = tk.Toplevel(self.widget)
        self.tip_window.wm_overrideredirect(True)
        self.tip_window.wm_geometry(f"+{x}+{y}")
        label = tk.Label(
            self.tip_window,
            text=self.text,
            justify=tk.LEFT,
            background="#ffffe0",
            relief=tk.SOLID,
            borderwidth=1,
            wraplength=400
        )
        label.pack(ipadx=1)

    def hide_tip(self):
        tw = self.tip_window
        self.tip_window = None
        if tw:
            tw.destroy()




def main():
    # root = tk.Tk()
    # app = DeviceManagerApp(root)
    # root.protocol("WM_DELETE_WINDOW", app.on_closing)
    # root.mainloop()
    root = tk.Tk()
    app = DeviceManagerApp(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    # atexit.register(app.save_ping_results)
    root.mainloop()


if __name__ == "__main__":
    main()
